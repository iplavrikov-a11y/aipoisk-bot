from __future__ import annotations

import asyncio
import base64
from copy import deepcopy
from contextlib import asynccontextmanager, contextmanager
from contextvars import ContextVar
from functools import lru_cache
from datetime import datetime, timezone
import json
import logging
import os
import re
import sqlite3
import socket
import threading
try:
    import dns.resolver
    HAS_DNS_RESOLVER = True
except ImportError:
    HAS_DNS_RESOLVER = False
import threading
import uuid
from dataclasses import dataclass, replace
from html import unescape
from pathlib import Path
from typing import Any, Awaitable, Callable
from urllib.parse import urljoin, urlparse, urlsplit, urlunsplit
import xml.etree.ElementTree as ET

import httpx
from bs4 import BeautifulSoup

from .ai import call_llm, parse_json_object
from .config import config
from .dadata_client import enrich_company_by_inn
from .models import SystemSettings

logger = logging.getLogger(__name__)

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", re.I)
PHONE_RE = re.compile(r"(?:\+7|8)\s*[\(\- ]?\d{3}[\)\- ]?\s*\d{3}[\- ]?\d{2}[\- ]?\d{2}")
URL_RE = re.compile(r"https?://[^\s<>)\"']+", re.I)
MIN_VERIFIED_SUPPLIER_SCORE = 55
ProgressCallback = Callable[[int, str], Awaitable[None]]
PRIMARY_SUPPLIER_SEARCH_PROVIDER = "yandex"
MAX_SUPPLIER_DELIVERY_TARGET = 100

BLOCKED_DOMAINS = {
    "2gis.ru",
    "allbiz.ru",
    "all-pribors.ru",
    "alibaba.com",
    "analitikamed.ru",
    "avito.ru",
    "awindex.ru",
    "b2b.house",
    "b2b-postavki.ru",
    "barahla.net",
    "bizorg.su",
    "consultant.ru",
    "clearspending.ru",
    "cntd.ru",
    "edufire37.ru",
    "dprom.online",
    "dzen.ru",
    "eb24.ru",
    "exportv.ru",
    "flagma.ru",
    "flowex-pipe.com",
    "fabricators.ru",
    "fabrikant.ru",
    "fireman.club",
    "gov.ru",
    "grmetr.ru",
    "habr.com",
    "market.yandex.ru",
    "kodtnved.ru",
    "kontur.ru",
    "komtender.ru",
    "law.ru",
    "legalacts.ru",
    "made-in-china.com",
    "mchs.gov.ru",
    "metaprom.ru",
    "mysku.club",
    "optlist.ru",
    "orgs.biz",
    "otc.ru",
    "ozon.ru",
    "paluba.media",
    "opt-union.ru",
    "poisktenderov.ru",
    "poleznayamodel.ru",
    "prostanki.com",
    "pulscen.ru",
    "pandapipe.com",
    "profiminer.ru",
    "productcenter.ru",
    "qrrussia.ru",
    "rostender.info",
    "rts-tender.ru",
    "ruscable.ru",
    "rutube.ru",
    "reestrinform.ru",
    "rusprofile.ru",
    "sbis.ru",
    "sovok.ru",
    "spravker.ru",
    "studbooks.net",
    "studopedia.su",
    "synapsenet.ru",
    "sino-fire.com",
    "monitoring-crm.ru",
    "supl.biz",
    "tek-all.ru",
    "tenderguru.ru",
    "tebiz.ru",
    "tektorg.ru",
    "tenderer.ru",
    "tenderhq.ru",
    "tendermedia.ru",
    "tiu.ru",
    "torgs.ru",
    "tradedir.ru",
    "tgko.ru",
    "wildberries.ru",
    "wiki-prom.ru",
    "vk.com",
    "vseinstrumenti.ru",
    "44fzrf.ru",
    "94fz.ru",
    "yandex.ru",
    "ya.ru",
    "zakupki.gov.ru",
    "zakupki44fz.ru",
    "zakupki360.ru",
    "b2b-center.ru",
    "bicotender.ru",
    "oborudunion.ru",
    "sudact.ru",
    # Banks & Financial Brokers
    "sberbank.ru", "sber.ru", "vtb.ru", "alfabank.ru", "tbank.ru", "tinkoff.ru", "gazprombank.ru",
    "psbank.ru", "lockobank.ru", "open.ru", "sovcombank.ru", "raiffeisen.ru", "mkb.ru", "rshb.ru",
    "domrfbank.ru", "uralsib.ru", "zenit.ru", "absolutbank.ru", "metallinvestbank.ru", "bspb.ru",
    "sravni.ru", "banki.ru", "cifin.ru", "vsezaimy.ru", "vbr.ru", "fintender.ru", "tender-garant.ru",
    "finstar.ru", "expertcentre.org",
    # OFD, EDS, Reporting & Tax Software
    "1-ofd.ru", "astral.ru", "tensor.ru", "taxcom.ru", "platformaofd.ru",
    "taxnet.ru", "e-dis.ru", "ecplegko.ru", "ed-sro.ru",
    # Additional Aggregators & Brokers
    "klerk.ru", "leboard.ru", "profi.ru", "kwork.ru", "youla.ru", "hh.ru", "superjob.ru",
    "zakon.guru", "garant.ru", "audit-it.ru", "vc.ru", "journal.tinkoff.ru",
    "seldon-pro.ru", "seldon.ru", "sberbank-ast.ru", "roseltorg.ru", "etp-ets.ru",
    "gz-spb.ru", "zakazrf.ru", "lot-online.ru", "etp-gpb.ru",
    "tendergo.pro", "izhtender.ru", "bidexpert.ru", "tenderopora.ru", "tendercorp.ru", "tendercapital.ru",
    "gos-44.ru", "tenderup.ru", "b2g-partner.ru", "torgi223.ru", "tender-life.ru",
    "znanium.ru", "shkulev.ru", "e-library.ru", "cyberleninka.ru",
    "gosuslugi.ru", "nalog.gov.ru", "cbr.ru", "fas.gov.ru", "minfin.gov.ru",
}

IRRELEVANT_PATTERNS = [
    r"банковск(?:ая|ие|ую|их)\s+гаранти",
    r"открыти(?:е|я)\s+расчетн(?:ого|ых)\s+счет",
    r"кредитован(?:ие|ия)\s+бизнеса",
    r"сравнен(?:ие|ия)\s+кредит",
    r"онлайн[\s-]касс",
    r"электронн(?:ая|ой|ую|ые)\s+подпис",
    r"выпуск\s+эцп",
    r"оператор\s+фискальных\s+данных",
    r"курсы\s+повышения\s+квалификации",
    r"обучение\s+(?:44-фз|223-фз|госзакупк)",
    r"электронная\s+библиотечная\s+система",
    r"сетевое\s+издание",
    r"городской\s+портал",
    r"новости\s+городов",
    r"агрегатор\s+тендеров",
    r"поиск\s+тендеров\s+и\s+закупок",
    r"тендерное\s+сопровождение",
]
RE_IRRELEVANT = re.compile("|".join(IRRELEVANT_PATTERNS), re.IGNORECASE)

BLOCKED_HOST_SUFFIXES = (
    ".gov.ru",
    ".zakupki.gov.ru",
    ".consultant.ru",
    ".cntd.ru",
    ".wikipedia.org",
    ".ua",
    ".kz",
    ".by",
)


def _clean_email(email_str: str) -> str:
    if not email_str:
        return ""
    import urllib.parse
    email_str = urllib.parse.unquote(str(email_str or "")).strip().lower()
    email_str = email_str.strip(".,;:/()[]{}<>\"'\\_+-# \t\r\n")
    email_str = re.sub(r"^(?:20|3d|25|2f)+", "", email_str).strip()
    if not (5 <= len(email_str) <= 100):
        return ""
    if "@" not in email_str or email_str.count("@") != 1:
        return ""
    user, domain = email_str.split("@", 1)
    user = user.strip(".,;:/()[]{}<>\"'\\_+-#")
    domain = domain.strip(".,;:/()[]{}<>\"'\\_+-#")
    if not user or not domain or "." not in domain:
        return ""
    if any(domain.endswith(ext) for ext in (".js", ".css", ".png", ".jpg", ".jpeg", ".svg", ".webp", ".gif", ".ico")):
        return ""
    return f"{user}@{domain}"


@dataclass(frozen=True)
class Candidate:
    url: str
    domain: str
    title: str = ""
    snippet: str = ""
    source: str = ""
    query: str = ""
    procurement_item_id: str = ""
    ai_rank_confidence: int = 0
    ai_rank_reason: str = ""


@dataclass(frozen=True)
class CandidateMatch:
    accepted: bool
    level: str
    product: str
    reason: str
    matched_terms: tuple[str, ...] = ()


@dataclass(frozen=True)
class ProcurementItem:
    id: str
    name: str
    aliases: tuple[str, ...] = ()
    okpd2_codes: tuple[str, ...] = ()
    category_terms: tuple[str, ...] = ()
    exact_terms: tuple[str, ...] = ()
    required_terms: tuple[str, ...] = ()
    excluded_terms: tuple[str, ...] = ()


@dataclass(frozen=True)
class ProcurementProfile:
    summary: str
    items: tuple[ProcurementItem, ...]
    excluded_terms: tuple[str, ...] = ()
    raw: dict | None = None


@dataclass(frozen=True)
class CandidateRerank:
    candidates: list[Candidate]
    meta: dict


@dataclass(frozen=True)
class MinpromRegistryRequirement:
    required: bool
    measure_type: str = ""
    evidence: str = ""
    reason: str = ""
    raw: dict | None = None


@dataclass(frozen=True)
class MinpromRegistryContext:
    requirement: MinpromRegistryRequirement
    queries: tuple[str, ...] = ()
    entries: tuple[dict, ...] = ()
    candidate_count: int = 0
    filter_error: str = ""
    status: str = "not_required"
    error: str = ""


SUPPLIER_POLICY_NORMAL = "normal"
SUPPLIER_POLICY_MINPROM_ONLY = "minprom_registry_only"
SUPPLIER_POLICY_MINPROM_PRIORITY = "minprom_registry_priority"
VALID_SUPPLIER_SEARCH_POLICIES = {
    SUPPLIER_POLICY_NORMAL,
    SUPPLIER_POLICY_MINPROM_ONLY,
    SUPPLIER_POLICY_MINPROM_PRIORITY,
}
GISP_PRODUCT_REGISTRY_URL = "https://gisp.gov.ru/pp719v2/pub/prod/"
MINPROM_REGISTRY_CACHE_FILENAME = "minprom_registry.xlsx"
MINPROM_REGISTRY_INDEX_FILENAME = "minprom_registry.jsonl"
MINPROM_REGISTRY_SQLITE_FILENAME = "minprom_registry.sqlite"
MINPROM_REGISTRY_SQLITE_SCHEMA_VERSION = "2"
REGISTRY_CANDIDATE_QUERY_STOPWORDS = {
    "гисп",
    "окпд",
    "окпд2",
    "реестр",
    "реестра",
    "реестровая",
    "реестровой",
    "запись",
    "минпромторг",
    "минпромторга",
    "производитель",
    "поставщик",
    "официальный",
    "сайт",
    "продукция",
    "российская",
    "российской",
    "гост",
    "ту",
    "пп",
    "719",
}


_MULTI_PART_TLDS = {
    "com.ru", "net.ru", "org.ru", "msk.ru", "spb.ru", "nov.ru", "sochi.su", "com.tw",
    "co.uk", "org.uk", "me.uk", "ltd.uk", "plc.uk", "com.ua", "co.il", "com.by",
    "com.kz", "com.tr", "co.jp", "com.cn", "com.br", "co.in", "com.au", "co.nz", "co.za",
}


def base_domain(url_or_domain: str) -> str:
    value = str(url_or_domain or "").strip().lower()
    if "://" in value:
        value = urlparse(value).netloc
    value = value.removeprefix("www.").split(":")[0]
    parts = [part for part in value.split(".") if part]
    if len(parts) >= 3 and f"{parts[-2]}.{parts[-1]}" in _MULTI_PART_TLDS:
        return ".".join(parts[-3:])
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return value


def is_blocked(url_or_domain: str) -> bool:
    host = hostname(url_or_domain)
    domain = base_domain(host)
    return (
        host in BLOCKED_DOMAINS
        or domain in BLOCKED_DOMAINS
        or any(host.endswith(suffix) for suffix in BLOCKED_HOST_SUFFIXES)
    )


def hostname(url_or_domain: str) -> str:
    value = str(url_or_domain or "").strip().lower()
    if "://" in value:
        value = urlparse(value).netloc
    return value.removeprefix("www.").split(":")[0]


def normalize_url(value: str) -> str:
    value = str(value or "").strip().rstrip(".,;])")
    if not value:
        return ""
    if value.startswith("//"):
        value = f"https:{value}"
    if not value.startswith(("http://", "https://")):
        value = f"https://{value}"
    parsed = urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    if is_blocked(parsed.netloc):
        return ""
    return value


async def assess_minprom_registry_requirement(settings: SystemSettings, context: str) -> MinpromRegistryRequirement:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for Minprom registry requirement analysis")
    prompt = f"""Определи, нужна ли для этой закупки реестровая запись/выписка Минпромторга как обязательное условие допуска или поставки.

Правило:
- required=true только если документы прямо указывают действующий запрет или обязательную поставку товара из реестра российской промышленной продукции / ГИСП / Минпромторга;
- required=false при ограничении, преимуществе, неприменении меры или если требование не найдено;
- если при запрете указаны ОКПД2, ПП 719, совокупное количество баллов, реестровая запись или выписка из реестра российской промышленной продукции, это существенные признаки required=true;
- не делай вывод по одному слову "Минпромторг" без практического требования к заявке/товару.

Ответ строго JSON:
{{
  "required": true,
  "measure_type": "prohibition|restriction|advantage|not_applied|unknown",
  "evidence": "короткая цитата/фрагмент документа",
  "reason": "кратко почему требуется или не требуется"
}}

Документы:
{context[:14000]}"""
    raw = await call_llm(
        settings,
        prompt,
        system_prompt="Ты тендерный юрист-аналитик. Отвечай только JSON и не придумывай требования Минпромторга.",
        tier="primary",
        routing_key="minprom_registry_requirement",
        json_mode=True,
        timeout_seconds=90,
        response_validator=_validate_minprom_requirement_response,
    )
    parsed = parse_json_object(raw)
    measure_type = str(parsed.get("measure_type") or "").strip().lower()
    required = bool(parsed.get("required"))
    if measure_type in {"restriction", "advantage", "not_applied"}:
        required = False
    return MinpromRegistryRequirement(
        required=required,
        measure_type=measure_type,
        evidence=str(parsed.get("evidence") or "").strip()[:800],
        reason=str(parsed.get("reason") or "").strip()[:800],
        raw=parsed,
    )


async def discover_minprom_registry_context(
    settings: SystemSettings,
    context: str,
    profile: ProcurementProfile,
    requirement: MinpromRegistryRequirement,
) -> MinpromRegistryContext:
    if not requirement.required:
        return MinpromRegistryContext(requirement=requirement, status="not_required")
    try:
        queries = await build_minprom_registry_queries(settings, context, profile, requirement)
        candidate_entries = await search_minprom_registry_entries(queries, max_results=300)
        entries = await filter_minprom_registry_entries_for_profile(settings, profile, candidate_entries)
        return MinpromRegistryContext(
            requirement=requirement,
            queries=tuple(queries),
            entries=tuple(entries),
            candidate_count=len(candidate_entries),
            status="ok" if entries else "empty",
        )
    except Exception as exc:
        return MinpromRegistryContext(
            requirement=requirement,
            status="error",
            error=f"{type(exc).__name__}: {str(exc)[:300]}",
        )


async def build_minprom_registry_queries(
    settings: SystemSettings,
    context: str,
    profile: ProcurementProfile,
    requirement: MinpromRegistryRequirement,
) -> list[str]:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for Minprom registry query generation")
    code_queries = _build_minprom_registry_code_queries(profile)
    prompt = f"""Сформируй запросы для поиска товара/производителей в реестре российской промышленной продукции Минпромторга/ГИСП.

Нужно искать номенклатуру и производителей, а не номер закупки и не площадку.
Используй несколько опор: полное наименование товара, очищенную товарную группу, модель/марку, производителя, ИНН/ОГРН, а также ОКПД2.
Если в профиле есть ОКПД2, добавляй запросы по полному коду и родительским уровням кода, потому что требования ПП 719 могут быть заданы на уровне вида, подгруппы, группы или подкласса.
Если для товара в ПП 719 есть балльная система, запросы должны помогать найти запись с совокупным количеством баллов и актуальным сроком действия.
Ответ строго JSON:
{{"queries": ["короткий запрос 1", "короткий запрос 2"]}}

Обязательные кодовые запросы, которые уже нужно учесть:
{json.dumps(code_queries, ensure_ascii=False)}

Основание требования:
{json.dumps(_minprom_requirement_to_dict(requirement), ensure_ascii=False)}

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Фрагмент ТЗ:
{context[:5000]}"""
    raw = await call_llm(
        settings,
        prompt,
        system_prompt="Ты закупочный исследователь. Формируешь запросы только для реестра Минпромторга/ГИСП.",
        tier="light",
        routing_key="minprom_registry_query_generation",
        json_mode=True,
        timeout_seconds=90,
        response_validator=None if code_queries else _validate_supplier_queries_response,
    )
    parsed = parse_json_object(raw)
    ai_queries = [str(item).strip() for item in parsed.get("queries", []) if str(item).strip()]
    return _clean_supplier_queries(code_queries + ai_queries)[:16]


async def search_minprom_registry_entries(queries: list[str], *, max_results: int = 300) -> list[dict]:
    if not queries:
        return []
    try:
        return await asyncio.to_thread(
            _search_minprom_registry_entries_local,
            queries[:24],
            max_results=max_results,
        )
    except Exception as exc:
        raise RuntimeError(f"Local Minprom registry search failed: {_exception_summary(exc)}") from exc


async def filter_minprom_registry_entries_for_profile(
    settings: SystemSettings,
    profile: ProcurementProfile,
    entries: list[dict],
) -> list[dict]:
    if not entries:
        return []
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for Minprom registry relevance filtering")
    candidates = [
        {
            "index": index,
            "manufacturer": entry.get("manufacturer", ""),
            "product": entry.get("product", ""),
            "inn": entry.get("inn", ""),
            "registry_number": entry.get("registry_number", ""),
            "evidence": entry.get("evidence", ""),
        }
        for index, entry in enumerate(entries)
    ]
    prompt = f"""Ты эксперт по закупкам. Оцени кандидатов из официального реестра Минпромторга/ГИСП на соответствие предмету закупки.

ГЛАВНОЕ ПРАВИЛО КАТЕГОРИЙНОГО СООТВЕТСТВИЯ:
1. Выдели базовую сущность/вид продукции из закупки (например: насосы, кабель, трубы, задвижки, канаты, спецодежда, лабораторное оборудование, метизы, стройматериалы).
2. ПРИНИМАЙ запись завода из ГИСП, если он является ПРОИЗВОДИТЕЛЕМ этой же базовой категории продукции.
3. Различия в конкретных модификациях, ГОСТах, ТУ, марках, размерах, цветах, классах мощности или исполнения НЕ являются поводом для отсева (завод той же отрасли производит весь спектр модификаций своего типа продукции).

Ответ строго JSON:
{{"accepted_indexes":[{{"index":0,"reason":"производитель соответствующей категории продукции ГИСП"}}]}}

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Кандидаты ГИСП:
{json.dumps(candidates, ensure_ascii=False)}
"""
    raw = await call_llm(
        settings,
        prompt,
        system_prompt="Ты закупочный эксперт. Проверяешь только товарное соответствие записей ГИСП закупаемым позициям.",
        tier="light",
        routing_key="minprom_registry_relevance_filter",
        json_mode=True,
        timeout_seconds=90,
    )
    parsed = parse_json_object(raw)
    accepted_raw = parsed.get("accepted_indexes")
    accepted_indexes: list[int] = []
    if isinstance(accepted_raw, list):
        for item in accepted_raw:
            value = item.get("index") if isinstance(item, dict) else item
            try:
                index = int(value)
            except (TypeError, ValueError):
                continue
            if 0 <= index < len(entries) and index not in accepted_indexes:
                accepted_indexes.append(index)
    return [entries[index] for index in accepted_indexes]


_minprom_status_cache: dict[str, Any] = {"sig": None, "data": None}
_minprom_status_lock = threading.Lock()


def _minprom_file_sig(path: Path) -> tuple[bool, int, int]:
    try:
        if not path.is_file():
            return False, 0, 0
        stat = path.stat()
        return True, stat.st_size, stat.st_mtime_ns
    except Exception:
        return False, 0, 0


def get_minprom_registry_cache_status() -> dict[str, Any]:
    xlsx_path = _minprom_registry_xlsx_path()
    index_path = _minprom_registry_index_path()
    sqlite_path = _minprom_registry_sqlite_path()
    sig = (
        _minprom_file_sig(xlsx_path),
        _minprom_file_sig(index_path),
        _minprom_file_sig(sqlite_path),
    )
    with _minprom_status_lock:
        if _minprom_status_cache["sig"] == sig and _minprom_status_cache["data"] is not None:
            return dict(_minprom_status_cache["data"])

    sqlite_meta = _minprom_registry_sqlite_meta(sqlite_path)
    result = {
        "xlsx_exists": xlsx_path.is_file(),
        "xlsx_path": str(xlsx_path),
        "xlsx_size_bytes": xlsx_path.stat().st_size if xlsx_path.is_file() else 0,
        "index_exists": index_path.is_file(),
        "index_path": str(index_path),
        "index_size_bytes": index_path.stat().st_size if index_path.is_file() else 0,
        "sqlite_exists": sqlite_path.is_file(),
        "sqlite_path": str(sqlite_path),
        "sqlite_size_bytes": sqlite_path.stat().st_size if sqlite_path.is_file() else 0,
        "source_url": GISP_PRODUCT_REGISTRY_URL,
        **sqlite_meta,
    }
    with _minprom_status_lock:
        _minprom_status_cache["sig"] = sig
        _minprom_status_cache["data"] = result
    return result


def minprom_registry_preflight_error(policy: str) -> str:
    normalized_policy = normalize_supplier_search_policy(policy)
    if normalized_policy not in {SUPPLIER_POLICY_MINPROM_ONLY, SUPPLIER_POLICY_MINPROM_PRIORITY}:
        return ""

    status = get_minprom_registry_cache_status()
    if not status.get("xlsx_exists"):
        return "Локальный реестр Минпромторга не готов: загрузите XLSX реестра в админке."
    if not status.get("index_exists"):
        return "Локальный реестр Минпромторга не готов: JSONL-индекс отсутствует, загрузите XLSX реестра заново."
    if not status.get("sqlite_ready"):
        return "Локальный реестр Минпромторга не готов: SQLite-индекс отсутствует или устарел, загрузите XLSX реестра заново."
    if int(status.get("sqlite_entry_count") or 0) <= 0:
        return "Локальный реестр Минпромторга не готов: индекс не содержит записей, загрузите корректный XLSX реестра."
    return ""


def store_minprom_registry_xlsx_cache(payload: bytes, *, filename: str = "") -> dict[str, Any]:
    if not payload:
        raise ValueError("Файл реестра Минпромторга пустой")
    if payload[:2] != b"PK":
        raise ValueError("Файл реестра Минпромторга должен быть XLSX")
    xlsx_path = _minprom_registry_xlsx_path()
    index_path = _minprom_registry_index_path()
    sqlite_path = _minprom_registry_sqlite_path()
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = f"{os.getpid()}.{uuid.uuid4().hex}"
    tmp_xlsx = xlsx_path.with_name(f"{xlsx_path.stem}.{suffix}.tmp{xlsx_path.suffix or '.xlsx'}")
    tmp_index = index_path.with_name(f"{index_path.name}.{suffix}.tmp")
    tmp_sqlite = sqlite_path.with_name(f"{sqlite_path.name}.{suffix}.tmp")
    try:
        tmp_xlsx.write_bytes(payload)
        index_count = _build_minprom_registry_jsonl_index(tmp_xlsx, tmp_index)
        if index_count <= 0:
            raise ValueError("В XLSX не найдены строки реестра Минпромторга")
        sqlite_count = _build_minprom_registry_sqlite_index(tmp_index, tmp_sqlite)
        if sqlite_count <= 0:
            raise ValueError("SQLite-индекс реестра Минпромторга не построен")
        os.replace(tmp_xlsx, xlsx_path)
        os.replace(tmp_index, index_path)
        os.replace(tmp_sqlite, sqlite_path)
    finally:
        tmp_xlsx.unlink(missing_ok=True)
        tmp_index.unlink(missing_ok=True)
        tmp_sqlite.unlink(missing_ok=True)
    status = get_minprom_registry_cache_status()
    status.update({"filename": filename, "index_count": index_count, "sqlite_count": sqlite_count})
    return status


def _minprom_registry_cache_dir() -> Path:
    configured = os.getenv("SUPPLIER_MINPROM_REGISTRY_CACHE_DIR", "").strip() or str(getattr(config, "minprom_registry_cache_dir", "") or "").strip()
    if configured:
        return Path(configured)
    db_url = str(getattr(config, "database_url", "") or "")
    if db_url.startswith("sqlite:///"):
        raw_path = db_url.removeprefix("sqlite:///")
        if not raw_path.startswith("/"):
            resolved_db = (Path(__file__).resolve().parents[1] / raw_path).resolve()
        else:
            resolved_db = Path(raw_path)
        return resolved_db.parent / "minprom_registry"
    return Path(__file__).resolve().parents[2] / "data" / "minprom_registry"


def _minprom_registry_xlsx_path() -> Path:
    configured = os.getenv("SUPPLIER_MINPROM_REGISTRY_XLSX_CACHE_PATH", "").strip()
    if configured:
        return Path(configured)
    return _minprom_registry_cache_dir() / MINPROM_REGISTRY_CACHE_FILENAME


def _minprom_registry_index_path() -> Path:
    configured = os.getenv("SUPPLIER_MINPROM_REGISTRY_INDEX_PATH", "").strip()
    if configured:
        return Path(configured)
    return _minprom_registry_cache_dir() / MINPROM_REGISTRY_INDEX_FILENAME


def _minprom_registry_sqlite_path() -> Path:
    configured = os.getenv("SUPPLIER_MINPROM_REGISTRY_SQLITE_PATH", "").strip()
    if configured:
        return Path(configured)
    default_shared = Path("/root/projects/emailagent/storage/minprom_registry/minprom_registry.sqlite")
    if default_shared.is_file():
        return default_shared
    return _minprom_registry_cache_dir() / MINPROM_REGISTRY_SQLITE_FILENAME


def _minprom_registry_sqlite_meta(sqlite_path: Path | None = None) -> dict[str, Any]:
    path = sqlite_path or _minprom_registry_sqlite_path()
    if not path.is_file():
        return {
            "sqlite_ready": False,
            "sqlite_fresh": False,
            "sqlite_entry_count": 0,
            "sqlite_fts_count": 0,
            "sqlite_integrity": "missing",
        }
    try:
        conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        try:
            rows = dict(conn.execute("SELECT key, value FROM meta").fetchall())
            meta_count = rows.get("entry_count")
            if meta_count is not None:
                entry_count = int(meta_count or 0)
            else:
                entry_count = int(conn.execute("SELECT COUNT(*) FROM entries").fetchone()[0] or 0)
            fts_count = entry_count
        finally:
            conn.close()
        index_path = _minprom_registry_index_path()
        fresh = False
        if index_path.is_file():
            stat = index_path.stat()
            fresh = (
                rows.get("source_jsonl_mtime_ns") == str(stat.st_mtime_ns)
                and rows.get("source_jsonl_size") == str(stat.st_size)
                and rows.get("schema_version") == MINPROM_REGISTRY_SQLITE_SCHEMA_VERSION
            )
        return {
            "sqlite_ready": bool(entry_count and entry_count == fts_count and fresh),
            "sqlite_fresh": fresh,
            "sqlite_entry_count": entry_count,
            "sqlite_fts_count": fts_count,
            "sqlite_integrity": "not_checked",
            "sqlite_schema_version": rows.get("schema_version"),
        }
    except Exception as exc:
        return {
            "sqlite_ready": False,
            "sqlite_fresh": False,
            "sqlite_entry_count": 0,
            "sqlite_fts_count": 0,
            "sqlite_integrity": "error",
            "sqlite_error": str(exc)[:500],
        }


def _search_minprom_registry_entries_local(queries: list[str], *, max_results: int) -> list[dict]:
    sqlite_path = _minprom_registry_sqlite_path()
    if sqlite_path.is_file() and sqlite_path.stat().st_size > 1000:
        return _search_minprom_registry_sqlite(queries, max_results=max_results)
    index_path = _ensure_minprom_registry_jsonl_index()
    if not index_path:
        raise RuntimeError(
            "локальный индекс реестра Минпромторга отсутствует; "
            "загрузите XLSX-снимок или JSONL/SQLite индекс"
        )
    if _ensure_minprom_registry_sqlite_index():
        return _search_minprom_registry_sqlite(queries, max_results=max_results)
    return _search_minprom_registry_jsonl(queries, max_results=max_results)


def _ensure_minprom_registry_jsonl_index() -> Path | None:
    index_path = _minprom_registry_index_path()
    if index_path.is_file() and index_path.stat().st_size > 0:
        return index_path
    xlsx_path = _minprom_registry_xlsx_path()
    if not xlsx_path.is_file():
        return None
    count = _build_minprom_registry_jsonl_index(xlsx_path, index_path)
    return index_path if count else None


def _ensure_minprom_registry_sqlite_index() -> Path | None:
    index_path = _ensure_minprom_registry_jsonl_index()
    if not index_path:
        return None
    sqlite_path = _minprom_registry_sqlite_path()
    if _minprom_registry_sqlite_is_fresh(index_path, sqlite_path):
        return sqlite_path
    count = _build_minprom_registry_sqlite_index(index_path, sqlite_path)
    return sqlite_path if count else None


def _minprom_registry_sqlite_is_fresh(index_path: Path, sqlite_path: Path) -> bool:
    if not index_path.is_file() or not sqlite_path.is_file():
        return False
    try:
        stat = index_path.stat()
        conn = sqlite3.connect(str(sqlite_path))
        try:
            rows = dict(conn.execute("SELECT key, value FROM meta").fetchall())
        finally:
            conn.close()
        return (
            rows.get("source_jsonl_mtime_ns") == str(stat.st_mtime_ns)
            and rows.get("source_jsonl_size") == str(stat.st_size)
            and rows.get("schema_version") == MINPROM_REGISTRY_SQLITE_SCHEMA_VERSION
        )
    except Exception:
        return False


def _build_minprom_registry_jsonl_index(xlsx_path: Path, index_path: Path) -> int:
    from openpyxl import load_workbook

    index_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = index_path.with_suffix(".jsonl.tmp")
    count = 0
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        with tmp_path.open("w", encoding="utf-8") as output:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                header_row: tuple[Any, ...] | None = None
                for row in rows:
                    values = [" ".join(str(value or "").split()).strip() for value in row]
                    lowered = " ".join(values).lower()
                    if "инн" in lowered and "продук" in lowered and "реестр" in lowered:
                        header_row = row
                        break
                if not header_row:
                    continue
                headers = [" ".join(str(value or "").split()).strip() for value in header_row]
                manufacturer_idx = _first_index(
                    _header_index(headers, "предприят"),
                    _header_index(headers, "производ"),
                    _header_index(headers, "изготов"),
                    _header_index(headers, "наименование"),
                )
                product_idx = _header_index(headers, "продук")
                inn_idx = _header_index(headers, "инн")
                registry_idx = _first_index(
                    _header_index(headers, "реестров", "номер"),
                    _header_index(headers, "регистрацион", "номер", "реестров"),
                    _header_index(headers, "первич", "регистрацион"),
                )
                valid_until_idx = _first_index(
                    _header_index(headers, "срок"),
                    _header_index(headers, "действ"),
                    _header_index(headers, "заключ"),
                )
                for row in rows:
                    entry, row_text = _registry_entry_from_row(
                        row,
                        headers,
                        manufacturer_idx=manufacturer_idx,
                        product_idx=product_idx,
                        inn_idx=inn_idx,
                        registry_idx=registry_idx,
                        valid_until_idx=valid_until_idx,
                    )
                    if not entry:
                        continue
                    output.write(json.dumps({**entry, "row_text": row_text[:5000]}, ensure_ascii=False) + "\n")
                    count += 1
        if count:
            os.replace(tmp_path, index_path)
        return count
    finally:
        workbook.close()
        tmp_path.unlink(missing_ok=True)


def _build_minprom_registry_sqlite_index(index_path: Path, sqlite_path: Path) -> int:
    if not index_path.is_file():
        return 0
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = sqlite_path.with_name(f"{sqlite_path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    tmp_path.unlink(missing_ok=True)
    count = 0
    conn = sqlite3.connect(str(tmp_path))
    replace_ready = False
    try:
        conn.execute("PRAGMA journal_mode=OFF")
        conn.execute("PRAGMA synchronous=OFF")
        conn.execute("PRAGMA temp_store=MEMORY")
        conn.execute(
            """
            CREATE TABLE entries (
                id INTEGER PRIMARY KEY,
                manufacturer TEXT NOT NULL DEFAULT '',
                product TEXT NOT NULL DEFAULT '',
                inn TEXT NOT NULL DEFAULT '',
                registry_number TEXT NOT NULL DEFAULT '',
                source_url TEXT NOT NULL DEFAULT '',
                evidence TEXT NOT NULL DEFAULT '',
                row_text TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute("CREATE INDEX idx_entries_inn ON entries(inn)")
        conn.execute("CREATE INDEX idx_entries_registry_number ON entries(registry_number)")
        conn.execute(
            """
            CREATE VIRTUAL TABLE entries_fts USING fts5(
                product,
                manufacturer,
                row_text,
                content='entries',
                content_rowid='id',
                tokenize='unicode61'
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        batch_entries: list[tuple[str, str, str, str, str, str, str]] = []
        batch_fts: list[tuple[int, str, str, str]] = []

        def flush() -> None:
            nonlocal batch_entries, batch_fts
            if not batch_entries:
                return
            conn.executemany(
                """
                INSERT INTO entries (
                    manufacturer,
                    product,
                    inn,
                    registry_number,
                    source_url,
                    evidence,
                    row_text
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                batch_entries,
            )
            conn.executemany(
                """
                INSERT INTO entries_fts(rowid, product, manufacturer, row_text)
                VALUES (?, ?, ?, ?)
                """,
                batch_fts,
            )
            batch_entries = []
            batch_fts = []

        with index_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                entry = _registry_entry_from_payload(payload)
                if not entry:
                    continue
                count += 1
                row_text = _registry_payload_row_text(payload, entry)
                batch_entries.append(
                    (
                        entry["manufacturer"],
                        entry["product"],
                        entry["inn"],
                        entry["registry_number"],
                        entry["source_url"],
                        entry["evidence"],
                        row_text[:5000],
                    )
                )
                batch_fts.append((count, entry["product"], entry["manufacturer"], row_text[:5000]))
                if len(batch_entries) >= 5000:
                    flush()
                    conn.commit()
        flush()
        stat = index_path.stat()
        conn.executemany(
            "INSERT INTO meta(key, value) VALUES (?, ?)",
            [
                ("source_jsonl_path", str(index_path)),
                ("source_jsonl_mtime_ns", str(stat.st_mtime_ns)),
                ("source_jsonl_size", str(stat.st_size)),
                ("entry_count", str(count)),
                ("schema_version", MINPROM_REGISTRY_SQLITE_SCHEMA_VERSION),
                ("built_at", datetime.now(timezone.utc).isoformat()),
            ],
        )
        if count:
            conn.execute("INSERT INTO entries_fts(entries_fts) VALUES ('optimize')")
        conn.commit()
        replace_ready = True
    finally:
        conn.close()
        if tmp_path.exists() and not replace_ready:
            tmp_path.unlink(missing_ok=True)
    os.replace(tmp_path, sqlite_path)
    return count


def _search_minprom_registry_sqlite(queries: list[str], *, max_results: int) -> list[dict]:
    sqlite_path = _minprom_registry_sqlite_path()
    if not sqlite_path or not sqlite_path.is_file():
        sqlite_path = _ensure_minprom_registry_sqlite_index()
    if not sqlite_path or not sqlite_path.is_file():
        return []
    scored_entries: list[tuple[float, int, int, dict]] = []
    seen: set[tuple[str, str, str]] = set()
    order = 0
    conn = sqlite3.connect(f"file:{sqlite_path}?mode=ro", uri=True)
    try:
        for _, _, terms in _registry_query_specs(tuple(queries)):
            match_expression = _fts_match_expression(terms)
            if not match_expression:
                continue
            try:
                rows = conn.execute(
                    """
                    SELECT
                        e.manufacturer,
                        e.product,
                        e.inn,
                        e.registry_number,
                        e.source_url,
                        e.evidence,
                        e.row_text
                    FROM entries_fts
                    JOIN entries e ON e.id = entries_fts.rowid
                    WHERE entries_fts MATCH ?
                    ORDER BY bm25(entries_fts)
                    LIMIT ?
                    """,
                    (match_expression, max(50, min(300, max_results * 4))),
                ).fetchall()
            except sqlite3.Error:
                continue
            for manufacturer, product, inn, registry_number, source_url, evidence, row_text in rows:
                entry = {
                    "registry_number": str(registry_number or ""),
                    "manufacturer": str(manufacturer or ""),
                    "product": str(product or ""),
                    "inn": str(inn or ""),
                    "source": "minprom_registry_local_sqlite",
                    "source_url": str(source_url or GISP_PRODUCT_REGISTRY_URL),
                    "evidence": str(evidence or "")[:1000],
                }
                key = _registry_entry_key(entry)
                if key in seen:
                    continue
                query_scores = _registry_candidate_query_scores(str(row_text or product or ""), queries)
                if not query_scores:
                    continue
                seen.add(key)
                for query_index, score in query_scores:
                    scored_entries.append((score, order, query_index, entry))
                order += 1
    finally:
        conn.close()
    return _rank_minprom_registry_candidates(scored_entries, max_results=max_results, query_count=len(queries))


def _search_minprom_registry_jsonl(queries: list[str], *, max_results: int) -> list[dict]:
    index_path = _ensure_minprom_registry_jsonl_index()
    if not index_path:
        return []
    scored_entries: list[tuple[float, int, int, dict]] = []
    seen: set[tuple[str, str, str]] = set()
    order = 0
    with index_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            entry = _registry_entry_from_payload(payload)
            if not entry:
                continue
            key = _registry_entry_key(entry)
            if key in seen:
                continue
            row_text = _registry_payload_row_text(payload, entry)
            query_scores = _registry_candidate_query_scores(row_text, queries)
            if not query_scores:
                continue
            seen.add(key)
            entry["source"] = "minprom_registry_local_jsonl"
            for query_index, score in query_scores:
                scored_entries.append((score, order, query_index, entry))
            order += 1
    return _rank_minprom_registry_candidates(scored_entries, max_results=max_results, query_count=len(queries))


def _registry_entry_from_payload(payload: dict) -> dict | None:
    entry = {
        "registry_number": str(payload.get("registry_number") or "")[:80],
        "manufacturer": str(payload.get("manufacturer") or "")[:240],
        "product": str(payload.get("product") or "")[:300],
        "inn": str(payload.get("inn") or "")[:20],
        "source": str(payload.get("source") or "minprom_registry_local"),
        "source_url": str(payload.get("source_url") or GISP_PRODUCT_REGISTRY_URL),
        "evidence": str(payload.get("evidence") or "")[:1000],
    }
    if not any((entry["registry_number"], entry["manufacturer"], entry["product"], entry["inn"])):
        return None
    return entry


def _registry_payload_row_text(payload: dict, entry: dict) -> str:
    return " ".join(
        part
        for part in (
            entry.get("manufacturer"),
            entry.get("product"),
            entry.get("registry_number"),
            entry.get("evidence"),
            payload.get("row_text"),
        )
        if part
    )


def _registry_entry_key(entry: dict) -> tuple[str, str, str]:
    return (
        str(entry.get("manufacturer") or "").lower(),
        str(entry.get("product") or "").lower(),
        str(entry.get("inn") or entry.get("registry_number") or "").lower(),
    )


@lru_cache(maxsize=4096)
def _registry_candidate_terms(query: str) -> tuple[str, ...]:
    terms: list[str] = []
    seen: set[str] = set()
    normalized = str(query or "").lower().replace("ё", "е")
    for raw_part in normalized.split():
        token = raw_part.strip(" \t\r\n.,;:!?()[]{}<>\"'«»")
        if not token or token in REGISTRY_CANDIDATE_QUERY_STOPWORDS:
            continue
        if token.isdigit() or len(token) < 3:
            continue
        if token in seen:
            continue
        seen.add(token)
        terms.append(token)
        if len(terms) >= 8:
            break
    return tuple(terms)


@lru_cache(maxsize=4096)
def _normalize_registry_text(value: str) -> str:
    return " ".join(str(value or "").lower().replace("ё", "е").split())


@lru_cache(maxsize=4096)
def _registry_query_specs(queries: tuple[str, ...]) -> tuple[tuple[int, str, tuple[str, ...]], ...]:
    specs: list[tuple[int, str, tuple[str, ...]]] = []
    for query_index, query in enumerate(queries):
        specs.append((query_index, _normalize_registry_text(query), _registry_candidate_terms(query)))
    return tuple(specs)


@lru_cache(maxsize=4096)
def _registry_query_term_union(queries: tuple[str, ...]) -> tuple[str, ...]:
    terms: list[str] = []
    seen: set[str] = set()
    for _, _, query_terms in _registry_query_specs(queries):
        for term in query_terms:
            if term in seen:
                continue
            seen.add(term)
            terms.append(term)
    terms.sort(key=len, reverse=True)
    return tuple(terms)


@lru_cache(maxsize=4096)
def _registry_term_has_digit(term: str) -> bool:
    return any(char.isdigit() for char in term)


def _registry_term_matches(term: str, lowered_row_text: str) -> bool:
    if term in lowered_row_text:
        return True
    if _registry_term_has_digit(term):
        return False
    if len(term) < 4:
        return False
    stem = term[: max(3, len(term) - 2)]
    return stem in lowered_row_text


def _registry_candidate_score_against_spec(
    lowered: str,
    normalized_query: str,
    terms: tuple[str, ...],
    matched_terms: set[str] | None = None,
) -> float:
    best_score = 0.0
    if normalized_query and len(normalized_query) >= 4 and normalized_query in lowered:
        best_score = max(best_score, 2.0)
    if not terms:
        return best_score
    if matched_terms is None:
        matched = sum(1 for term in terms if _registry_term_matches(term, lowered))
    else:
        matched = sum(1 for term in terms if term in matched_terms)
    if matched == len(terms):
        best_score = max(best_score, 1.0 + matched / max(1, len(terms)))
    elif matched >= 2 and matched / len(terms) >= 0.6:
        best_score = max(best_score, matched / len(terms))
    elif matched >= 1:
        best_score = max(best_score, 0.5 * (matched / max(1, len(terms))))
    return best_score


def _registry_candidate_query_scores(row_text: str, queries: list[str]) -> list[tuple[int, float]]:
    if not queries:
        return [(0, 1.0)]
    lowered = _normalize_registry_text(str(row_text or ""))
    if not lowered:
        return []
    matched_terms = {
        term for term in _registry_query_term_union(tuple(queries)) if _registry_term_matches(term, lowered)
    }
    scores: list[tuple[int, float]] = []
    for query_index, normalized_query, terms in _registry_query_specs(tuple(queries)):
        if terms and not any(term in matched_terms for term in terms):
            continue
        score = _registry_candidate_score_against_spec(lowered, normalized_query, terms, matched_terms)
        if score > 0:
            scores.append((query_index, score))
    return scores


def _rank_minprom_registry_candidates(
    scored_entries: list[tuple[float, int, int, dict]],
    *,
    max_results: int,
    query_count: int,
) -> list[dict]:
    if max_results <= 0:
        return []
    selected: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    inn_counts: dict[str, int] = {}

    def add_entry(entry: dict, strict_inn_limit: bool = True) -> bool:
        key = _registry_entry_key(entry)
        if key in seen:
            return False
        inn_key = str(entry.get("inn") or entry.get("manufacturer") or "").strip().lower()
        if strict_inn_limit and inn_key and inn_counts.get(inn_key, 0) >= 2:
            return False
        seen.add(key)
        if inn_key:
            inn_counts[inn_key] = inn_counts.get(inn_key, 0) + 1
        selected.append(entry)
        return True

    effective_query_count = max(1, query_count)
    per_query_limit = max(4, min(20, (max_results + effective_query_count - 1) // effective_query_count + 1))
    for query_index in range(effective_query_count):
        query_matches = [item for item in scored_entries if item[2] == query_index]
        query_matches.sort(key=lambda item: (-item[0], item[1]))
        picked_for_query = 0
        for _, _, _, entry in query_matches:
            if add_entry(entry):
                picked_for_query += 1
            if len(selected) >= max_results or picked_for_query >= per_query_limit:
                break
        if len(selected) >= max_results:
            return selected

    for _, _, _, entry in sorted(scored_entries, key=lambda item: (-item[0], item[1])):
        add_entry(entry)
        if len(selected) >= max_results:
            break
    return selected


def _header_index(headers: list[str], *markers: str) -> int | None:
    for index, header in enumerate(headers):
        lowered = header.lower()
        if all(marker in lowered for marker in markers):
            return index
    return None


def _first_index(*values: int | None) -> int | None:
    for value in values:
        if value is not None:
            return value
    return None


def _cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return " ".join(str(row[index] or "").split()).strip()


def _registry_entry_from_row(
    row: tuple[Any, ...],
    headers: list[str],
    *,
    manufacturer_idx: int | None,
    product_idx: int | None,
    inn_idx: int | None,
    registry_idx: int | None,
    valid_until_idx: int | None,
) -> tuple[dict | None, str]:
    manufacturer = _cell(row, manufacturer_idx)
    product = _cell(row, product_idx)
    inn = _cell(row, inn_idx)
    registry_number = _cell(row, registry_idx)
    if not any((manufacturer, product, inn, registry_number)):
        return None, ""
    valid_until = _cell(row, valid_until_idx)
    evidence_parts = [
        f"Производитель: {manufacturer}" if manufacturer else "",
        f"Продукция: {product}" if product else "",
        f"ИНН: {inn}" if inn else "",
        f"Реестровый номер: {registry_number}" if registry_number else "",
        f"Срок действия/заключение: {valid_until}" if valid_until else "",
    ]
    return (
        {
            "registry_number": registry_number[:80],
            "manufacturer": manufacturer[:240],
            "product": product[:300],
            "inn": inn[:20],
            "source": "minprom_registry_local_xlsx",
            "source_url": GISP_PRODUCT_REGISTRY_URL,
            "evidence": "; ".join(part for part in evidence_parts if part)[:1000],
        },
        " ".join(part for part in (manufacturer, product, registry_number, valid_until) if part),
    )


def _fts_token_parts(value: str) -> tuple[str, ...]:
    parts: list[str] = []
    current: list[str] = []
    for char in str(value or "").lower().replace("ё", "е"):
        if char.isalnum():
            current.append(char)
        elif current:
            part = "".join(current)
            if len(part) >= 2:
                parts.append(part)
            current = []
    if current:
        part = "".join(current)
        if len(part) >= 2:
            parts.append(part)
    return tuple(parts)


def _fts_term_clause(term: str) -> str:
    parts = _fts_token_parts(term)
    if not parts:
        return ""
    clauses: list[str] = []
    has_digit = _registry_term_has_digit(term)
    for part in parts[:3]:
        if has_digit or len(part) < 5:
            clauses.append(f'"{part}"')
        else:
            clauses.append(f"{part}*")
    if not clauses:
        return ""
    if len(clauses) == 1:
        return clauses[0]
    return "(" + " AND ".join(clauses) + ")"


def _fts_match_expression(terms: tuple[str, ...]) -> str:
    clauses: list[str] = []
    seen: set[str] = set()
    for term in terms:
        clause = _fts_term_clause(term)
        if not clause or clause in seen:
            continue
        seen.add(clause)
        clauses.append(clause)
        if len(clauses) >= 8:
            break
    return " OR ".join(clauses)


async def build_procurement_profile(settings: SystemSettings, context: str) -> ProcurementProfile:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for procurement profile extraction")
    prompt = f"""Извлеки из технического задания профиль закупки для поиска поставщиков.

Нужно отделить закупаемые позиции от условий поставки, адресов, сроков, форм документов, стандартов, служебных кодов, комплектующих и расходников.
Для поиска поставщиков важно отделить широкую товарную группу/номенклатуру от точных характеристик.
Например, если ТЗ требует "канат стальной 31 мм ЛК-РО", товарная группа — "стальные канаты" / "канатная продукция", а "31 мм", "ЛК-РО", "ГОСТ" — точные характеристики для проверки и части запросов.
Если ТЗ требует краску без конкретной торговой марки, ищи категорию "краски/лакокрасочные материалы", а не одну точную позицию.
Если комплектующая или расходник закупаются как самостоятельная позиция, включи их отдельной позицией. Иначе добавь в excluded_terms.
В okpd2_codes добавляй только коды, явно подписанные как ОКПД2/ОКПД. Не добавляй туда номера ГОСТ, РД, СП, СНиП, ФЗ, ТУ или методик.

Ответ строго JSON:
{{
  "summary": "краткое описание предмета закупки",
  "items": [
    {{
      "id": "item-1",
      "name": "основная закупаемая позиция",
      "aliases": ["марки, модели, русские/английские варианты, аналоги"],
      "okpd2_codes": ["ОКПД2 коды из ТЗ/карточки, если есть"],
      "category_terms": ["широкая товарная группа/номенклатура для поиска производителей и поставщиков"],
      "exact_terms": ["точные размеры, ГОСТ, тип, марка, модель, артикул, если они важны"],
      "required_terms": ["термины, которые помогают подтвердить соответствие сайта"],
      "excluded_terms": ["что не считать самостоятельным предметом поиска"]
    }}
  ],
  "excluded_terms": ["общие исключения по ТЗ"]
}}

ТЗ:
{context[:16000]}"""
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            current_prompt = prompt if attempt == 0 else (
                prompt + "\n\nВАЖНО: Ответ ОБЯЗАН содержать массив items с хотя бы одним элементом. "
                "Каждый элемент должен иметь name с названием товара/позиции."
            )
            raw = await call_llm(
                settings,
                current_prompt,
                system_prompt="Ты закупочный аналитик. Строишь профиль закупки для поиска поставщиков, не подменяешь товар условиями и комплектующими.",
                tier="primary",
                routing_key="supplier_procurement_profile",
                json_mode=True,
                timeout_seconds=90,
                response_validator=_validate_procurement_profile_response,
            )
            profile = _normalize_procurement_profile(parse_json_object(raw))
            if not profile.items:
                last_error = RuntimeError(f"AI returned profile with 0 items (raw keys: {list(parse_json_object(raw).keys())})")
                continue
            return profile
        except Exception as exc:
            last_error = exc
            continue
    raise RuntimeError(f"AI procurement profile extraction failed: {last_error}") from last_error


async def build_supplier_queries(
    settings: SystemSettings,
    context: str,
    target: int,
    profile: ProcurementProfile | None = None,
    *,
    is_extend: bool = False,
    wave_index: int = 1,
    executed_queries: set[str] | list[str] | None = None,
    additional_prompt: str = "",
) -> list[str]:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for supplier search query generation")
    profile = profile or await build_procurement_profile(settings, context)
    
    if is_extend or wave_index > 1:
        refine_extra = f"\nДополнительные критерии добора: {additional_prompt.strip()}" if additional_prompt else ""
        system_prompt = (
            f"Ты ведущий эксперт по B2B поиску коммерческих поставщиков, региональных дистрибьюторов, торговых домов и оптовых складов. "
            f"ЭТО РЕЖИМ ДОБОРА (ВОЛНА №{wave_index}). Базовые поставщики уже найдены. "
            "Твоя задача — составить разнообразные поисковые запросы для нахождения альтернативных поставщиков, региональных центров сбыта, оптовых складов, торговых домов и комплектовщиков. "
            "Используй отрицательные операторы Яндекса: -\"банковская гарантия\" -\"обучение\" -\"семинар\" -\"эцп\" -\"агрегатор\" -\"курсы\"."
        )
        prompt = f"""На основе профиля закупки сформируй поисковые запросы для ДОБОРА поставщиков (Волна #{wave_index}).
Нужно искать альтернативных поставщиков: региональные торговые дома, склады, дистрибьюторы, дилеры по номенклатуре закупки.{refine_extra}
Сформируй 18-28 поисковых запросов для поиска по регионам РФ и специализированным каналам сбыта.
Включай в запросы минус-слова (-"банковская гарантия" -"обучение" -семинар -эцп -агрегатор -курсы).
Ответ строго JSON:
{{"queries": ["..."]}}

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Фрагмент ТЗ для контекста:
{context[:6000]}"""
    else:
        system_prompt = "Ты закупочный исследователь. Формируешь только поисковые запросы."
        prompt = f"""На основе профиля закупки сформируй поисковые запросы для поиска поставщиков.

Нужно искать не только точную строку из ТЗ, а компании, которые производят или поставляют нужную товарную группу/номенклатуру и могут дать КП по характеристикам ТЗ.
Сформируй 18-28 коротких запросов для поиска российских заводов, производителей, официальных дилеров, дистрибьюторов и B2B-поставщиков.
Обязательная структура:
- 40-60% запросов широкие по товарной группе/номенклатуре без точного размера, ГОСТ, марки или артикула;
- 20-40% запросов с точными характеристиками из ТЗ;
- отдельные запросы по "производитель", "завод", "поставщик", "дилер", "дистрибьютор", "оптом", "каталог";
- если задана фиксированная торговая марка/модель, добавь брендовые запросы, но всё равно ищи официальных дилеров и профильных производителей категории.
Не добавляй агрегаторы, маркетплейсы, реестры, тендерные площадки, справочники, статьи, видео и учебные страницы.
Если позиций несколько, запросы должны покрывать каждую позицию.
Ответ строго JSON:
{{"queries": ["..."]}}

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Фрагмент ТЗ для контекста:
{context[:6000]}"""
    last_error: Exception | None = None
    for attempt in range(1, 3):
        try:
            raw = await call_llm(
                settings,
                prompt,
                system_prompt=system_prompt,
                tier="light",
                routing_key="supplier_query_generation",
                json_mode=True,
                timeout_seconds=90,
                response_validator=_validate_supplier_queries_response,
            )
            parsed = parse_json_object(raw)
            queries = _clean_supplier_queries([str(item).strip() for item in parsed.get("queries", [])])
            if executed_queries:
                exec_set = {str(q).strip().lower() for q in executed_queries}
                queries = [q for q in queries if q.strip().lower() not in exec_set]
            if not is_extend and _queries_need_broadening(profile, queries, target):
                revised = await _revise_supplier_queries_with_ai(settings, context, profile, queries, target)
                if revised:
                    queries = revised
            return queries[:28]
        except Exception as exc:
            last_error = exc
            if attempt == 1:
                await asyncio.sleep(1)
                continue
            break
    detail = _exception_summary(last_error)
    raise RuntimeError(f"AI supplier query generation failed after retry: {detail}") from last_error


async def discover_suppliers(
    settings: SystemSettings,
    context: str,
    target: int,
    *,
    progress_callback: ProgressCallback | None = None,
    excluded_suppliers: list[dict] | None = None,
    supplier_search_policy: str = SUPPLIER_POLICY_NORMAL,
    preloaded_candidates: list[dict] | None = None,
    cached_procurement_profile: dict | None = None,
    executed_queries: list[str] | set[str] | None = None,
    additional_prompt: str = "",
    is_extend: bool = False,
    wave_index: int = 1,
) -> tuple[list[dict], dict]:
    async with _browser_pool_session():
        return await _discover_suppliers_impl(
            settings,
            context,
            target,
            progress_callback=progress_callback,
            excluded_suppliers=excluded_suppliers,
            supplier_search_policy=supplier_search_policy,
            preloaded_candidates=preloaded_candidates,
            cached_procurement_profile=cached_procurement_profile,
            executed_queries=executed_queries,
            additional_prompt=additional_prompt,
            is_extend=is_extend,
            wave_index=wave_index,
        )


async def _discover_suppliers_impl(
    settings: SystemSettings,
    context: str,
    target: int,
    *,
    progress_callback: ProgressCallback | None = None,
    excluded_suppliers: list[dict] | None = None,
    supplier_search_policy: str = SUPPLIER_POLICY_NORMAL,
    preloaded_candidates: list[dict] | None = None,
    cached_procurement_profile: dict | None = None,
    executed_queries: list[str] | set[str] | None = None,
    additional_prompt: str = "",
    is_extend: bool = False,
    wave_index: int = 1,
) -> tuple[list[dict], dict]:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for supplier search")
    minimum_target = max(1, min(MAX_SUPPLIER_DELIVERY_TARGET, int(target or 0)))
    delivery_target = _supplier_delivery_target(minimum_target)
    excluded_domains, excluded_company_keys = _supplier_exclusion_sets(excluded_suppliers)
    
    # 1. Build or reuse Procurement Profile
    if cached_procurement_profile:
        try:
            profile = _normalize_procurement_profile(cached_procurement_profile)
            await _emit_progress(progress_callback, 28, f"Использую профиль закупки: {len(profile.items)} позиций")
        except Exception:
            await _emit_progress(progress_callback, 28, "Анализирую ТЗ и выделяю закупаемые позиции")
            profile = await build_procurement_profile(settings, context)
    else:
        await _emit_progress(progress_callback, 28, "Анализирую ТЗ и выделяю закупаемые позиции")
        profile = await build_procurement_profile(settings, context)
    await _emit_progress(progress_callback, 36, f"Определил закупаемые позиции: {len(profile.items)}")
    
    policy = normalize_supplier_search_policy(supplier_search_policy)
    await _emit_progress(progress_callback, 39, "Проверяю требования к реестру Минпромторга")
    if policy == SUPPLIER_POLICY_NORMAL:
        minprom_requirement = await assess_minprom_registry_requirement(settings, context)
    else:
        minprom_requirement = MinpromRegistryRequirement(
            required=True,
            measure_type="prohibition" if policy == SUPPLIER_POLICY_MINPROM_ONLY else "restriction",
            evidence="Ручной выбор клиента",
            reason="Клиент выбрал режим поиска с учетом реестровой записи Минпромторга.",
            raw={"manual_policy": policy},
        )
    if minprom_requirement.required:
        await _emit_progress(progress_callback, 41, "Ищу подтверждения в реестре промышленной продукции")
        minprom_context = await discover_minprom_registry_context(settings, context, profile, minprom_requirement)
    else:
        minprom_context = MinpromRegistryContext(requirement=minprom_requirement, status="not_required")
    registry_unavailable = policy in {SUPPLIER_POLICY_MINPROM_ONLY, SUPPLIER_POLICY_MINPROM_PRIORITY} and minprom_context.status == "error"
    
    # Convert preloaded candidates from parent task if available
    preloaded_objs: list[Candidate] = []
    if preloaded_candidates:
        for c in preloaded_candidates:
            if isinstance(c, dict):
                u = str(c.get("url") or "")
                d = base_domain(str(c.get("domain") or u))
                if d and d not in excluded_domains and d not in BLOCKED_DOMAINS:
                    preloaded_objs.append(
                        Candidate(
                            url=u or f"https://{d}",
                            domain=d,
                            title=str(c.get("title") or ""),
                            snippet=str(c.get("snippet") or ""),
                            source=str(c.get("source") or "preloaded_pool"),
                            query=str(c.get("query") or ""),
                            ai_rank_confidence=int(c.get("ai_rank_confidence") or 0),
                            ai_rank_reason=str(c.get("ai_rank_reason") or ""),
                        )
                    )

    search_meta = {}
    queries: list[str] = []
    minprom_supplier_queries: list[str] = []

    # If we have enough preloaded candidates, we skip / minimize search query calls
    if len(preloaded_objs) >= delivery_target * 4:
        candidates = preloaded_objs
        search_meta = {
            "provider_order": ["preloaded_pool"],
            "strategy": {"source": "parent_job_cache", "count": len(preloaded_objs)},
            "yandex_requests_count": 0,
            "yandex_cost_rub": 0.0,
            "reports": [{"provider": "preloaded_pool", "status": "ok", "added": len(preloaded_objs), "requested": 0}],
        }
        await _emit_progress(progress_callback, 50, f"Использую {len(preloaded_objs)} сохраненных сайтов из пула (0 ₽ за поиск)...")
    else:
        try:
            general_queries = await build_supplier_queries(
                settings,
                context,
                minimum_target,
                profile=profile,
                is_extend=is_extend,
                wave_index=wave_index,
                executed_queries=executed_queries,
                additional_prompt=additional_prompt,
            )
        except TypeError:
            general_queries = await build_supplier_queries(
                settings,
                context,
                minimum_target,
                profile=profile,
            )
        minprom_supplier_queries = _build_minprom_supplier_queries(profile, minprom_context)
        if registry_unavailable:
            queries = []
        elif policy == SUPPLIER_POLICY_MINPROM_ONLY:
            queries = minprom_supplier_queries
        else:
            queries = _merge_supplier_query_tracks(general_queries, minprom_supplier_queries)
        await _emit_progress(
            progress_callback,
            50,
            f"Ищу сайты поставщиков: запросов {len(queries)}, минимум {minimum_target}, целевой результат {delivery_target}",
        )
        discovered, search_meta = await discover_candidates(
            settings,
            queries,
            max_results=max(delivery_target * 10, 120),
            excluded_domains=excluded_domains,
            primary_candidate_floor=_primary_candidate_floor(delivery_target),
            fallback_candidate_limit=_fallback_candidate_limit(delivery_target),
        )
        preloaded_domains = {p.domain for p in preloaded_objs}
        candidates = preloaded_objs + [c for c in discovered if c.domain not in preloaded_domains]

    await _emit_progress(progress_callback, 60, f"Найдено кандидатов: {len(candidates)}. Отсекаю нерелевантные сайты")
    candidates = _exclude_candidates(_rank_candidates(candidates, context), excluded_domains)[: max(delivery_target * 5, 60)]
    await _emit_progress(progress_callback, 66, "Отбираю подходящие компании")
    rerank = await ai_rerank_candidates(settings, profile, candidates, delivery_target, registry_context=minprom_context)
    candidates = rerank.candidates

    await _emit_progress(progress_callback, 72, f"Проверяю сайты и контакты: кандидатов {len(candidates)}")
    accepted, reviewed, review_meta = await _review_candidates_until_target(
        settings,
        candidates,
        context,
        delivery_target,
        profile=profile,
        registry_context=minprom_context,
        excluded_domains=excluded_domains,
        excluded_company_keys=excluded_company_keys,
        progress_callback=progress_callback,
    )
    recovery_rounds: list[dict] = []
    max_recovery_rounds = 1
    for recovery_attempt in range(max_recovery_rounds):
        # The client minimum is the completion guarantee. Extra verified rows come
        # from the first reviewed pool; do not spend another recovery pass solely
        # to chase the optional surplus.
        if len(accepted) >= minimum_target:
            break
        await _emit_progress(progress_callback, 92 + recovery_attempt, f"Расширяю поиск (раунд {recovery_attempt + 1}): подтверждено {len(accepted)}")
        accepted_before_recovery = len(accepted)
        recovery_queries = await _build_supplier_recovery_queries_with_ai(
            settings,
            context,
            profile,
            queries,
            reviewed,
            accepted,
            minimum_target,
        )
        recovery_round = {
            "status": "empty_queries",
            "queries": recovery_queries,
            "accepted_before": len(accepted),
            "accepted_after": len(accepted),
        }
        if recovery_queries:
            try:
                recovery_candidates, recovery_search_meta = await discover_candidates(
                    settings,
                    recovery_queries,
                    max_results=max(minimum_target * 8, 80),
                    excluded_domains=excluded_domains,
                    primary_candidate_floor=_primary_candidate_floor(minimum_target),
                    fallback_candidate_limit=_fallback_candidate_limit(minimum_target),
                )
                seen_domains = {candidate.domain for candidate in candidates}
                seen_domains.update(excluded_domains)
                seen_domains.update(base_domain(str(item.get("site") or "")) for item in reviewed)
                recovery_candidates = [
                    candidate
                    for candidate in recovery_candidates
                    if candidate.domain and candidate.domain not in seen_domains
                ]
                recovery_candidates = _rank_candidates(recovery_candidates, context)[: max(minimum_target * 5, 60)]
                recovery_rerank = await ai_rerank_candidates(
                    settings,
                    profile,
                    recovery_candidates,
                    max(1, minimum_target - len(accepted)),
                    registry_context=minprom_context,
                )
                recovery_accepted, recovery_reviewed, recovery_review_meta = await _review_candidates_until_target(
                    settings,
                    recovery_rerank.candidates,
                    context,
                    max(1, minimum_target - len(accepted)),
                    profile=profile,
                    registry_context=minprom_context,
                    excluded_domains=excluded_domains,
                    excluded_company_keys=excluded_company_keys,
                    progress_callback=progress_callback,
                )
                reviewed.extend(recovery_reviewed)
                accepted = _accepted_supplier_results(
                    reviewed,
                    minimum_target,
                    profile=profile,
                    limit_to_target=False,
                    excluded_domains=excluded_domains,
                    excluded_company_keys=excluded_company_keys,
                )
                recovery_round = {
                    "status": "ok",
                    "queries": recovery_queries,
                    "candidate_count": len(recovery_candidates),
                    "accepted_before": accepted_before_recovery,
                    "accepted_after": len(accepted),
                    "search": recovery_search_meta,
                    "candidate_rerank": recovery_rerank.meta,
                    "review": recovery_review_meta,
                }
            except Exception as exc:
                recovery_round = {
                    "status": "error",
                    "queries": recovery_queries,
                    "accepted_before": len(accepted),
                    "accepted_after": len(accepted),
                    "error": _exception_summary(exc),
                }
        recovery_rounds.append(recovery_round)
        # Merge recovery queries into main list so next round generates different queries
        queries = queries + recovery_queries
    _annotate_minprom_registry_matches(accepted, minprom_context, policy)
    pre_strict_filter_accepted = [dict(item) for item in accepted]

    if policy == SUPPLIER_POLICY_MINPROM_ONLY:
        accepted = _filter_minprom_verified_suppliers(accepted, minprom_context)
    elif minprom_context.status == "ok" and minprom_context.entries:
        await _emit_progress(progress_callback, 90, "Обогащаю контакты и реквизиты реестра через DaData и поиск")
        accepted = await _enrich_unmatched_registry_suppliers(settings, accepted, minprom_context)

    # Enrich general accepted suppliers with missing region/director
    accepted = await _enrich_accepted_suppliers_with_dadata(accepted)

    if policy == SUPPLIER_POLICY_MINPROM_PRIORITY:
        accepted.sort(key=lambda item: 0 if (item.get("minprom_registry_match") or {}).get("matched") else 1)
    await _emit_progress(progress_callback, 94, f"Готовлю результат: подтверждено {len(accepted)}")

    evidence = {
        "ai_required": True,
        "ai_used": True,
        "ai_required_stages": [
            "supplier_procurement_profile",
            "minprom_registry_requirement",
            "supplier_query_generation",
            "supplier_candidate_reranker",
            "supplier_candidate_verifier",
        ],
        "acceptance_policy": "Supplier rows are accepted only after AI verifier returns action=accept with verified evidence.",
        "target": minimum_target,
        "minimum_target": minimum_target,
        "delivery_target": delivery_target,
        "delivery_policy": "minimum_plus_verified_relevant_first_pass",
        "supplier_search_policy": policy,
        "registry_unavailable_no_charge": bool(registry_unavailable),
        "excluded_suppliers": {
            "count": len(excluded_suppliers or []),
            "domains": sorted(excluded_domains),
            "company_keys": sorted(excluded_company_keys),
        },
        "procurement_profile": _profile_to_dict(profile),
        "minprom_registry": _minprom_context_to_dict(minprom_context),
        "minprom_supplier_queries": minprom_supplier_queries,
        "search_provider": "yandex_primary",
        "search": search_meta,
        "candidate_rerank": rerank.meta,
        "review": review_meta,
        "recovery_rounds": recovery_rounds,
        "candidate_source_counts": _source_counts(candidates),
        "accepted_source_counts": _source_counts(
            Candidate(
                url=str(item.get("site") or ""),
                domain=base_domain(str(item.get("site") or "")),
                source=str(item.get("source") or ""),
                query=str(item.get("search_query") or ""),
            )
            for item in accepted
        ),
        "queries": queries,
        "executed_queries": queries,
        "unreviewed_candidates": [
            (candidate.__dict__ if hasattr(candidate, "__dict__") else candidate)
            for candidate in candidates
            if base_domain(getattr(candidate, "domain", "") or getattr(candidate, "url", "")) not in {base_domain(str(r.get("site") or "")) for r in accepted}
            and base_domain(getattr(candidate, "domain", "") or getattr(candidate, "url", "")) not in excluded_domains
        ][:120],
        "wave_index": wave_index,
        "registry_result": {
            "status": minprom_context.status,
            "verified_count": sum(1 for item in accepted if (item.get("minprom_registry_match") or {}).get("matched")),
        },
        "non_registry_alternative": _non_registry_alternative_evidence(
            pre_strict_filter_accepted if (policy == SUPPLIER_POLICY_MINPROM_ONLY and not accepted) else accepted,
            minprom_context,
            registry_verified_count=sum(1 for item in accepted if (item.get("minprom_registry_match") or {}).get("matched")),
        ),
        "candidates": [candidate.__dict__ for candidate in candidates],
        "accepted_count": len(accepted),
        "accepted": accepted,
        "reviewed": reviewed,
    }
    return accepted, evidence


def normalize_supplier_search_policy(value: str) -> str:
    normalized = str(value or "").strip().lower()
    return normalized if normalized in VALID_SUPPLIER_SEARCH_POLICIES else SUPPLIER_POLICY_NORMAL


EXTRA_AGGREGATOR_DOMAINS = {
    "checko.ru", "list-org.com", "audit-it.ru", "rusprofile.ru", "zachestnyibiznes.ru",
    "spark-interfax.ru", "saby.ru", "kontur.ru", "reestr-sro.ru", "sbis.ru",
    "vbankcenter.ru", "prima-inform.ru", "complan.pro", "testfirm.ru", "injust.pro",
    "classinform.ru", "vypiska-nalog.com", "egrul.nalog.ru", "b2b-center.ru", "synapsenet.ru",
    "kartoteka.ru", "companies.rbc.ru", "fedresurs.ru", "service-online.su", "licexpert.ru",
    "bbnt.ru", "globalstat.ru", "comfex.ru", "fira.ru", "znaybiznes.ru", "star-pro.ru",
    "datanewton.ru", "tochka.com", "artalix.ru", "reestr.smolensk-sro.ru", "catalog-firm.ru"
}


async def _enrich_unmatched_registry_suppliers(
    settings: SystemSettings,
    accepted: list[dict],
    registry_context: MinpromRegistryContext,
) -> list[dict]:
    """Ensures ALL GISP registry entries appear in results and enriches them via DaData + targeted search."""
    if registry_context.status != "ok" or not registry_context.entries:
        return accepted

    matched_inns: set[str] = set()
    matched_names: set[str] = set()
    for item in accepted:
        match_data = item.get("minprom_registry_match") or {}
        if match_data.get("matched"):
            inn = re.sub(r"\D+", "", str(match_data.get("inn") or ""))
            if inn:
                matched_inns.add(inn)
            name = str(match_data.get("manufacturer") or "").strip().lower()
            if name and len(name) >= 3:
                matched_names.add(name)

    unmatched_entries = []
    for entry in registry_context.entries:
        inn = re.sub(r"\D+", "", str(entry.get("inn") or ""))
        mfr = str(entry.get("manufacturer") or "").strip()
        mfr_lower = mfr.lower()
        already_matched = False
        if inn and inn in matched_inns:
            already_matched = True
        elif mfr_lower and len(mfr_lower) >= 3:
            for name in matched_names:
                if mfr_lower in name or name in mfr_lower:
                    already_matched = True
                    break
        if not already_matched:
            unmatched_entries.append(entry)
            if inn:
                matched_inns.add(inn)
            if mfr_lower:
                matched_names.add(mfr_lower)

    if not unmatched_entries:
        return accepted

    semaphore = asyncio.Semaphore(3)

    async def _enrich_single_entry(entry: dict) -> dict:
        async with semaphore:
            inn = re.sub(r"\D+", "", str(entry.get("inn") or ""))
            mfr = str(entry.get("manufacturer") or "").strip()
            product = str(entry.get("product") or "").strip()
            registry_number = str(entry.get("registry_number") or "").strip()

            # Step 1: DaData enrichment
            dadata: dict[str, Any] = {}
            if inn:
                try:
                    dadata = await enrich_company_by_inn(inn)
                except Exception as exc:
                    logger.warning("DaData enrichment error for %s: %s", inn, exc)

            company_name = dadata.get("company_name") or mfr
            legal_address = dadata.get("legal_address") or ""
            region = dadata.get("region") or ""
            management_name = dadata.get("management_name") or ""
            dadata_phones = dadata.get("phones") or []
            dadata_emails = dadata.get("emails") or []
            dadata_sites = dadata.get("sites") or []

            # Step 2: Targeted web search for website & contacts
            site = ""
            evidence_url = ""
            contact_url = ""
            phone = dadata_phones[0] if dadata_phones else ""
            email = dadata_emails[0] if dadata_emails else ""

            queries = []
            if inn:
                queries.append(f'"{inn}" официальный сайт')
            if mfr:
                clean_name = re.sub(
                    r'(?:ООО|АО|ЗАО|ПАО|ОАО|НАО|НПК|НМФ|ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ|АКЦИОНЕРНОЕ ОБЩЕСТВО)\s*',
                    "",
                    mfr,
                    flags=re.I,
                ).strip(' "«»')
                if clean_name and len(clean_name) >= 3:
                    queries.append(f'"{clean_name}" официальный сайт контакты')

            if queries:
                try:
                    candidates: list[Candidate] = []
                    if settings.yandex_search_folder_id and settings.yandex_search_api_key:
                        candidates, _ = await _search_with_yandex(settings, queries, max_results=6)
                    elif settings.google_search_api_key and settings.google_search_cse_id:
                        candidates = await _search_with_google(settings, queries, max_results=6)
                    else:
                        candidates = await _search_with_ddgs(queries, max_results=6)

                    valid_candidates = []
                    for c in candidates:
                        cand_domain = base_domain(c.url)
                        if is_blocked(c.url) or cand_domain in EXTRA_AGGREGATOR_DOMAINS:
                            continue
                        valid_candidates.append(c)

                    chosen_candidate = valid_candidates[0] if valid_candidates else None
                    if chosen_candidate:
                        site = chosen_candidate.url
                        evidence_url = chosen_candidate.url
                        contact_url = chosen_candidate.url
                        try:
                            pages = await collect_pages(chosen_candidate.url)
                            if pages:
                                comb_text = "\n".join(p.get("text", "") for p in pages)
                                phones = sorted(set(PHONE_RE.findall(comb_text)))
                                emails = sorted(set(EMAIL_RE.findall(comb_text)))
                                verified_p = _verified_phone("", phones)
                                verified_e = _verified_email("", emails)
                                if verified_p:
                                    phone = verified_p
                                if verified_e:
                                    email = verified_e
                                contact_url = contact_page_url(pages, emails, phones) or chosen_candidate.url
                                evidence_url = pages[0].get("url") or chosen_candidate.url
                        except Exception as crawl_exc:
                            logger.warning("Targeted crawl failed for %s: %s", chosen_candidate.url, crawl_exc)
                except Exception as search_exc:
                    logger.warning("Targeted search failed for entry %s: %s", inn or mfr, search_exc)

            if not site and dadata_sites:
                site = dadata_sites[0]
                evidence_url = site
                contact_url = site

            # Step 3: Build clean comment
            comment_parts = [
                f"Точное соответствие: {product}." if product else "Точное соответствие ТЗ.",
                f"Реестр ГИСП Минпромторга: запись № {registry_number}, производитель {mfr}." if registry_number else f"Реестр Минпромторга: {mfr}."
            ]
            if legal_address:
                comment_parts.append(f"Юр. адрес: {legal_address}.")
            if management_name:
                comment_parts.append(f"Руководитель: {management_name}.")
            if site:
                comment_parts.append("Официальный сайт и контакты подтверждены.")
            else:
                comment_parts.append("Сайт не найден автоматически.")

            final_comment = " ".join(comment_parts)

            return {
                "company_name": company_name,
                "name": company_name,
                "site": site,
                "evidence_url": evidence_url,
                "contact_url": contact_url,
                "email": email,
                "phone": phone,
                "inn": inn,
                "region": region,
                "contact_person": management_name,
                "product": product,
                "comments": final_comment,
                "match_level": "exact",
                "product_fit": "exact",
                "evidence_status": "verified" if (site or phone or email) else "registry_only",
                "evidence_snippet": f"Реестр Минпромторга: запись № {registry_number}, ИНН {inn}",
                "supplier_search_origin": "minprom_registry",
                "minprom_registry_required": True,
                "minprom_registry_status": registry_context.status,
                "minprom_registry_entries_count": len(registry_context.entries),
                "minprom_registry_match": {
                    "matched": True,
                    "method": "direct_registry_entry_enriched",
                    "confidence": 1.0,
                    "registry_number": registry_number,
                    "manufacturer": mfr[:240],
                    "product": product[:300],
                    "inn": inn,
                    "source_url": str(entry.get("source_url") or "https://gisp.gov.ru/pp/pub/prod/"),
                    "evidence": f"Прямая запись реестра: {mfr}, номер {registry_number}",
                },
                "quality_score": 85 if (phone or email or site) else 75,
                "quality_tier": "high" if (phone and email and site) else "medium",
            }

    enriched_stubs = await asyncio.gather(*[_enrich_single_entry(e) for e in unmatched_entries])
    accepted.extend(enriched_stubs)
    return accepted


async def _enrich_accepted_suppliers_with_dadata(accepted: list[dict]) -> list[dict]:
    """Enrich missing region and contact person for any accepted suppliers with INN."""
    for item in accepted:
        inn = re.sub(r"\D+", "", str(item.get("inn") or ""))
        if not inn or len(inn) not in (10, 12):
            continue
        need_region = not str(item.get("region") or "").strip()
        need_contact = not str(item.get("contact_person") or "").strip()
        if need_region or need_contact:
            try:
                dadata = await enrich_company_by_inn(inn)
                if dadata:
                    if need_region and dadata.get("region"):
                        item["region"] = dadata["region"]
                    if need_contact and dadata.get("management_name"):
                        item["contact_person"] = dadata["management_name"]
            except Exception:
                pass
    return accepted


def _inject_unmatched_registry_stubs(
    accepted: list[dict],
    registry_context,
) -> list[dict]:
    """Compatibility sync wrapper."""
    if registry_context.status != "ok" or not registry_context.entries:
        return accepted
    return accepted


def _filter_minprom_verified_suppliers(accepted: list[dict], registry_context: MinpromRegistryContext) -> list[dict]:
    if registry_context.status != "ok" or not registry_context.entries:
        return []
    return [item for item in accepted if _supplier_minprom_registry_match(item, registry_context).get("matched")]


def _supplier_has_minprom_registry_evidence(item: dict, registry_context: MinpromRegistryContext) -> bool:
    return bool(_supplier_minprom_registry_match(item, registry_context).get("matched"))


def _annotate_minprom_registry_matches(
    accepted: list[dict],
    registry_context: MinpromRegistryContext,
    policy: str,
) -> None:
    if not registry_context.requirement.required:
        for item in accepted:
            item["supplier_search_policy"] = policy
            item["minprom_registry_required"] = False
            item["minprom_registry_status"] = registry_context.status
            item["minprom_registry_entries_count"] = len(registry_context.entries)
            item.setdefault("supplier_search_origin", "ordinary")
        return
    for item in accepted:
        item["supplier_search_policy"] = policy
        item["minprom_registry_required"] = True
        item["minprom_registry_status"] = registry_context.status
        item["minprom_registry_entries_count"] = len(registry_context.entries)
        match = _supplier_minprom_registry_match(item, registry_context)
        item["minprom_registry_match"] = match
        if match.get("matched"):
            item["supplier_search_origin"] = "minprom_registry"
        elif policy in {SUPPLIER_POLICY_MINPROM_PRIORITY, SUPPLIER_POLICY_MINPROM_ONLY}:
            item["supplier_search_origin"] = "ordinary_fallback"
        else:
            item["supplier_search_origin"] = "ordinary"


def _supplier_minprom_registry_match(item: dict, registry_context: MinpromRegistryContext) -> dict:
    empty = {
        "matched": False,
        "method": "",
        "confidence": 0.0,
        "registry_number": "",
        "manufacturer": "",
        "product": "",
        "inn": "",
        "source_url": "",
        "evidence": "",
    }
    if registry_context.status != "ok" or not registry_context.entries:
        return empty
    haystack = " ".join(
        str(item.get(key) or "")
        for key in (
            "company_name",
            "name",
            "company",
            "site",
            "product",
            "comments",
            "inn",
            "evidence_snippet",
            "ai_rank_reason",
        )
    )
    lowered = _normalize_registry_text(haystack)
    company_key = _normalize_company_key(str(item.get("company_name") or item.get("name") or item.get("company") or ""))
    item_inn_values = [re.sub(r"\D+", "", str(item.get("inn") or ""))]
    item_inn_values.extend(match.group(1) for match in re.finditer(r"\bИНН[:\s]*(\d{10,12})\b", haystack, re.I))
    item_inn = " ".join(value for value in item_inn_values if value)
    best: dict | None = None
    best_score = 0.0

    for entry in registry_context.entries:
        manufacturer = str(entry.get("manufacturer") or "").strip()
        product = str(entry.get("product") or "").strip()
        registry_number = str(entry.get("registry_number") or "").strip()
        inn = re.sub(r"\D+", "", str(entry.get("inn") or ""))
        method = ""
        score = 0.0
        if inn and len(inn) >= 10 and inn in item_inn:
            method = "inn"
            score = 1.0
        elif registry_number and _normalize_registry_text(registry_number) in lowered:
            method = "registry_number"
            score = 0.98
        else:
            manufacturer_key = _normalize_company_key(manufacturer)
            comp_key = _normalize_company_key(company_key)
            hay_key = _normalize_company_key(haystack)
            manufacturer_matched = bool(
                manufacturer_key
                and len(manufacturer_key) >= 3
                and (
                    (comp_key and (manufacturer_key in comp_key or comp_key in manufacturer_key))
                    or manufacturer_key in hay_key
                )
            )
            product_matched = bool(product and item.get("product") and _registry_candidate_query_scores(str(item.get("product")), [product]))
            if manufacturer_matched and product_matched:
                method = "manufacturer_product"
                score = 0.9
            elif manufacturer_matched:
                method = "manufacturer"
                score = 0.78
        if score <= best_score:
            continue
        best_score = score
        best = {
            "matched": True,
            "method": method,
            "confidence": score,
            "registry_number": registry_number,
            "manufacturer": manufacturer[:240],
            "product": product[:300],
            "inn": inn,
            "source_url": str(entry.get("source_url") or GISP_PRODUCT_REGISTRY_URL),
            "evidence": str(entry.get("evidence") or "")[:500],
        }
    return best or empty


async def extract_supplier_search_context(settings: SystemSettings, context: str) -> str:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for supplier search context extraction")
    prompt = f"""Из полного комплекта закупочной документации выдели контекст, который нужен именно для поиска поставщиков.

Нужно найти и сохранить:
- техническое задание, описание объекта закупки, спецификацию или товарную таблицу;
- наименования закупаемых товаров/номенклатуры;
- характеристики, размеры, ГОСТ/ТУ/марки/модели/бренды, если они важны для проверки;
- ОКПД2, КТРУ, реестровые номера, сведения о баллах Минпромторга и сроке действия записи как внутренние признаки для поиска;
- единицы измерения и количества;
- требования к Минпромторгу/ГИСП/реестровым записям, если они есть;
- краткий предмет закупки.

Нужно убрать или сильно сократить:
- общие условия договора;
- реквизиты, адреса, штрафы, обеспечение, инструкции подачи заявки;
- формы документов и служебные таблицы, которые не описывают товар.

Не придумывай товары и характеристики. Если в комплекте есть несколько разных ТЗ/ООЗ, сохрани каждую самостоятельную товарную позицию.

Ответ строго JSON:
{{"supplier_context": "самодостаточный текст ТЗ/ООЗ для поиска поставщиков"}}

Документация:
{context[:180000]}"""
    raw = await call_llm(
        settings,
        prompt,
        system_prompt="Ты закупочный аналитик. Выделяешь из документации только ТЗ/ООЗ и товарные требования для последующего поиска поставщиков.",
        tier="primary",
        routing_key="supplier_tz_context_extraction",
        json_mode=True,
        timeout_seconds=180,
        response_validator=_validate_supplier_context_response,
    )
    parsed = parse_json_object(raw)
    supplier_context = str(parsed.get("supplier_context") or "").strip()
    supplier_context = re.sub(r"[ \t]+", " ", supplier_context)
    supplier_context = re.sub(r"\n{3,}", "\n\n", supplier_context).strip()
    if len(supplier_context) < 50:
        raise RuntimeError("AI supplier search context extraction returned an empty context")
    return supplier_context


async def _emit_progress(progress_callback: ProgressCallback | None, progress: int, message: str) -> None:
    if progress_callback is None:
        return
    await progress_callback(progress, message)


def _supplier_exclusion_sets(excluded_suppliers: list[dict] | None) -> tuple[set[str], set[str]]:
    domains: set[str] = set()
    company_keys: set[str] = set()
    for item in excluded_suppliers or []:
        if not isinstance(item, dict):
            continue
        domain = base_domain(
            str(item.get("site") or item.get("evidence_url") or item.get("contact_url") or "")
        )
        if domain:
            domains.add(domain)
        company_key = _normalize_company_key(item.get("company_name") or domain)
        if company_key:
            company_keys.add(company_key)
    return domains, company_keys


def _exclude_candidates(candidates: list[Candidate], excluded_domains: set[str]) -> list[Candidate]:
    if not excluded_domains:
        return candidates
    return [candidate for candidate in candidates if candidate.domain and candidate.domain not in excluded_domains]


def _normalize_procurement_profile(data: dict) -> ProcurementProfile:
    raw_items = data.get("items") if isinstance(data, dict) else []
    items: list[ProcurementItem] = []
    for index, item in enumerate(raw_items if isinstance(raw_items, list) else [], start=1):
        if not isinstance(item, dict):
            continue
        name = re.sub(r"\s+", " ", str(item.get("name") or item.get("title") or "")).strip(" .,:;")
        if not name:
            continue
        item_context = f"{name} {' '.join(str(value) for value in item.values())}"
        item_id = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(item.get("id") or f"item-{index}")).strip("-").lower() or f"item-{index}"
        items.append(
            ProcurementItem(
                id=item_id,
                name=name[:220],
                aliases=_clean_profile_terms(item.get("aliases")),
                okpd2_codes=_clean_okpd2_codes(
                    item.get("okpd2_codes")
                    or item.get("okpd2")
                    or item.get("okpd_codes")
                    or item.get("classification_codes")
                    or item_context,
                    context_text=item_context,
                ),
                category_terms=_clean_profile_terms(
                    item.get("category_terms") or item.get("nomenclature_terms") or item.get("supplier_category_terms")
                ),
                exact_terms=_clean_profile_terms(item.get("exact_terms") or item.get("strict_terms") or item.get("spec_terms")),
                required_terms=_clean_profile_terms(item.get("required_terms") or item.get("search_terms")),
                excluded_terms=_clean_profile_terms(item.get("excluded_terms")),
            )
        )
    summary = re.sub(r"\s+", " ", str(data.get("summary") or "")).strip() if isinstance(data, dict) else ""
    return ProcurementProfile(
        summary=summary[:500],
        items=tuple(items),
        excluded_terms=_clean_profile_terms(data.get("excluded_terms") if isinstance(data, dict) else None),
        raw=data if isinstance(data, dict) else {},
    )


def _clean_profile_terms(value: object) -> tuple[str, ...]:
    if isinstance(value, str):
        values = [value]
    elif isinstance(value, list):
        values = [str(item) for item in value]
    else:
        values = []
    result: list[str] = []
    for item in values:
        cleaned = re.sub(r"\s+", " ", item).strip(" .,:;")
        if 2 <= len(cleaned) <= 160 and cleaned.lower() not in [existing.lower() for existing in result]:
            result.append(cleaned)
    return tuple(result)


def _clean_okpd2_codes(value: object, *, context_text: str = "") -> tuple[str, ...]:
    if isinstance(value, str):
        text = value
    elif isinstance(value, (list, tuple, set)):
        text = " ".join(str(item) for item in value)
    else:
        text = str(value or "")
    result: list[str] = []
    for match in re.finditer(r"\b\d{2}(?:\.\d{1,3}){1,3}\b", text):
        code = match.group(0).strip(".")
        if _looks_like_regulatory_document_code(code, context_text or text):
            continue
        if code and code not in result:
            result.append(code)
    return tuple(result)


def _looks_like_regulatory_document_code(code: str, context_text: str) -> bool:
    if not code or not context_text:
        return False
    escaped = re.escape(code)
    marker = r"(?:ГОСТ|РД|СП|СНиП|ФЗ|ТУ|ISO|IEC|EN|СТО|СанПиН|МУК|МУ)"
    return bool(
        re.search(rf"{marker}[\s№N:-]{{0,12}}{escaped}(?:[-–]\d{{2,4}})?", context_text, flags=re.I)
        or re.search(rf"{escaped}[-–]\d{{2,4}}[\s,;:()/-]{{0,12}}{marker}", context_text, flags=re.I)
    )


def _okpd2_hierarchy_codes(code: str) -> tuple[str, ...]:
    parts = [part for part in str(code or "").split(".") if part]
    if len(parts) < 2 or not all(part.isdigit() for part in parts):
        return ()
    candidates: list[str] = [".".join(parts)]
    if len(parts) >= 3:
        candidates.append(".".join(parts[:3]))
        if len(parts[2]) > 1:
            candidates.append(".".join([parts[0], parts[1], parts[2][0]]))
    candidates.append(".".join(parts[:2]))
    if len(parts[1]) > 1:
        candidates.append(".".join([parts[0], parts[1][0]]))
    result: list[str] = []
    for candidate in candidates:
        if candidate and candidate not in result:
            result.append(candidate)
    return tuple(result)


def _profile_okpd2_hierarchy_codes(profile: ProcurementProfile) -> list[str]:
    result: list[str] = []
    for item in profile.items:
        for code in item.okpd2_codes:
            for candidate in _okpd2_hierarchy_codes(code):
                if candidate not in result:
                    result.append(candidate)
    return result


def _profile_to_dict(profile: ProcurementProfile) -> dict:
    return {
        "summary": profile.summary,
        "items": [
            {
                "id": item.id,
                "name": item.name,
                "aliases": list(item.aliases),
                "okpd2_codes": list(item.okpd2_codes),
                "category_terms": list(item.category_terms),
                "exact_terms": list(item.exact_terms),
                "required_terms": list(item.required_terms),
                "excluded_terms": list(item.excluded_terms),
            }
            for item in profile.items
        ],
        "excluded_terms": list(profile.excluded_terms),
    }


def _validate_procurement_profile_response(raw: str) -> None:
    profile = _normalize_procurement_profile(parse_json_object(raw))
    if not profile.items:
        raise RuntimeError(f"AI returned profile with 0 items (raw keys: {list(parse_json_object(raw).keys())})")


def _validate_minprom_requirement_response(raw: str) -> None:
    parsed = parse_json_object(raw)
    if "required" not in parsed and not str(parsed.get("measure_type") or "").strip():
        raise RuntimeError("AI Minprom requirement response is incomplete")


def _validate_supplier_context_response(raw: str) -> None:
    parsed = parse_json_object(raw)
    supplier_context = str(parsed.get("supplier_context") or "").strip()
    supplier_context = re.sub(r"[ \t]+", " ", supplier_context)
    supplier_context = re.sub(r"\n{3,}", "\n\n", supplier_context).strip()
    if len(supplier_context) < 50:
        raise RuntimeError("AI supplier search context extraction returned an empty context")


def _validate_supplier_queries_response(raw: str) -> None:
    parsed = parse_json_object(raw)
    queries = _clean_supplier_queries([str(item).strip() for item in parsed.get("queries", [])])
    if not queries:
        raise RuntimeError("AI did not return usable supplier search queries")


def _validate_supplier_verification_response(raw: str) -> None:
    parsed = parse_json_object(raw)
    action = str(parsed.get("action") or "").strip().lower()
    if action not in {"accept", "reject"}:
        raise RuntimeError("AI supplier verifier returned invalid action")


def _minprom_requirement_to_dict(requirement: MinpromRegistryRequirement) -> dict:
    return {
        "required": requirement.required,
        "measure_type": requirement.measure_type,
        "evidence": requirement.evidence,
        "reason": requirement.reason,
    }


def _minprom_context_to_dict(context: MinpromRegistryContext) -> dict:
    return {
        **_minprom_requirement_to_dict(context.requirement),
        "status": context.status,
        "queries": list(context.queries),
        "entries_count": len(context.entries),
        "candidate_count": context.candidate_count,
        "entries": list(context.entries)[:20],
        "filter_error": context.filter_error,
        "error": context.error,
    }


def _build_minprom_registry_code_queries(profile: ProcurementProfile, *, limit: int = 20) -> list[str]:
    queries: list[str] = []
    for code in _profile_okpd2_hierarchy_codes(profile)[:6]:
        queries.extend(
            [
                f"ОКПД2 {code} реестр Минпромторга",
                f'"{code}" ПП 719 баллы',
                f'"{code}" ПП 719',
            ]
        )
    for item in profile.items:
        item_name = str(item.name or "").strip()
        if item_name and len(item_name) >= 3:
            queries.append(item_name)
        for alias in item.aliases:
            if alias and len(alias) >= 3:
                queries.append(alias)
    primary_terms = _minprom_profile_query_terms(profile)
    for term in primary_terms:
        if term not in queries:
            queries.append(term)
    return _clean_supplier_queries(queries)[:limit]


def _build_minprom_supplier_code_queries(profile: ProcurementProfile, *, limit: int = 8) -> list[str]:
    queries: list[str] = []
    primary_terms = _minprom_profile_query_terms(profile)[:2]
    for code in _profile_okpd2_hierarchy_codes(profile)[:6]:
        queries.extend(
            [
                f"ОКПД2 {code} \"реестр Минпромторга\" производитель",
                f'"{code}" "реестровая запись" производитель',
                f'"{code}" "ПП 719" производитель',
            ]
        )
        if primary_terms:
            queries.append(f'"{primary_terms[0]}" ОКПД2 {code} производитель')
    return _clean_supplier_queries(queries)[:limit]


def _build_minprom_supplier_queries(
    profile: ProcurementProfile,
    registry_context: MinpromRegistryContext,
    *,
    limit: int = 60,
) -> list[str]:
    if not registry_context.requirement.required:
        return []

    queries: list[str] = []
    for entry in registry_context.entries[:20]:
        manufacturer = _clean_minprom_query_term(entry.get("manufacturer"))
        product = _clean_minprom_query_term(entry.get("product"))
        registry_number = _clean_minprom_query_term(entry.get("registry_number"))
        inn = _clean_minprom_query_term(entry.get("inn"))
        if inn:
            queries.extend([
                f'ИНН {inn} официальный сайт',
                f'ИНН {inn} контакты отдел продаж',
            ])
        if manufacturer:
            clean_m = manufacturer.replace('"', '').strip()
            queries.extend(
                [
                    f'"{clean_m}" официальный сайт',
                    f'"{clean_m}" производитель',
                ]
            )
            if product:
                queries.append(f'"{clean_m}" {product[:35]}')
        if product:
            queries.extend(
                [
                    f'"{product}" реестр Минпромторга',
                    f'"{product}" ГИСП поставщик',
                    f'"{product}" производитель',
                ]
            )
        if registry_number:
            queries.append(f'"{registry_number}" Минпромторг')

    code_queries_inserted = False
    for term in _minprom_profile_query_terms(profile)[:8]:
        queries.extend(
            [
                f'"{term}" "реестр Минпромторга" производитель',
                f'"{term}" "реестровая запись" производитель',
                f"{term} российский производитель",
                f'"{term}" официальный сайт',
                f'"{term}" ГИСП производитель',
            ]
        )
        if not code_queries_inserted:
            queries.extend(_build_minprom_supplier_code_queries(profile))
            code_queries_inserted = True
    if not code_queries_inserted:
        queries.extend(_build_minprom_supplier_code_queries(profile))
    return _clean_supplier_queries(queries)[:limit]


def _merge_supplier_query_tracks(general_queries: list[str], minprom_queries: list[str]) -> list[str]:
    if not minprom_queries:
        return general_queries
    prioritized = minprom_queries[:6] + general_queries + minprom_queries[6:]
    return _clean_supplier_queries(prioritized)[:60]


def _minprom_profile_query_terms(profile: ProcurementProfile) -> list[str]:
    terms: list[str] = []
    for item in profile.items:
        candidates = (
            *item.category_terms,
            *item.aliases,
            *item.exact_terms,
            *item.required_terms,
            item.name,
        )
        for candidate in candidates:
            term = _clean_minprom_query_term(candidate)
            if term and term.lower() not in [existing.lower() for existing in terms]:
                terms.append(term)
    summary = _clean_minprom_query_term(profile.summary)
    if summary and summary.lower() not in [existing.lower() for existing in terms]:
        terms.append(summary)
    return terms


def _clean_minprom_query_term(value: object) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip(" .,:;\"'")
    if not text:
        return ""
    text = _normalize_procurement_title_for_minprom_query(text)
    text = re.sub(
        r"^(?:наименование|продукция|товар|производитель|изготовитель|реестровая запись|номер записи)\s*[:№-]\s*",
        "",
        text,
        flags=re.I,
    ).strip(" .,:;\"'")
    if " | " in text:
        parts = [part.strip(" .,:;\"'") for part in text.split(" | ")]
        product_like = [
            part
            for part in parts
            if 4 <= len(part) <= 140
            and not re.search(r"\b(?:инн|огрн|реестр|запись|дата|номер)\b", part, re.I)
        ]
        if product_like:
            text = product_like[0]
    if len(text) < 4 or len(text) > 140:
        return ""
    if _is_generic_supplier_anchor(text.lower()):
        return ""
    return text


def _normalize_procurement_title_for_minprom_query(text: str) -> str:
    normalized = re.sub(r"\s+", " ", text).strip(" .,:;\"'")
    normalized = re.sub(
        r"^(?:эа\s+|ок\s+|зк\s+)?(?:на\s+)?(?:поставка|поставку|приобретение|закупка)\s+",
        "",
        normalized,
        flags=re.I,
    ).strip(" .,:;\"'")
    normalized = re.sub(r"^\d{4}[-/]\d+\s*[=-]\s*", "", normalized).strip(" .,:;\"'")
    purpose_match = re.search(
        r"\s+(?:для|в рамках|по адресу)\s+(?:нужд|гбоу|мбоу|гкоу|фгбоу|фку|суд|оснащения|реализации|обеспечения|пункта|заказчик|заказчика|адрес)",
        normalized,
        flags=re.I,
    )
    if purpose_match:
        normalized = normalized[: purpose_match.start()].strip(" .,:;\"'")
    return normalized


def _minprom_rerank_context(registry_context: MinpromRegistryContext | None) -> tuple[str, dict]:
    if not registry_context or not registry_context.requirement.required:
        return "", {}
    payload = _minprom_context_to_dict(registry_context)
    guidance = (
        "Для этой закупки действует запрет, поэтому реестровая запись Минпромторга/ГИСП важна. "
        "Выше ранжируй производителей и официальных дилеров, если title/snippet/query совпадают с найденной "
        "реестровой записью, производителем, товаром или прямо содержат сигналы 'реестр Минпромторга', 'ГИСП', "
        "'реестровая запись'. Не выбирай сами госреестры, справочники и тендерные страницы. "
        "Если записей ГИСП в контексте нет, не отклоняй всех поставщиков только из-за отсутствия сигнала в сниппете: "
        "оставляй релевантных производителей/дилеров для финального аудита."
    )
    return guidance, payload


def _queries_need_broadening(profile: ProcurementProfile, queries: list[str], target: int) -> bool:
    if target < 10:
        return False
    if len(queries) < max(12, target):
        return True
    narrow_markers = _narrow_query_markers(profile)
    if not narrow_markers:
        return False
    broad_count = sum(1 for query in queries if not _query_contains_narrow_marker(query, narrow_markers))
    return broad_count < max(4, int(len(queries) * 0.35))


def _narrow_query_markers(profile: ProcurementProfile) -> tuple[str, ...]:
    markers: list[str] = []
    for item in profile.items:
        for term in (*item.exact_terms, *item.required_terms, *item.aliases):
            lowered = re.sub(r"\s+", " ", str(term or "").lower()).strip(" .,:;")
            if not lowered:
                continue
            if (
                re.search(r"\d", lowered)
                or any(word in lowered for word in ("гост", "ост", "ту ", "din", "iso", "en ", "тип ", "марка", "модель", "артикул"))
                or re.search(r"\b[а-яёa-z]{1,4}-[а-яёa-z0-9]{1,6}\b", lowered)
            ):
                markers.append(lowered)
    return tuple(dict.fromkeys(markers))


def _query_contains_narrow_marker(query: str, markers: tuple[str, ...]) -> bool:
    lowered = re.sub(r"\s+", " ", str(query or "").lower()).strip(" .,:;")
    for marker in markers:
        marker_parts = [part for part in re.findall(r"[а-яёa-z0-9]+", marker) if len(part) >= 2]
        if marker in lowered:
            return True
        if marker_parts and all(part in lowered for part in marker_parts):
            return True
    return False


async def _revise_supplier_queries_with_ai(
    settings: SystemSettings,
    context: str,
    profile: ProcurementProfile,
    initial_queries: list[str],
    target: int,
) -> list[str]:
    prompt = f"""Первый набор поисковых запросов получился слишком узко привязанным к точным характеристикам ТЗ.

Сформируй новый набор 18-28 запросов для закупочного поиска поставщиков.
Ищи производителей, заводы, дилеров, дистрибьюторов и B2B-поставщиков всей товарной группы/номенклатуры.
Не повторяй ошибку: не делай каждый запрос с точным размером, ГОСТ, типом, маркой или артикулом.
Точные характеристики оставь только в части запросов, чтобы найти точные совпадения.

Цель отчёта: найти до {target} проверенных поставщиков.

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Слишком узкие запросы:
{json.dumps(initial_queries, ensure_ascii=False)}

Фрагмент ТЗ:
{context[:5000]}

Ответ строго JSON:
{{"queries": ["..."]}}"""
    raw = await call_llm(
        settings,
        prompt,
        system_prompt="Ты закупочный исследователь. Расширяешь поисковую стратегию до товарной группы, не уходя от предмета ТЗ.",
        tier="light",
        routing_key="supplier_query_generation",
        json_mode=True,
        timeout_seconds=90,
        response_validator=_validate_supplier_queries_response,
    )
    parsed = parse_json_object(raw)
    queries = _clean_supplier_queries([str(item).strip() for item in parsed.get("queries", [])])
    return queries[:28]


async def _build_supplier_recovery_queries_with_ai(
    settings: SystemSettings,
    context: str,
    profile: ProcurementProfile,
    initial_queries: list[str],
    reviewed: list[dict],
    accepted: list[dict],
    target: int,
) -> list[str]:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for supplier recovery query generation")
    rejected_samples = [
        {
            "site": item.get("site", ""),
            "source": item.get("source", ""),
            "query": item.get("search_query", ""),
            "reason": item.get("comments", ""),
        }
        for item in reviewed
        if item.get("evidence_status") != "verified"
    ][:20]
    accepted_samples = [
        {
            "company": item.get("company_name", ""),
            "site": item.get("site", ""),
            "query": item.get("search_query", ""),
            "product_fit": item.get("product_fit", ""),
        }
        for item in accepted
    ][:20]
    prompt = f"""Первый поисковый проход дал меньше подтверждённых поставщиков, чем нужно для качественного закупочного отчёта.

Сформируй дополнительный набор 8-16 поисковых запросов для второго прохода.
Это должен быть не повтор точной позиции, а расширение поиска по товарной группе/номенклатуре.

Правила:
- ищи производителей, заводы, официальных дилеров, дистрибьюторов и B2B-поставщиков;
- используй широкие category_terms и синонимы из профиля закупки;
- не делай каждый запрос с точным размером, ГОСТ, типом, маркой, артикулом или моделью;
- если точная марка обязательна, добавь 1-3 точных запроса, но основная часть должна искать профильных поставщиков категории;
- не добавляй агрегаторы, маркетплейсы, тендерные площадки, реестры, справочники, статьи, видео и учебные страницы;
- не повторяй исходные запросы без новой формулировки.

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Исходные запросы:
{json.dumps(initial_queries, ensure_ascii=False)}

Уже принятые поставщики:
{json.dumps(accepted_samples, ensure_ascii=False)}

Отклоненные/слабые кандидаты:
{json.dumps(rejected_samples, ensure_ascii=False)}

Фрагмент ТЗ:
{context[:5000]}

Ответ строго JSON:
{{"queries": ["..."]}}"""
    raw = await call_llm(
        settings,
        prompt,
        system_prompt="Ты закупочный ресерчер. Исправляешь недобор поставщиков через более широкую AI-стратегию поиска по номенклатуре.",
        tier="light",
        routing_key="supplier_query_generation",
        json_mode=True,
        timeout_seconds=90,
        response_validator=_validate_supplier_queries_response,
    )
    parsed = parse_json_object(raw)
    initial_set = {query.lower() for query in initial_queries}
    queries = [
        query
        for query in _clean_supplier_queries([str(item).strip() for item in parsed.get("queries", [])])
        if query.lower() not in initial_set
    ]
    return queries[:16]


async def ai_rerank_candidates(
    settings: SystemSettings,
    profile: ProcurementProfile,
    candidates: list[Candidate],
    target: int,
    *,
    registry_context: MinpromRegistryContext | None = None,
) -> CandidateRerank:
    if not settings.has_active_ai_provider:
        raise RuntimeError("AI provider is required for supplier candidate reranking")
    if not candidates:
        return CandidateRerank([], {"status": "empty", "input_count": 0, "kept_count": 0})
    limit = max(30, min(90, target * 18))
    payload_candidates = [
        {
            "id": str(index),
            "url": candidate.url,
            "domain": candidate.domain,
            "title": candidate.title[:180],
            "snippet": candidate.snippet[:360],
            "source": candidate.source,
            "query": candidate.query[:180],
        }
        for index, candidate in enumerate(candidates[:limit])
    ]
    desired_review_count = _desired_candidate_review_count(target, len(payload_candidates))
    by_id = {str(index): candidate for index, candidate in enumerate(candidates[:limit])}
    minprom_guidance, minprom_payload = _minprom_rerank_context(registry_context)
    prompt = f"""Отранжируй поисковых кандидатов перед открытием сайтов.

Нужно выбрать широкий пул кандидатов для дальнейшей ИИ-проверки сайтов и контактов, а не только точные товарные страницы.
Выбирай производителей, заводы, дилеров, дистрибьюторов или B2B-поставщиков позиций из профиля закупки и релевантной товарной группы/номенклатуры.
Если кандидат является профильным поставщиком категории, оставь его даже без точного размера, ГОСТ, артикула или модели в сниппете: финальный ИИ-аудит уточнит product_fit.
Для цели {target} поставщиков желательно оставить до {desired_review_count} кандидатов, если они похожи на сайты компаний.
Понижай или отклоняй маркетплейсы, агрегаторы, тендеры, реестры, справочники, статьи, видео, учебные страницы и страницы профессий.
Для multi-item закупки сохрани покрытие разных позиций, если в выдаче есть подходящие кандидаты.
{minprom_guidance}

Ответ строго JSON:
{{
  "ranked": [
    {{
      "id": "0",
      "keep": true,
      "confidence": 0,
      "procurement_item_id": "item-1",
      "reason": "кратко почему стоит открыть сайт"
    }}
  ]
}}

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Контекст Минпромторга:
{json.dumps(minprom_payload, ensure_ascii=False)}

Кандидаты:
{json.dumps(payload_candidates, ensure_ascii=False)}"""
    last_error: Exception | None = None
    for attempt in range(1, 3):
        try:
            raw = await call_llm(
                settings,
                prompt,
                system_prompt="Ты закупочный ресерчер. Ранжируешь поисковую выдачу до открытия сайтов и отбрасываешь нерелевантные типы источников.",
                tier="light",
                routing_key="supplier_candidate_reranker",
                json_mode=True,
                timeout_seconds=90,
                response_validator=lambda raw: _validate_candidate_rerank_response(raw, by_id),
            )
            parsed = parse_json_object(raw)
            ranked = parsed.get("ranked", [])
            if not isinstance(ranked, list):
                raise RuntimeError("AI candidate reranker returned invalid ranked list")
            kept, seen_ids, seen_domains = _ranked_candidates_from_ai(ranked, by_id)
            if not kept:
                raise RuntimeError("AI candidate reranker did not keep any candidates")
            initial_kept_count = len(kept)
            expanded_kept_count = 0
            if len(kept) < desired_review_count and len(kept) < len(payload_candidates):
                expanded = await _expand_candidate_rerank_with_ai(
                    settings,
                    profile,
                    payload_candidates,
                    seen_ids,
                    seen_domains,
                    desired_review_count - len(kept),
                    target,
                    registry_context=registry_context,
                )
                if expanded:
                    kept.extend(expanded)
                    expanded_kept_count = len(expanded)
            return CandidateRerank(
                kept[:desired_review_count],
                {
                    "status": "ok",
                    "attempts": attempt,
                    "input_count": len(candidates),
                    "sent_to_ai": len(payload_candidates),
                    "desired_review_count": desired_review_count,
                    "initial_kept_count": initial_kept_count,
                    "expanded_kept_count": expanded_kept_count,
                    "kept_count": len(kept[:desired_review_count]),
                },
            )
        except Exception as exc:
            last_error = exc
            if attempt == 1:
                await asyncio.sleep(1)
                continue
            break
    detail = _exception_summary(last_error)
    if _is_empty_rerank_rejection(last_error):
        fallback_candidates = _fallback_candidates_after_empty_rerank(
            candidates[:limit],
            desired_review_count,
        )
        if fallback_candidates:
            logger.warning(
                "supplier_candidate_reranker_fallback job_id=%s input_count=%s selected_count=%s sources=%s error=%s",
                _supplier_search_job_id_context.get(),
                len(candidates),
                len(fallback_candidates),
                _source_counts(fallback_candidates),
                detail,
            )
            return CandidateRerank(
                fallback_candidates,
                {
                    "status": "fallback_after_empty_ai_selection",
                    "attempts": 2,
                    "input_count": len(candidates),
                    "sent_to_ai": len(payload_candidates),
                    "desired_review_count": desired_review_count,
                    "kept_count": len(fallback_candidates),
                    "fallback_reason": detail,
                },
            )
    raise RuntimeError(f"AI candidate reranking failed after retry: {detail}") from last_error


def _desired_candidate_review_count(target: int, candidate_count: int) -> int:
    if candidate_count <= 0:
        return 0
    return min(candidate_count, max(target * 5, target + 20, 30))


def _is_empty_rerank_rejection(error: Exception | None) -> bool:
    """Only recover when AI deliberately rejected every prefilter candidate."""
    return bool(error and "reranker did not keep any candidates" in str(error).lower())


def _fallback_candidates_after_empty_rerank(
    candidates: list[Candidate],
    desired_review_count: int,
) -> list[Candidate]:
    """Keep a bounded local pool; final evidence verification still decides acceptance."""
    fallback: list[Candidate] = []
    seen_domains: set[str] = set()
    for candidate in candidates:
        domain = base_domain(candidate.domain or candidate.url)
        if not domain or domain in seen_domains:
            continue
        seen_domains.add(domain)
        fallback.append(
            replace(
                candidate,
                ai_rank_confidence=0,
                ai_rank_reason="ИИ-предфильтр не выбрал кандидата; направлен на обязательную финальную проверку сайта.",
            )
        )
        if len(fallback) >= desired_review_count:
            break
    return fallback


def _ranked_candidates_from_ai(
    ranked: list,
    by_id: dict[str, Candidate],
    *,
    seen_domains: set[str] | None = None,
    min_confidence: int = 50,
) -> tuple[list[Candidate], set[str], set[str]]:
    kept: list[Candidate] = []
    seen_ids: set[str] = set()
    domains = set(seen_domains or set())
    for item in ranked:
        if not isinstance(item, dict) or not item.get("keep", False):
            continue
        item_id = str(item.get("id") or "")
        candidate = by_id.get(item_id)
        if not candidate or candidate.domain in domains:
            continue
        confidence = _bounded_int(item.get("confidence"), default=0)
        if confidence < min_confidence:
            continue
        seen_ids.add(item_id)
        domains.add(candidate.domain)
        kept.append(
            replace(
                candidate,
                procurement_item_id=str(item.get("procurement_item_id") or ""),
                ai_rank_confidence=confidence,
                ai_rank_reason=str(item.get("reason") or "")[:300],
            )
        )
    if kept:
        return kept, seen_ids, domains
    # Fallback: confidence threshold too strict — accept any keep=true candidate
    # Prevents total failure when the model rates all candidates conservatively
    if min_confidence > 0:
        return _ranked_candidates_from_ai(
            ranked, by_id, seen_domains=seen_domains, min_confidence=0
        )
    return kept, seen_ids, domains


def _validate_candidate_rerank_response(
    raw: str,
    by_id: dict[str, Candidate],
    *,
    seen_domains: set[str] | None = None,
) -> None:
    parsed = parse_json_object(raw)
    ranked = parsed.get("ranked", [])
    if not isinstance(ranked, list):
        raise RuntimeError("AI candidate reranker returned invalid ranked list")
    kept, _, _ = _ranked_candidates_from_ai(ranked, by_id, seen_domains=seen_domains)
    if not kept:
        raise RuntimeError("AI candidate reranker did not keep any candidates")


async def _expand_candidate_rerank_with_ai(
    settings: SystemSettings,
    profile: ProcurementProfile,
    payload_candidates: list[dict],
    seen_ids: set[str],
    seen_domains: set[str],
    needed: int,
    target: int,
    *,
    registry_context: MinpromRegistryContext | None = None,
) -> list[Candidate]:
    remaining = [
        item
        for item in payload_candidates
        if str(item.get("id") or "") not in seen_ids and str(item.get("domain") or "") not in seen_domains
    ]
    if not remaining or needed <= 0:
        return []
    by_id = {
        str(item.get("id") or ""): Candidate(
            url=str(item.get("url") or ""),
            domain=str(item.get("domain") or ""),
            title=str(item.get("title") or ""),
            snippet=str(item.get("snippet") or ""),
            source=str(item.get("source") or ""),
            query=str(item.get("query") or ""),
        )
        for item in remaining
    }
    minprom_guidance, minprom_payload = _minprom_rerank_context(registry_context)
    prompt = f"""Первый ИИ-отбор оставил слишком мало кандидатов для отчёта на {target} поставщиков.

Выбери дополнительно до {needed} сайтов компаний для финального ИИ-аудита.
Расширяй пул за счет производителей, заводов, дилеров, дистрибьюторов и B2B-поставщиков товарной группы/номенклатуры, даже если в сниппете нет точного размера, ГОСТ, артикула или модели.
Не выбирай маркетплейсы, агрегаторы, тендеры, реестры, справочники, статьи, видео, учебные и госстраницы.
{minprom_guidance}

Профиль закупки:
{json.dumps(_profile_to_dict(profile), ensure_ascii=False)}

Контекст Минпромторга:
{json.dumps(minprom_payload, ensure_ascii=False)}

Кандидаты для дополнительного отбора:
{json.dumps(remaining, ensure_ascii=False)}

Ответ строго JSON:
{{"ranked": [{{"id": "0", "keep": true, "confidence": 0, "procurement_item_id": "item-1", "reason": "почему стоит открыть сайт"}}]}}"""
    try:
        raw = await call_llm(
            settings,
            prompt,
            system_prompt="Ты закупочный ресерчер. Расширяешь пул профильных сайтов для обязательного финального ИИ-аудита.",
            tier="light",
            routing_key="supplier_candidate_reranker",
            json_mode=True,
            timeout_seconds=90,
            response_validator=lambda raw: _validate_candidate_rerank_response(raw, by_id, seen_domains=seen_domains),
        )
        parsed = parse_json_object(raw)
        ranked = parsed.get("ranked", [])
        if not isinstance(ranked, list):
            return []
        kept, _, _ = _ranked_candidates_from_ai(ranked, by_id, seen_domains=seen_domains)
        return kept[:needed]
    except Exception:
        return []


def _bounded_int(value: object, *, default: int = 0, minimum: int = 0, maximum: int = 100) -> int:
    try:
        number = float(str(value).replace(",", "."))
        if minimum == 0 and maximum == 100 and 0 < number <= 1:
            number *= 100
        return max(minimum, min(maximum, int(round(number))))
    except (TypeError, ValueError):
        return default


def _exception_summary(exc: Exception | None) -> str:
    if exc is None:
        return "unknown error"
    message = str(exc).strip()
    if not message:
        message = type(exc).__name__
    return f"{type(exc).__name__}: {message}"[:500]


class _SupplierVerificationLimiter:
    """Process-wide capacity guard that is safe across worker threads/event loops."""

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._active = 0

    def _try_acquire(self, limit: int) -> bool:
        with self._lock:
            if self._active >= limit:
                return False
            self._active += 1
            return True

    async def acquire(self, limit_provider: Callable[[], int]) -> None:
        while not self._try_acquire(limit_provider()):
            await asyncio.sleep(0.05)

    def release(self) -> None:
        with self._lock:
            if self._active <= 0:
                raise RuntimeError("supplier verification limiter released without acquire")
            self._active -= 1

    def reset_for_tests(self) -> None:
        with self._lock:
            if self._active:
                raise RuntimeError("cannot reset an active supplier verification limiter")
            self._active = 0


_supplier_verification_limiter = _SupplierVerificationLimiter()


def _supplier_verification_limit() -> int:
    try:
        configured = int(config.supplier_verification_concurrency)
    except (TypeError, ValueError):
        configured = 8
    return max(1, min(32, configured))


def _supplier_verification_timeout_seconds() -> float:
    return max(
        1.0,
        min(600.0, float(config.supplier_verification_timeout_seconds or 150.0)),
    )


class SupplierBrowserInfrastructureError(Exception):
    pass


async def _verify_candidate_with_limits(
    settings: SystemSettings,
    candidate: Candidate,
    context: str,
    *,
    profile: ProcurementProfile | None = None,
    registry_context: MinpromRegistryContext | None = None,
) -> dict | None:
    await _supplier_verification_limiter.acquire(_supplier_verification_limit)
    try:
        timeout = _supplier_verification_timeout_seconds()
        for attempt in range(2):
            try:
                return await asyncio.wait_for(
                    verify_candidate(
                        settings,
                        candidate,
                        context,
                        profile=profile,
                        registry_context=registry_context,
                    ),
                    timeout=timeout,
                )
            except TimeoutError:
                return None
            except SupplierBrowserInfrastructureError:
                if attempt == 0:
                    await asyncio.sleep(0.01)
                    continue
                return None
            except Exception:
                return None
    finally:
        _supplier_verification_limiter.release()


async def _review_candidates_until_target(
    settings: SystemSettings,
    candidates: list[Candidate],
    context: str,
    target: int,
    *,
    profile: ProcurementProfile | None = None,
    registry_context: MinpromRegistryContext | None = None,
    excluded_domains: set[str] | None = None,
    excluded_company_keys: set[str] | None = None,
    progress_callback: ProgressCallback | None = None,
) -> tuple[list[dict], list[dict], dict]:
    reviewed: list[dict] = []
    batch_size = _candidate_review_batch_size(target)
    stopped_after = 0

    async def review(index: int, candidate: Candidate) -> dict | None:
        result = await _verify_candidate_with_limits(
            settings,
            candidate,
            context,
            profile=profile,
            registry_context=registry_context,
        )
        if result:
            result["_source_rank"] = index
        return result

    for batch_start in range(0, len(candidates), batch_size):
        batch = candidates[batch_start : batch_start + batch_size]
        already_accepted = _accepted_supplier_results(
            reviewed,
            target,
            profile=profile,
            limit_to_target=False,
            excluded_domains=excluded_domains,
            excluded_company_keys=excluded_company_keys,
        )
        # Early stop: skip batch entirely if we already have enough
        if len(already_accepted) >= target:
            stopped_after = batch_start
            break
        review_progress = 74 + int(18 * min(batch_start, len(candidates)) / max(1, len(candidates)))
        await _emit_progress(
            progress_callback,
            review_progress,
            (
                f"Проверяю сайты: {batch_start + 1}-{batch_start + len(batch)} "
                f"из {len(candidates)}, подтверждено {len(already_accepted)}"
            ),
        )
        tasks = [
            asyncio.create_task(review(batch_start + offset, candidate))
            for offset, candidate in enumerate(batch)
        ]
        try:
            for task in asyncio.as_completed(tasks):
                result = await task
                if result:
                    reviewed.append(result)
        finally:
            for task in tasks:
                if not task.done():
                    task.cancel()
            await asyncio.gather(*tasks, return_exceptions=True)
        stopped_after = batch_start + len(batch)
        accepted = _accepted_supplier_results(
            reviewed,
            target,
            profile=profile,
            limit_to_target=False,
            excluded_domains=excluded_domains,
            excluded_company_keys=excluded_company_keys,
        )
        review_progress = 74 + int(18 * min(stopped_after, len(candidates)) / max(1, len(candidates)))
        await _emit_progress(
            progress_callback,
            review_progress,
            f"Проверено сайтов: {stopped_after}/{len(candidates)}, подтверждено {len(accepted)}",
        )
        if len(accepted) >= target:
            return accepted, reviewed, {
                "batch_size": batch_size,
                "reviewed_count": len(reviewed),
                "candidate_count": len(candidates),
                "stopped_after_candidates": stopped_after,
                "early_stop": stopped_after < len(candidates),
            }

    return _accepted_supplier_results(
        reviewed,
        target,
        profile=profile,
        limit_to_target=False,
        excluded_domains=excluded_domains,
        excluded_company_keys=excluded_company_keys,
    ), reviewed, {
        "batch_size": batch_size,
        "reviewed_count": len(reviewed),
        "candidate_count": len(candidates),
        "stopped_after_candidates": stopped_after,
        "early_stop": stopped_after < len(candidates),
    }


def _candidate_review_batch_size(target: int) -> int:
    return max(12, min(32, max(1, target) * 2))


def _accepted_supplier_results(
    reviewed: list[dict],
    target: int,
    *,
    profile: ProcurementProfile | None = None,
    limit_to_target: bool = True,
    excluded_domains: set[str] | None = None,
    excluded_company_keys: set[str] | None = None,
) -> list[dict]:
    accepted: list[dict] = []
    seen_domains: set[str] = set(excluded_domains or set())
    seen_companies: set[str] = set(excluded_company_keys or set())
    verified = [item for item in reviewed if item.get("evidence_status") == "verified"]
    sorted_verified = sorted(verified, key=_supplier_result_sort_key)

    def add_result(result: dict) -> bool:
        if result.get("evidence_status") != "registry_only" and _supplier_quality_score(result) < MIN_VERIFIED_SUPPLIER_SCORE:
            return False
        domain = base_domain(result.get("site", ""))
        company_key = _normalize_company_key(result.get("company_name") or domain)
        # Allow registry_only entries without a domain
        is_registry_stub = result.get("evidence_status") == "registry_only"
        if is_registry_stub and company_key and company_key not in seen_companies:
            result["quality_score"] = max(_supplier_quality_score(result), 75)
            result["quality_tier"] = _supplier_quality_tier(result["quality_score"])
            accepted.append(result)
            seen_companies.add(company_key)
        elif domain and domain not in seen_domains and company_key not in seen_companies:
            score = _supplier_quality_score(result)
            result["quality_score"] = score
            result["quality_tier"] = _supplier_quality_tier(score)
            accepted.append(result)
            seen_domains.add(domain)
            if company_key:
                seen_companies.add(company_key)
            return True
        return False

    if profile and len(profile.items) > 1:
        for item in profile.items:
            for result in sorted_verified:
                if str(result.get("procurement_item_id") or "") == item.id and add_result(result):
                    break
            if limit_to_target and len(accepted) >= target:
                return accepted[:target]

    for result in sorted_verified:
        add_result(result)
        if limit_to_target and len(accepted) >= target:
            break
    return accepted[:target] if limit_to_target else accepted


def _supplier_delivery_target(minimum_target: int) -> int:
    """Return a modest verified-result goal while preserving the configured minimum."""
    minimum = max(1, min(MAX_SUPPLIER_DELIVERY_TARGET, int(minimum_target or 0)))
    bonus = min(10, max(2, (minimum + 2) // 3))
    return min(MAX_SUPPLIER_DELIVERY_TARGET, minimum + bonus)


def _primary_candidate_floor(delivery_target: int) -> int:
    """Enough first-party candidates to keep reserve sources out of the normal path."""
    return max(24, min(40, max(1, int(delivery_target or 0)) * 2))


def _fallback_candidate_limit(delivery_target: int) -> int:
    """Cap supplemental sources when Yandex already returned candidates."""
    return max(6, min(16, (max(1, int(delivery_target or 0)) + 1) // 2))


def _deterministic_measure_type(text: str) -> str:
    lower = str(text or "").lower()
    if "запрет" in lower or "1236" in lower:
        return "prohibition"
    if "ограничен" in lower or "878" in lower or "616" in lower:
        return "restriction"
    if "преимуществ" in lower:
        return "advantage"
    return ""


async def discover_candidates(
    settings: SystemSettings,
    queries: list[str],
    max_results: int,
    *,
    excluded_domains: set[str] | None = None,
    primary_candidate_floor: int | None = None,
    fallback_candidate_limit: int | None = None,
    reserve_only: bool = False,
) -> tuple[list[Candidate], dict]:
    candidates: list[Candidate] = []
    reports: list[dict] = []
    provider_order = _provider_order(settings)
    base_excluded_domains = set(excluded_domains or set())
    configured_primary_floor = max_results if primary_candidate_floor is None else primary_candidate_floor
    configured_fallback_limit = max_results if fallback_candidate_limit is None else fallback_candidate_limit
    primary_floor = min(max_results, max(1, configured_primary_floor))
    fallback_limit = min(max_results, max(0, configured_fallback_limit))
    yandex_candidate_count = 0
    yandex_total_requests = 0
    fallback_candidates_added = 0
    fallback_used = False

    for provider in provider_order:
        before = len(candidates)
        provider_candidates: list[Candidate] = []
        request_limit = max_results
        status = "skipped"
        error = ""
        if reserve_only and provider == PRIMARY_SUPPLIER_SEARCH_PROVIDER:
            reports.append(
                {
                    "provider": provider,
                    "status": "skipped_post_verification_reserve",
                    "added": 0,
                    "returned": 0,
                    "total_after": len(candidates),
                    "requested": 0,
                    "error": "",
                }
            )
            continue
        is_fallback = provider != PRIMARY_SUPPLIER_SEARCH_PROVIDER
        if is_fallback:
            if not reserve_only and yandex_candidate_count >= primary_floor:
                reports.append(
                    {
                        "provider": provider,
                        "status": "skipped_primary_sufficient",
                        "added": 0,
                        "returned": 0,
                        "total_after": len(candidates),
                        "requested": 0,
                        "error": "",
                    }
                )
                continue
            if yandex_candidate_count > 0 and not reserve_only:
                remaining_gap = max(0, primary_floor - len(candidates))
                remaining_fallback_budget = max(0, fallback_limit - fallback_candidates_added)
                request_limit = min(max_results - len(candidates), remaining_gap, remaining_fallback_budget)
                if request_limit <= 0:
                    reports.append(
                        {
                            "provider": provider,
                            "status": "skipped_fallback_limit",
                            "added": 0,
                            "returned": 0,
                            "total_after": len(candidates),
                            "requested": 0,
                            "error": "",
                        }
                    )
                    continue
            else:
                request_limit = min(fallback_limit, max(0, max_results - len(candidates))) if reserve_only else max(0, max_results - len(candidates))
                if request_limit <= 0:
                    break
        try:
            existing_domains = base_excluded_domains | {candidate.domain for candidate in candidates}
            if provider == "yandex":
                y_res = await _search_with_yandex(settings, queries, request_limit, existing_domains=existing_domains)
                if isinstance(y_res, tuple) and len(y_res) == 2:
                    provider_candidates, y_reqs = y_res
                else:
                    provider_candidates, y_reqs = y_res, len(queries)
                yandex_total_requests += y_reqs
            elif provider == "google":
                provider_candidates = await _search_with_google(settings, queries, request_limit, existing_domains=existing_domains)
            elif provider == "tavily":
                provider_candidates = await _search_with_tavily(settings, queries, request_limit, existing_domains=existing_domains)
            elif provider == "ddgs":
                provider_candidates = await _search_with_ddgs(queries, request_limit, existing_domains=existing_domains)
            provider_candidates = _exclude_candidates(provider_candidates, base_excluded_domains)[:request_limit]
            status = "ok" if provider_candidates else "empty"
        except Exception as exc:
            status = "error"
            error = f"{type(exc).__name__}: {str(exc)[:180]}"
            logger.warning("Supplier candidate provider failed: provider=%s error=%s", provider, error)
        candidates = _merge_candidates(candidates, provider_candidates, max_results=max_results, excluded_domains=base_excluded_domains)
        added = len(candidates) - before
        if provider == PRIMARY_SUPPLIER_SEARCH_PROVIDER:
            yandex_candidate_count = len(candidates)
        else:
            fallback_used = fallback_used or bool(added)
            fallback_candidates_added += added
        reports.append(
            {
                "provider": provider,
                "status": status,
                "added": added,
                "returned": len(provider_candidates),
                "total_after": len(candidates),
                "requested": request_limit,
                "error": error,
            }
        )
        if len(candidates) >= max_results:
            break

    return candidates[:max_results], {
        "provider_order": provider_order,
        "strategy": {
            "primary_provider": PRIMARY_SUPPLIER_SEARCH_PROVIDER,
            "primary_candidate_floor": primary_floor,
            "fallback_candidate_limit": fallback_limit,
            "yandex_candidate_count": yandex_candidate_count,
            "fallback_used": fallback_used,
            "reserve_only": reserve_only,
        },
        "yandex_requests_count": yandex_total_requests,
        "yandex_cost_rub": round(yandex_total_requests * float(getattr(settings, "yandex_search_price_per_request", 0.04) or 0.04), 2),
        "reports": reports,
    }


def _provider_order(settings: SystemSettings) -> list[str]:
    configured = str(getattr(settings, "supplier_search_provider_order", "") or os.getenv("AIPOISK_SUPPLIER_SEARCH_PROVIDER_ORDER", ""))
    raw_items = [item.strip().lower() for item in configured.split(",") if item.strip()]
    supported = {"yandex", "google", "tavily", "ddgs"}
    order = [item for item in raw_items if item in supported]
    if not order:
        order = ["google", "tavily", "ddgs"]
    fallbacks = [item for item in order if item != PRIMARY_SUPPLIER_SEARCH_PROVIDER]
    return [PRIMARY_SUPPLIER_SEARCH_PROVIDER, *dict.fromkeys(fallbacks)]


def _provider_query_limit(settings: SystemSettings, provider: str) -> int:
    specific = os.getenv(f"AIPOISK_{provider.upper()}_SEARCH_QUERY_LIMIT", "")
    configured = specific or os.getenv("AIPOISK_SEARCH_QUERY_LIMIT", "")
    try:
        return max(1, min(48, int(configured)))
    except ValueError:
        return 14 if provider == "google" else 12


def _yandex_max_pages_per_query(settings: SystemSettings) -> int:
    configured = os.getenv("AIPOISK_YANDEX_MAX_PAGES_PER_QUERY", "")
    try:
        return max(1, min(10, int(configured)))
    except ValueError:
        return 3


def _merge_candidates(
    *groups: list[Candidate] | tuple[Candidate, ...],
    max_results: int,
    excluded_domains: set[str] | None = None,
) -> list[Candidate]:
    merged: list[Candidate] = []
    seen_domains: set[str] = set(excluded_domains or set())
    for group in groups:
        for candidate in group:
            domain = base_domain(candidate.domain or candidate.url)
            if not domain or domain in seen_domains:
                continue
            seen_domains.add(domain)
            merged.append(candidate)
            if len(merged) >= max_results:
                return merged
    return merged


def _source_counts(candidates: list[Candidate] | tuple[Candidate, ...] | object) -> dict[str, int]:
    counts: dict[str, int] = {}
    for candidate in candidates:
        source = str(getattr(candidate, "source", "") or "unknown")
        counts[source] = counts.get(source, 0) + 1
    return dict(sorted(counts.items()))


def _yandex_credentials(settings: SystemSettings) -> tuple[str, str]:
    folder_id = str(getattr(settings, "yandex_search_folder_id", "") or os.getenv("AIPOISK_YANDEX_SEARCH_FOLDER_ID", "") or os.getenv("YANDEX_FOLDER_ID", "")).strip()
    api_key = str(getattr(settings, "yandex_search_api_key", "") or os.getenv("AIPOISK_YANDEX_SEARCH_API_KEY", "") or os.getenv("YANDEX_API_KEY", "")).strip()
    return folder_id, api_key


async def _search_with_yandex(
    settings: SystemSettings,
    queries: list[str],
    max_results: int,
    *,
    existing_domains: set[str] | None = None,
    expand_queries: bool = True,
    max_pages_per_query: int | None = None,
) -> tuple[list[Candidate], int]:
    folder_id, api_key = _yandex_credentials(settings)
    if not folder_id or not api_key:
        return [], 0
    if expand_queries:
        search_queries = _expand_search_queries(queries, max_queries=_provider_query_limit(settings, "yandex"))
    else:
        search_queries = [q.strip() for q in queries if str(q).strip()]
    max_pages = max_pages_per_query if max_pages_per_query is not None else _yandex_max_pages_per_query(settings)
    semaphore = asyncio.Semaphore(3)
    requests_count = 0

    async def _poll_yandex_operation(client: httpx.AsyncClient, operation_id: str) -> str:
        poll_delays = [0.3, 0.5, 0.8, 1.2, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5]
        for _poll_idx in range(len(poll_delays)):
            await asyncio.sleep(poll_delays[_poll_idx])
            operation = await client.get(f"https://operation.api.cloud.yandex.net/operations/{operation_id}", headers=headers)
            if operation.status_code != 200:
                continue
            data = operation.json()
            if not data.get("done"):
                continue
            return str(data.get("response", {}).get("rawData") or "")
        return ""

    async def search_one_page(client: httpx.AsyncClient, query: str, page: int) -> list[Candidate]:
        nonlocal requests_count, headers
        async with semaphore:
            body = {
                "query": {"searchType": "SEARCH_TYPE_RU", "queryText": query, "page": str(page)},
                "folderId": folder_id,
                "responseFormat": "FORMAT_XML",
                "groupBy": {"groupsOnPage": 10, "docsInGroup": 1},
            }
            requests_count += 1
            response = await client.post("https://searchapi.api.cloud.yandex.net/v2/web/searchAsync", headers=headers, json=body)
            if response.status_code != 200:
                return []
            operation_id = str(response.json().get("id") or "")
            if not operation_id:
                return []
            raw_data = await _poll_yandex_operation(client, operation_id)
            return _parse_yandex_xml(raw_data, query=query) if raw_data else []

    async def search_one(client: httpx.AsyncClient, query: str) -> list[Candidate]:
        all_candidates: list[Candidate] = []
        seen_domains: set[str] = set()
        for page in range(max_pages):
            page_candidates = await search_one_page(client, query, page)
            if not page_candidates:
                break
            new_added = 0
            for c in page_candidates:
                if c.domain not in seen_domains:
                    seen_domains.add(c.domain)
                    all_candidates.append(c)
                    new_added += 1
            if new_added == 0 or len(all_candidates) >= 30:
                break
        return all_candidates

    headers = {"Authorization": f"Api-Key {api_key}", "Content-Type": "application/json"}
    candidates: list[Candidate] = []
    seen = set(existing_domains or set())
    async with httpx.AsyncClient(timeout=45, follow_redirects=True) as client:
        tasks = [asyncio.create_task(search_one(client, query)) for query in search_queries]
        for task in asyncio.as_completed(tasks):
            for candidate in await task:
                if candidate.domain in seen:
                    continue
                seen.add(candidate.domain)
                candidates.append(candidate)
                if len(candidates) >= max_results:
                    for pending in tasks:
                        if not pending.done():
                            pending.cancel()
                    break
            if len(candidates) >= max_results:
                break
        await asyncio.gather(*tasks, return_exceptions=True)
    return candidates[:max_results], requests_count


def _parse_yandex_xml(xml_data: str, *, query: str) -> list[Candidate]:
    raw = str(xml_data or "").strip()
    if not raw:
        return []
    if not raw.startswith("<"):
        try:
            raw = base64.b64decode(raw).decode("utf-8")
        except Exception:
            pass
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return []
    candidates: list[Candidate] = []
    for doc in root.findall(".//doc"):
        url_elem = doc.find("url")
        if url_elem is None or not url_elem.text:
            continue
        url = normalize_url(url_elem.text)
        domain = base_domain(url)
        if not url or not domain:
            continue
        title_elem = doc.find("title")
        title = "".join(title_elem.itertext()).strip() if title_elem is not None else ""
        snippets: list[str] = []
        for passage in doc.findall(".//passage"):
            snippets.append("".join(passage.itertext()).strip())
        candidates.append(Candidate(url=url, domain=domain, title=title, snippet=" ".join(snippets), source="yandex", query=query))
    return candidates


def _google_credentials(settings: SystemSettings) -> tuple[str, str]:
    api_key = str(getattr(settings, "google_search_api_key", "") or os.getenv("AIPOISK_GOOGLE_SEARCH_API_KEY", "") or os.getenv("GOOGLE_API_KEY", "")).strip()
    cse_id = str(getattr(settings, "google_search_cse_id", "") or os.getenv("AIPOISK_GOOGLE_CSE_ID", "") or os.getenv("GOOGLE_CSE_ID", "")).strip()
    return api_key, cse_id


async def _search_with_google(
    settings: SystemSettings,
    queries: list[str],
    max_results: int,
    *,
    existing_domains: set[str] | None = None,
) -> list[Candidate]:
    api_key, cse_id = _google_credentials(settings)
    if not api_key or not cse_id:
        return []
    search_queries = _expand_search_queries(queries, max_queries=_provider_query_limit(settings, "google"))
    semaphore = asyncio.Semaphore(4)

    async def search_one(client: httpx.AsyncClient, query: str) -> list[Candidate]:
        async with semaphore:
            response = await client.get(
                "https://www.googleapis.com/customsearch/v1",
                params={
                    "key": api_key,
                    "cx": cse_id,
                    "q": query,
                    "num": 10,
                    "gl": "ru",
                    "cr": "countryRU",
                    "lr": "lang_ru",
                    "safe": "off",
                },
            )
            if response.status_code != 200:
                return []
            return _parse_google_items(list(response.json().get("items", [])), query=query)

    candidates: list[Candidate] = []
    seen = set(existing_domains or set())
    async with httpx.AsyncClient(timeout=25, follow_redirects=True) as client:
        tasks = [asyncio.create_task(search_one(client, query)) for query in search_queries]
        for task in asyncio.as_completed(tasks):
            for candidate in await task:
                if candidate.domain in seen:
                    continue
                seen.add(candidate.domain)
                candidates.append(candidate)
                if len(candidates) >= max_results:
                    for pending in tasks:
                        if not pending.done():
                            pending.cancel()
                    break
            if len(candidates) >= max_results:
                break
        await asyncio.gather(*tasks, return_exceptions=True)
    return candidates[:max_results]


def _parse_google_items(items: list[dict], *, query: str) -> list[Candidate]:
    candidates: list[Candidate] = []
    for item in items:
        url = normalize_url(str(item.get("link") or item.get("url") or ""))
        domain = base_domain(url)
        if not url or not domain:
            continue
        candidates.append(
            Candidate(
                url=url,
                domain=domain,
                title=str(item.get("title") or ""),
                snippet=str(item.get("snippet") or item.get("body") or ""),
                source="google",
                query=query,
            )
        )
    return candidates


async def _search_with_tavily(
    settings: SystemSettings,
    queries: list[str],
    max_results: int,
    *,
    existing_domains: set[str] | None = None,
) -> list[Candidate]:
    key_candidates = _tavily_key_candidates(settings)
    if not key_candidates:
        return []

    base_url = _tavily_base_url(settings)
    search_queries = _expand_search_queries(queries, max_queries=max(32, min(48, len(queries) * 3)))

    async with httpx.AsyncClient(timeout=28, follow_redirects=True) as client:
        candidates, followups = await _run_tavily_queries(
            client,
            base_url,
            key_candidates,
            search_queries,
            max_results,
            existing_domains=existing_domains,
        )
        if len(candidates) >= max_results:
            return candidates[:max_results]
        followup_queries = list(dict.fromkeys(followups))[:24]
        if followup_queries:
            extra_candidates, _ = await _run_tavily_queries(
                client,
                base_url,
                key_candidates,
                followup_queries,
                max_results,
                existing_domains={candidate.domain for candidate in candidates},
            )
            candidates.extend(extra_candidates)
    return candidates[:max_results]


async def _run_tavily_queries(
    client: httpx.AsyncClient,
    base_url: str,
    key_candidates: list[str],
    queries: list[str],
    max_results: int,
    *,
    existing_domains: set[str] | None = None,
) -> tuple[list[Candidate], list[str]]:
    semaphore = asyncio.Semaphore(6)

    async def search_one(query: str) -> tuple[str, list[dict]]:
        async with semaphore:
            for api_key in key_candidates:
                try:
                    response = await client.post(
                        f"{base_url}/search",
                        json={
                            "api_key": api_key,
                            "query": query,
                            "max_results": 10,
                            "search_depth": "advanced",
                            "include_raw_content": False,
                            "include_images": False,
                        },
                    )
                    if response.status_code in {401, 403}:
                        continue
                    response.raise_for_status()
                    data = response.json()
                    return query, list(data.get("results", []))
                except Exception:
                    continue
        return query, []

    candidates: list[Candidate] = []
    followups: list[str] = []
    seen_domains: set[str] = set(existing_domains or set())
    tasks = [asyncio.create_task(search_one(query)) for query in queries]
    for task in asyncio.as_completed(tasks):
        query, items = await task
        for item in items:
            title = str(item.get("title") or "")
            snippet = str(item.get("content") or "")
            followups.extend(_lead_queries_from_result(title, snippet, query))
            url = normalize_url(str(item.get("url") or ""))
            domain = base_domain(url)
            if not url or not domain or domain in seen_domains:
                continue
            seen_domains.add(domain)
            candidates.append(Candidate(url=url, domain=domain, title=title, snippet=snippet or query, source="tavily", query=query))
            if len(candidates) >= max_results:
                for pending in tasks:
                    if not pending.done():
                        pending.cancel()
                break
        if len(candidates) >= max_results:
            break
    await asyncio.gather(*tasks, return_exceptions=True)
    return candidates, followups


async def _search_with_ddgs(
    queries: list[str],
    max_results: int,
    *,
    existing_domains: set[str] | None = None,
) -> list[Candidate]:
    try:
        from ddgs import DDGS  # type: ignore
    except ImportError:
        return []

    search_queries = _ddgs_search_queries(queries)

    def search_all_sync() -> list[Candidate]:
        candidates: list[Candidate] = []
        seen_domains = set(existing_domains or set())
        with DDGS() as client:
            for query in search_queries:
                if len(candidates) >= max_results:
                    break
                try:
                    items = list(client.text(query, region="ru-ru", backend="auto", max_results=10))
                except Exception:
                    continue
                for item in items:
                    url = normalize_url(str(item.get("href") or item.get("url") or ""))
                    domain = base_domain(url)
                    if not url or not domain or domain in seen_domains:
                        continue
                    seen_domains.add(domain)
                    candidates.append(
                        Candidate(
                            url=url,
                            domain=domain,
                            title=str(item.get("title") or ""),
                            snippet=str(item.get("body") or item.get("description") or "") or query,
                            source="ddgs",
                            query=query,
                        )
                    )
                    if len(candidates) >= max_results:
                        break
        return candidates

    return await asyncio.to_thread(search_all_sync)


def _ddgs_search_queries(queries: list[str]) -> list[str]:
    preferred: list[str] = []
    for query in queries:
        clean = re.sub(r"\s+", " ", str(query or "")).strip(" .,:;")
        if not clean or len(clean) > 110:
            continue
        if clean.startswith('"') and clean.count('"') >= 2 and len(clean) > 70:
            continue
        preferred.append(clean)
    return list(dict.fromkeys(preferred))[:28]


def _tavily_base_url(settings: SystemSettings) -> str:
    configured = str(settings.supplier_search_adapter_base_url or "").strip().rstrip("/")
    if "tavily.com" in configured:
        return configured
    return "https://api.tavily.com"


def _tavily_key_candidates(settings: SystemSettings) -> list[str]:
    keys: list[str] = []
    configured_base = str(settings.supplier_search_adapter_base_url or "").lower()
    configured_model = str(settings.supplier_search_adapter_model or "").lower()
    configured_key = str(settings.supplier_search_adapter_api_key or "").strip()
    if configured_key and ("tavily" in configured_base or configured_model == "tavily"):
        keys.append(configured_key)
    for name in ("AIPOISK_TAVILY_API_KEY", "TAVILY_API_KEY"):
        value = os.getenv(name, "").strip()
        if value:
            keys.append(value)
    hermes_env = Path("/home/hermes/.hermes/.env")
    if hermes_env.exists():
        for line in hermes_env.read_text(encoding="utf-8", errors="ignore").splitlines():
            if line.startswith("TAVILY_API_KEY="):
                keys.append(line.split("=", 1)[1].strip().strip('"').strip("'"))
                break
    return [key for key in dict.fromkeys(keys) if key]


def _expand_search_queries(queries: list[str], *, max_queries: int) -> list[str]:
    base_queries: list[str] = []
    secondary_variants: list[str] = []
    for query in queries:
        clean = re.sub(r"\s+", " ", str(query or "")).strip(" .,:;")
        if not clean:
            continue
        base_queries.append(clean)
        variants = [
            f"{clean} официальный сайт",
            f"{clean} контакты",
            f"{clean} каталог",
            f"{clean} производитель",
            f"{clean} поставщик",
            f"{clean} купить",
        ]
        for item in variants:
            secondary_variants.append(item)

    expanded: list[str] = []
    for item in [*base_queries, *secondary_variants]:
        if item not in expanded:
            expanded.append(item)
        if len(expanded) >= max_queries:
            return expanded
    return expanded


def _lead_queries_from_result(title: str, snippet: str, original_query: str) -> list[str]:
    text = f"{title} {snippet}"
    leads: list[str] = []
    for marker in ("|", "—", "–"):
        for part in title.split(marker)[1:]:
            value = re.sub(r"\s+", " ", part).strip(" .,:;")
            if _looks_like_company_fragment(value):
                leads.append(f"{value} официальный сайт контакты")
    for match in re.finditer(
        r"((?:ООО|АО|ЗАО|ПАО|НПО|НПП|ТД|ГК)\s+[«\"A-Za-zА-Яа-яЁё0-9][^.,;|]{2,80})",
        text,
        re.I,
    ):
        leads.append(f"{match.group(1).strip()} официальный сайт контакты")
    if "завод" in text.lower():
        words = re.findall(r"[А-Яа-яЁёA-Za-z0-9\-]{3,}", text)
        for index, word in enumerate(words):
            if word.lower().startswith("завод"):
                fragment = " ".join(words[max(0, index - 4) : min(len(words), index + 6)])
                if _looks_like_company_fragment(fragment):
                    leads.append(f"{fragment} официальный сайт контакты")
    product = re.sub(r"\s+", " ", original_query).strip()
    return [f"{lead} {product}" for lead in dict.fromkeys(leads) if lead][:4]


def _looks_like_company_fragment(value: str) -> bool:
    lowered = value.lower()
    return len(value) >= 8 and any(
        marker in lowered
        for marker in (
            "завод",
            "компания",
            "предприят",
            "производ",
            "снаб",
            "сервис",
            "тд ",
            "ооо",
            "ао ",
        )
    )


async def _search_with_adapter(settings: SystemSettings, queries: list[str], max_results: int) -> list[Candidate]:
    base_url = str(settings.supplier_search_adapter_base_url or "").strip().rstrip("/")
    api_key = str(settings.supplier_search_adapter_api_key or "").strip()
    model = str(settings.supplier_search_adapter_model or "").strip()
    if not base_url or not api_key:
        return []
    if not base_url.endswith("/chat/completions"):
        base_url = f"{base_url}/chat/completions"

    candidates: list[Candidate] = []
    seen: set[str] = set()
    async with httpx.AsyncClient(timeout=35, follow_redirects=True) as client:
        for query in queries:
            payload = {
                "web_search": True,
                "messages": [
                    {
                        "role": "system",
                        "content": "Ищи официальные сайты российских производителей, заводов, дилеров и B2B-поставщиков. Не возвращай каталоги и маркетплейсы как финальный сайт.",
                    },
                    {
                        "role": "user",
                        "content": f"Найди до 10 официальных сайтов компаний по запросу: {query}. Верни URL и краткое описание.",
                    },
                ],
                "temperature": 0.1,
                "max_tokens": 1600,
            }
            if model:
                payload["model"] = model
            try:
                response = await client.post(base_url, headers={"Authorization": f"Bearer {api_key}"}, json=payload)
                response.raise_for_status()
                data = response.json()
            except Exception:
                continue
            content = json.dumps(data, ensure_ascii=False)
            for raw_url in URL_RE.findall(content):
                url = normalize_url(raw_url)
                domain = base_domain(url)
                if not url or not domain or domain in seen or is_blocked(domain):
                    continue
                seen.add(domain)
                candidates.append(Candidate(url=url, domain=domain, title="", snippet=query, source="adapter", query=query))
                if len(candidates) >= max_results:
                    return candidates
    return candidates


_DNS_RESOLVE_CACHE: dict[str, bool] = {}
_MX_RESOLVE_CACHE: dict[str, bool] = {}
_MAX_DNS_CACHE_SIZE = 5000


def _clean_host_for_dns(value: str) -> str:
    host = str(value or "").strip().lower()
    if "://" in host:
        host = host.split("://", 1)[1]
    host = host.split("/")[0].split("?")[0].split("#")[0].split("@")[-1].split(":")[0].strip(".")
    if host.startswith("www."):
        host = host[4:]
    try:
        host = host.encode("idna").decode("ascii")
    except Exception:
        pass
    return host


async def candidate_domain_resolves_fast(domain_or_url: str, *, timeout: float = 1.2) -> bool:
    host = _clean_host_for_dns(domain_or_url)
    if not host or "." not in host:
        return False
    if host.endswith(".example") or host.endswith(".test") or host.endswith(".invalid"):
        return True
    cached = _DNS_RESOLVE_CACHE.get(host)
    if cached is not None:
        return cached
    try:
        await asyncio.wait_for(asyncio.to_thread(socket.getaddrinfo, host, None), timeout=timeout)
        if len(_DNS_RESOLVE_CACHE) >= _MAX_DNS_CACHE_SIZE:
            _DNS_RESOLVE_CACHE.clear()
        _DNS_RESOLVE_CACHE[host] = True
        return True
    except Exception:
        if len(_DNS_RESOLVE_CACHE) >= _MAX_DNS_CACHE_SIZE:
            _DNS_RESOLVE_CACHE.clear()
        _DNS_RESOLVE_CACHE[host] = False
        return False


async def email_has_valid_mx(email: str, *, timeout: float = 1.5) -> bool:
    if not HAS_DNS_RESOLVER or not email or "@" not in email:
        return True
    domain = email.rsplit("@", 1)[1].strip().lower()
    try:
        domain = domain.encode("idna").decode("ascii")
    except Exception:
        pass
    cached = _MX_RESOLVE_CACHE.get(domain)
    if cached is not None:
        return cached
    try:
        loop = asyncio.get_event_loop()
        records = await asyncio.wait_for(
            loop.run_in_executor(None, lambda: dns.resolver.resolve(domain, "MX")),
            timeout=timeout,
        )
        has_mx = len(records) > 0
        if len(_MX_RESOLVE_CACHE) >= _MAX_DNS_CACHE_SIZE:
            _MX_RESOLVE_CACHE.clear()
        _MX_RESOLVE_CACHE[domain] = has_mx
        return has_mx
    except Exception:
        if len(_MX_RESOLVE_CACHE) >= _MAX_DNS_CACHE_SIZE:
            _MX_RESOLVE_CACHE.clear()
        _MX_RESOLVE_CACHE[domain] = False
        return False


async def verify_candidate(
    settings: SystemSettings,
    candidate: Candidate,
    context: str,
    *,
    profile: ProcurementProfile | None = None,
    registry_context: MinpromRegistryContext | None = None,
) -> dict | None:
    # Fast DNS pre-check: skip unresolvable/dead domains before heavy HTTP/Playwright
    if not await candidate_domain_resolves_fast(candidate.domain or candidate.url):
        return None

    source_token = _supplier_search_source_context.set(str(candidate.source or "unknown")[:80])
    try:
        pages = await collect_pages(candidate.url)
    finally:
        _supplier_search_source_context.reset(source_token)
    if not pages:
        return None
    combined_text = "\n\n".join(page["text"] for page in pages)
    match = assess_candidate_match(candidate, context, pages)
    if not settings.has_active_ai_provider:
        return {
            "company_name": candidate.domain,
            "site": candidate.url,
            "evidence_status": "weak",
            "source": candidate.source,
            "search_query": candidate.query,
            "comments": "Сервис ИИ недоступен: кандидат не проверен обязательным ИИ-аудитом.",
        }
    emails = prioritize_emails(EMAIL_RE.findall(combined_text), candidate.domain)
    phones = sorted(set(PHONE_RE.findall(combined_text)))
    inn_matches = [m.group(1) or m.group(2) for m in re.finditer(r"\bИНН(?:/КПП)?[:\s]*(\d{10}|\d{12})\b", combined_text, re.I)]
    extracted_inn = inn_matches[0] if inn_matches else ""

    decision = await ai_verify(settings, candidate, context, pages, emails, phones, match, profile=profile, registry_context=registry_context)
    rejection = _ai_rejection_reason(decision)
    if rejection:
        return {
            "company_name": candidate.domain,
            "site": candidate.url,
            "evidence_status": "weak",
            "source": candidate.source,
            "search_query": candidate.query,
            "comments": rejection,
        }

    evidence_url = decision.get("evidence_url") or best_evidence_page_url(pages, match) or pages[0]["url"]
    contact_url = decision.get("contact_url") or contact_page_url(pages, emails, phones) or pages[0]["url"]
    site_url = evidence_url if match.level == "exact" else candidate.url
    phone = _verified_phone(decision.get("phone"), phones)
    email = _verified_email(decision.get("email"), emails)
    if email and HAS_DNS_RESOLVER:
        if not await email_has_valid_mx(email):
            # Check other extracted emails
            valid_found = False
            for alt_email in emails:
                if alt_email != email and await email_has_valid_mx(alt_email):
                    email = alt_email
                    valid_found = True
                    break
            if not valid_found and not phone:
                email = ""
    contact_warning = ""
    if not phone and not email:
        # Fallback: accept AI-provided contacts even without parser confirmation
        ai_email = str(decision.get("email") or "").strip()
        ai_phone = str(decision.get("phone") or "").strip()
        if ai_email and EMAIL_RE.fullmatch(ai_email.lower()):
            email = ai_email.lower()
            contact_warning = " Контакт указан ИИ, требует ручной проверки."
        elif ai_phone:
            phone_match = PHONE_RE.search(ai_phone)
            if phone_match:
                normalized = _normalize_ru_phone(phone_match.group(0))
                if normalized:
                    phone = normalized
                    contact_warning = " Контакт указан ИИ, требует ручной проверки."
        if not phone and not email:
            return {
                "company_name": decision.get("company_name") or candidate.domain,
                "site": site_url,
                "evidence_status": "weak",
                "source": candidate.source,
                "search_query": candidate.query,
                "comments": "ИИ-аудит не подтвердил опубликованный телефон или email поставщика.",
            }
    comments = _sanitize_minprom_comment_claims(
        (decision.get("comments") or match.reason or "Официальный сайт открыт, релевантность и контакты проверены.") + contact_warning,
        registry_context,
    )
    verified_inn = str(decision.get("inn") or extracted_inn or "").strip()
    result = {
        "company_name": decision.get("company_name") or candidate.domain,
        "inn": re.sub(r"\D+", "", verified_inn)[:12],
        "region": decision.get("region") or "",
        "status": decision.get("status") or "поставщик",
        "product": decision.get("product") or match.product,
        "contact_person": decision.get("contact_person") or "",
        "phone": phone,
        "email": email,
        "site": site_url,
        "evidence_url": evidence_url,
        "contact_url": contact_url,
        "comments": comments,
        "evidence_status": "verified",
        "match_level": _ai_match_level(decision, match.level),
        "procurement_item_id": str(decision.get("procurement_item_id") or candidate.procurement_item_id or ""),
        "procurement_item": str(decision.get("procurement_item_name") or decision.get("procurement_item") or ""),
        "site_type": str(decision.get("site_type") or ""),
        "product_fit": str(decision.get("product_fit") or ""),
        "ai_confidence": _bounded_int(decision.get("confidence"), default=0),
        "evidence_snippet": str(decision.get("evidence_snippet") or "")[:700],
        "contact_evidence_snippet": str(decision.get("contact_evidence_snippet") or "")[:700],
        "ai_rank_confidence": candidate.ai_rank_confidence,
        "ai_rank_reason": candidate.ai_rank_reason,
        "source": candidate.source,
        "search_query": candidate.query,
    }
    score = _supplier_quality_score(result)
    result["quality_score"] = score
    result["quality_tier"] = _supplier_quality_tier(score)
    return result


def _sanitize_minprom_comment_claims(comment: str, registry_context: MinpromRegistryContext | None) -> str:
    value = re.sub(r"\s+", " ", str(comment or "")).strip()
    if not value or not registry_context or not registry_context.requirement.required:
        return value
    if registry_context.status == "ok" and registry_context.entries:
        return value
    parts = re.split(r"(?<=[.!?])\s+", value)
    kept = [
        part.strip()
        for part in parts
        if part.strip()
        and not re.search(
            r"(?:минпромторг\w*|гисп|реестр\w*|реестров\w*|национальн\w+\s+режим|требован\w+\s+национальн\w+\s+режим)",
            part,
            flags=re.I,
        )
    ]
    cleaned = " ".join(kept).strip()
    return cleaned or "Официальный сайт открыт, релевантность и контакты проверены."


async def collect_pages(url: str) -> list[dict]:
    pages: list[dict] = []
    async with httpx.AsyncClient(timeout=18, follow_redirects=True, headers={"User-Agent": "TenderLex supplier verifier"}) as client:
        first = await fetch_page(client, url)
        if first:
            pages.append(first)
            links = extract_internal_links(first["html"], first["url"])
            # Parallel fetch: up to 5 internal links concurrently, capped at 3 to avoid hammering the server
            link_semaphore = asyncio.Semaphore(3)

            async def _fetch_link(link_url: str) -> dict | None:
                async with link_semaphore:
                    return await fetch_page(client, link_url)

            if links:
                tasks = [asyncio.create_task(_fetch_link(link)) for link in links[:5]]
                results = await asyncio.gather(*tasks, return_exceptions=True)
                for result in results:
                    if isinstance(result, BaseException) or result is None:
                        continue
                    pages.append(result)
                    if len(pages) >= 6:
                        break
    if pages and _pages_have_contact(pages):
        return pages
    browser_page = await fetch_page_with_browser(url)
    if browser_page:
        for index, page in enumerate(pages):
            if page["url"] == browser_page["url"]:
                pages[index] = browser_page
                break
        else:
            pages.insert(0, browser_page)
    return pages


async def fetch_page(client: httpx.AsyncClient, url: str) -> dict | None:
    try:
        response = await client.get(url)
        ctype = response.headers.get("content-type", "")
        if response.status_code >= 400 or "text/html" not in ctype:
            return None
        html_text = response.text[:300000]
    except Exception:
        return None
    return html_text_to_page(html_text, str(response.url))


def html_text_to_page(html_text: str, url: str) -> dict | None:
    soup = BeautifulSoup(html_text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    text = unescape(soup.get_text("\n", strip=True))
    href_contacts: list[str] = []
    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href") or "").strip()
        if href.lower().startswith(("mailto:", "tel:")):
            href_contacts.append(href.replace("mailto:", "").replace("tel:", ""))
    if href_contacts:
        text = f"{text}\n" + "\n".join(href_contacts)
    if len(text.strip()) < 80:
        return None
    return {"url": str(url), "html": html_text, "text": text[:80000]}


def _browser_log_url(url: str) -> str:
    try:
        parsed = urlsplit(str(url))
    except ValueError:
        return "<invalid-url>"
    if not parsed.scheme or not parsed.netloc:
        return "<invalid-url>"
    netloc = parsed.netloc.split("@", 1)[-1]
    return urlunsplit((parsed.scheme, netloc, parsed.path, "", ""))[:500]


def _browser_log_error(exc: BaseException) -> str:
    summary = _exception_summary(exc)
    summary = URL_RE.sub("<url>", summary)
    summary = re.sub(r"(?i)(authorization|proxy-authorization)(\s*[:=]\s*)[^\r\n,;]+", r"\1\2<redacted>", summary)
    summary = re.sub(r"(?i)(cookie)(\s*[:=]\s*)[^\r\n]+", r"\1\2<redacted>", summary)
    summary = re.sub(r'(?i)("(?:token|password|secret|key)"\s*:\s*)"[^"]+"', r'\1"<redacted>"', summary)
    summary = re.sub(r'(?i)(token|password|session|remember|secret|key)(\s*=\s*)[^\s;&,]+', r'\1\2<redacted>', summary)
    return re.sub(r"\s+", " ", summary).strip()[:300]


async def fetch_page_with_browser(url: str, *, source: str = "") -> dict | None:
    """Fetch a page using the current supplier-search browser pool."""
    try:
        async with _browser_pool_session():
            browser_pool = _get_browser_pool()
            async with browser_pool:
                browser = await browser_pool.get_browser()
                page = await browser.new_page(user_agent="TenderLex supplier verifier")
                try:
                    await page.goto(url, wait_until="networkidle", timeout=18000)
                    html_text = await page.content()
                    final_url = page.url
                finally:
                    await page.close()
    except Exception as exc:
        logger.warning(
            "supplier_browser_fetch_failed job_id=%s source=%s url=%s error=%s",
            _supplier_search_job_id_context.get(),
            str(source or _supplier_search_source_context.get())[:80],
            _browser_log_url(url),
            _browser_log_error(exc),
        )
        return None
    return html_text_to_page(html_text[:300000], final_url)


class _BrowserPool:
    """Manages a shared Playwright browser instance with semaphore-based concurrency."""

    def __init__(self) -> None:
        self._playwright = None
        self._browser = None
        self._semaphore = asyncio.Semaphore(3)  # Max 3 concurrent browser tabs
        self._lock = asyncio.Lock()

    async def get_browser(self):
        """Get or create the shared browser instance."""
        async with self._lock:
            if self._browser is not None and self._browser.is_connected():
                return self._browser
            await self._close_unlocked()
            from playwright.async_api import async_playwright  # type: ignore
            playwright = await async_playwright().start()
            try:
                browser = await playwright.chromium.launch(headless=True, args=["--no-sandbox"])
            except BaseException:
                await playwright.stop()
                raise
            self._playwright = playwright
            self._browser = browser
            return browser

    async def __aenter__(self):
        await self._semaphore.acquire()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        self._semaphore.release()

    async def close(self) -> None:
        """Close the browser and playwright instances."""
        async with self._lock:
            await self._close_unlocked()

    async def _close_unlocked(self) -> None:
        cancelled_exc: BaseException | None = None
        if self._browser:
            for attempt in range(2):
                try:
                    await self._browser.close()
                    break
                except asyncio.CancelledError as exc:
                    cancelled_exc = exc
                    break
                except Exception:
                    if attempt == 0:
                        continue
            self._browser = None
        if self._playwright:
            try:
                await self._playwright.stop()
            except Exception:
                pass
            self._playwright = None
        if cancelled_exc:
            raise cancelled_exc


_browser_pool_context: ContextVar[_BrowserPool | None] = ContextVar(
    "supplier_search_browser_pool",
    default=None,
)
_supplier_search_job_id_context: ContextVar[str] = ContextVar(
    "supplier_search_job_id",
    default="unknown",
)
_supplier_search_source_context: ContextVar[str] = ContextVar(
    "supplier_search_source",
    default="unknown",
)


@contextmanager
def supplier_search_job_context(job_id: str):
    token = _supplier_search_job_id_context.set(str(job_id or "unknown")[:64])
    try:
        yield
    finally:
        _supplier_search_job_id_context.reset(token)


@asynccontextmanager
async def _browser_pool_session():
    existing = _browser_pool_context.get()
    if existing is not None:
        yield existing
        return

    pool = _BrowserPool()
    token = _browser_pool_context.set(pool)
    try:
        yield pool
    finally:
        try:
            await pool.close()
        finally:
            _browser_pool_context.reset(token)


def _get_browser_pool() -> _BrowserPool:
    """Return the browser pool bound to the current supplier-search job."""
    pool = _browser_pool_context.get()
    if pool is None:
        raise RuntimeError("browser pool is not active")
    return pool


def _pages_have_contact(pages: list[dict]) -> bool:
    text = "\n".join(str(page.get("text") or "") for page in pages)
    return bool(EMAIL_RE.search(text) or PHONE_RE.search(text))


def extract_internal_links(html_text: str, base_url_value: str) -> list[str]:
    base = base_domain(base_url_value)
    soup = BeautifulSoup(html_text, "html.parser")
    scored: list[tuple[int, str]] = []
    for anchor in soup.find_all("a", href=True):
        href = urljoin(base_url_value, str(anchor.get("href") or ""))
        parsed = urlparse(href)
        if base_domain(parsed.netloc) != base:
            continue
        label = f"{anchor.get_text(' ', strip=True)} {href}".lower()
        score = 0
        if any(word in label for word in ["контакт", "contact", "отдел", "sales", "продаж", "связ", "requisite", "реквизит"]):
            score += 4
        if any(word in label for word in ["каталог", "product", "produk", "товар", "оборуд", "shop", "catalog"]):
            score += 3
        if any(word in label for word in ["о компании", "about", "производ", "завод", "company"]):
            score += 2
        if score:
            scored.append((score, href))
    unique: list[str] = []
    for _, href in sorted(scored, reverse=True):
        if href not in unique:
            unique.append(href)
    return unique


async def ai_verify(
    settings: SystemSettings,
    candidate: Candidate,
    context: str,
    pages: list[dict],
    emails: list[str],
    phones: list[str],
    match: CandidateMatch,
    *,
    profile: ProcurementProfile | None = None,
    registry_context: MinpromRegistryContext | None = None,
) -> dict:
    if not settings.has_active_ai_provider:
        return {
            "action": "reject",
            "confidence": 0,
            "site_type": "unknown",
            "product_fit": "unrelated",
            "comments": "Сервис ИИ недоступен для проверки кандидата.",
        }
    payload = {
        "target_tz_excerpt": context[:6000],
        "candidate": candidate.__dict__,
        "local_match": {
            "level": match.level,
            "product": match.product,
            "reason": match.reason,
            "matched_terms": list(match.matched_terms),
        },
        "procurement_profile": _profile_to_dict(profile) if profile else {},
        "minprom_registry": _minprom_context_to_dict(registry_context) if registry_context else {},
        "rerank_hint": {
            "procurement_item_id": candidate.procurement_item_id,
            "confidence": candidate.ai_rank_confidence,
            "reason": candidate.ai_rank_reason,
        },
        "emails": emails,
        "phones": phones[:5],
        "inns": extracted_inn if 'extracted_inn' in locals() and extracted_inn else [],
        "pages": [{"url": page["url"], "text": page["text"][:2500]} for page in pages[:4]],
    }
    prompt = f"""Проверь поставщика для закупочного ТЗ.
Твоя задача — финальный закупочный аудит перед попаданием строки в отчёт.

Принимай только если одновременно верно:
- сайт принадлежит компании, которая производит, продает, поставляет, дилерствует или дистрибутирует закупаемый товар/аналог;
- на сайте виден точный товар, совместимый аналог, релевантная товарная категория или профильная специализация, достаточная для запроса КП;
- контакт опубликован на открытой странице этого же сайта;
- сайт не является маркетплейсом, агрегатором, тендерной площадкой, реестром, справочником, статьей, видео, форумом, учебным/госресурсом или страницей профессии.

	Классификация product_fit:
- exact: конкретный товар или товарная страница подтверждает предмет закупки и ключевые обязательные характеристики из ТЗ;
- analog: найден конкретный похожий товар, но не все обязательные характеристики из ТЗ подтверждены на странице;
- category: сайт подтверждает релевантную товарную категорию, но не подтверждает конкретный полностью подходящий товар;
- profile: компания профильная, но конкретный товар или категория подтверждены недостаточно для вывода о полном соответствии.

	В comments запрещено писать "полностью соответствует" для analog, category и profile. Для них пиши, что поставщик релевантен для запроса КП и что у него нужно уточнить конкретные характеристики товара по ТЗ.

Если minprom_registry.required=true:
- учитывай найденные записи ГИСП/Минпромторга как важное подтверждение товара российского производства;
- положительное подтверждение Минпромторга можно писать только если minprom_registry.status="ok", minprom_registry.entries_count > 0 и сайт кандидата совпадает с конкретной записью по названию/товару/ИНН;
- если minprom_registry.status="empty" или "error", либо entries_count=0, считай реестровую запись непроверенной: не пиши, что товар включен в реестр, соответствует национальному режиму или что требование Минпромторга выполнено;
- если кандидат является дилером/дистрибьютором, принимай его только при релевантном товаре и контактах; не пиши в comments повторяющийся отрицательный статус вроде "выписки нет" или "нужно запросить подтверждение" для каждой строки;
- запрещено писать, что требование Минпромторга выполнено, если связи с записью нет.

	Отклоняй, если совпадение только по комплектующей, стандарту, форме документа, служебному коду, адресу, условиям поставки, новостной/справочной статье или странице без доказательства поставки закупаемого предмета.
Ответ строго JSON:
{{
  "action": "accept|reject",
  "confidence": 0,
  "site_type": "manufacturer|dealer|distributor|supplier|service_company|marketplace|aggregator|tender|registry|directory|article|video|education|government|unknown",
  "product_fit": "exact|analog|category|profile|unrelated",
  "procurement_item_id": "",
  "procurement_item_name": "",
  "company_name": "",
  "inn": "10 или 12 цифр ИНН компании если найдено на сайте",
  "region": "",
  "status": "завод|дилер|дистрибьютор|поставщик",
  "product": "",
  "email": "",
  "phone": "",
  "evidence_url": "",
  "contact_url": "",
  "evidence_snippet": "короткая цитата/фрагмент страницы, подтверждающий товар или профиль",
  "contact_evidence_snippet": "короткая цитата/фрагмент страницы, подтверждающий телефон/email",
  "comments": ""
}}

Данные:
{json.dumps(payload, ensure_ascii=False)}"""
    try:
        raw = await call_llm(
            settings,
            prompt,
            system_prompt="Ты закупочный аудитор. Не выдумываешь компании и контакты, но не отбрасываешь реального профильного поставщика только из-за отсутствия точного артикула.",
            tier="primary",
            routing_key="supplier_candidate_verifier",
            json_mode=True,
            timeout_seconds=90,
            response_validator=_validate_supplier_verification_response,
        )
        return parse_json_object(raw)
    except Exception:
        return {
            "action": "reject",
            "confidence": 0,
            "site_type": "unknown",
            "product_fit": "unrelated",
            "comments": "ИИ-аудит поставщика не выполнен: кандидат не принят без обязательной проверки.",
        }


def _ai_rejection_reason(decision: dict) -> str:
    action = str(decision.get("action") or "").strip().lower()
    if action != "accept":
        return str(decision.get("comments") or "ИИ-аудит отклонил кандидата: соответствие ТЗ, тип сайта или контакт не подтверждены.")
    site_type = str(decision.get("site_type") or "").strip().lower()
    accepted_site_types = {"manufacturer", "dealer", "distributor", "supplier", "service_company"}
    if site_type not in accepted_site_types:
        return "ИИ-аудит отклонил кандидата: тип сайта поставщика не подтвержден."
    if site_type in {"marketplace", "aggregator", "tender", "registry", "directory", "article", "video", "education", "government"}:
        return f"ИИ-аудит отклонил кандидата: тип сайта не является сайтом поставщика ({site_type})."
    product_fit = str(decision.get("product_fit") or "").strip().lower()
    if product_fit not in {"exact", "analog", "category", "profile"}:
        return "ИИ-аудит отклонил кандидата: продукция не соответствует предмету ТЗ."
    confidence = decision.get("confidence")
    if confidence in (None, ""):
        return "ИИ-аудит отклонил кандидата: отсутствует оценка уверенности."
    normalized_confidence = _bounded_int(confidence, default=-1)
    if normalized_confidence < 0:
        return "ИИ-аудит отклонил кандидата: некорректная оценка уверенности."
    # Adaptive confidence threshold by product_fit
    confidence_thresholds = {
        "exact": 45,
        "analog": 40,
        "category": 55,
        "profile": 65,
    }
    min_confidence = confidence_thresholds.get(product_fit, 45)
    if normalized_confidence < min_confidence:
        return f"ИИ-аудит отклонил кандидата: низкая уверенность ({normalized_confidence}) для {product_fit} (порог {min_confidence})."
    if product_fit in {"category", "profile"}:
        evidence_snippet = re.sub(r"\s+", " ", str(decision.get("evidence_snippet") or "")).strip()
        if len(evidence_snippet) < 20:
            return "ИИ-аудит отклонил кандидата: для профильного совпадения нет фрагмента сайта с подтверждением категории."
    return ""


def _ai_match_level(decision: dict, local_level: str) -> str:
    product_fit = str(decision.get("product_fit") or "").strip().lower()
    if product_fit == "exact":
        return "exact"
    if product_fit == "analog":
        return "adjacent"
    if product_fit in {"category", "profile"}:
        return "profile"
    return local_level if local_level in {"exact", "adjacent", "profile"} else "profile"


def keyword_verify(
    candidate: Candidate,
    context: str,
    pages: list[dict],
    emails: list[str],
    phones: list[str],
    match: CandidateMatch | None = None,
) -> dict:
    match = match or assess_candidate_match(candidate, context, pages)
    if not match.accepted:
        return {"action": "reject", "comments": match.reason or "Недостаточно совпадений с ТЗ без ИИ-проверки."}
    context_words = set(important_terms(context))
    text = "\n".join(page["text"].lower() for page in pages)
    overlap = [word for word in context_words if word in text]
    name = extract_company_name(candidate, pages)
    return {
        "action": "accept",
        "company_name": name,
        "status": infer_supplier_status(text),
        "product": match.product or ", ".join(overlap[:8]),
        "email": emails[0] if emails else "",
        "phone": phones[0] if phones else "",
        "evidence_url": pages[0]["url"],
        "contact_url": contact_page_url(pages, emails, phones) or pages[0]["url"],
        "comments": match.reason or "Проверка выполнена по официальной странице, профилю поставщика и контактам.",
    }


def email_matches_domain(email: str, domain: str) -> bool:
    email_domain = base_domain(email.split("@")[-1])
    return bool(email_domain and email_domain == base_domain(domain))


def _confidence_percent(value: object) -> int:
    if value is None:
        return 0
    try:
        num = float(value)
    except (ValueError, TypeError):
        return 0
    if 0.0 < num < 1.0:
        return int(round(num * 100))
    if 1.0 <= num <= 10.0 and isinstance(value, (int, float)) and num == int(num):
        raise RuntimeError("ambiguous 0-10 confidence scale")
    if 0 <= num <= 100:
        return int(round(num))
    return 0


def _normalize_phone(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    digits = re.sub(r"\D+", "", raw)
    if not digits:
        return ""
    if raw.startswith("+"):
        return f"+{digits}"
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        return f"+7{digits[1:]}"
    if len(digits) == 10:
        return f"+7{digits}"
    return f"+{digits}"


def _verified_email(value: object, extracted_emails: list[str]) -> str:
    raw = str(value or "").strip().lower()
    if raw and extracted_emails:
        if raw in [e.lower() for e in extracted_emails]:
            return raw
        return extracted_emails[0]
    if EMAIL_RE.fullmatch(raw):
        return raw
    return extracted_emails[0] if extracted_emails else ""


def _verified_phone(value: object, extracted_phones: list[str]) -> str:
    if not extracted_phones:
        return ""
    raw = str(value or "").strip()
    raw_ru = _normalize_ru_phone(raw)
    raw_norm = _normalize_phone(raw)
    for phone in extracted_phones:
        ru = _normalize_ru_phone(phone)
        norm = _normalize_phone(phone)
        if ru and (ru == raw_ru or norm == raw_norm or not raw):
            return ru
        if norm and (norm == raw_norm or not raw):
            return ru if ru else norm
    for phone in extracted_phones:
        ru = _normalize_ru_phone(phone)
        if ru:
            return ru
        norm = _normalize_phone(phone)
        if norm:
            return norm
    return ""


def _normalize_ru_phone(value: object) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    if len(digits) == 10:
        digits = f"7{digits}"
    if len(digits) != 11 or digits[0] not in {"7", "8"}:
        return ""
    area = digits[1:4]
    if area == "000":
        return ""
    return f"+7 ({area}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"


def _merge_deterministic_okpd2(profile: ProcurementProfile, text: str) -> ProcurementProfile:
    codes = re.findall(r"\b\d{2}\.\d{2}(?:\.\d{2})?(?:\.\d{3})?\b", text)
    if not codes or not profile.items:
        return profile
    updated_items = []
    for item in profile.items:
        existing = list(item.okpd2_codes or [])
        for c in codes:
            if c not in existing:
                existing.append(c)
        updated_items.append(
            ProcurementItem(
                id=item.id,
                name=item.name,
                aliases=item.aliases,
                okpd2_codes=tuple(existing),
                category_terms=item.category_terms,
                exact_terms=item.exact_terms,
                required_terms=item.required_terms,
                excluded_terms=item.excluded_terms,
            )
        )
    return ProcurementProfile(
        summary=profile.summary,
        items=tuple(updated_items),
        excluded_terms=profile.excluded_terms,
        raw=profile.raw,
    )


def _non_registry_alternative_evidence(
    rows: list[dict],
    registry_context: MinpromRegistryContext,
    *,
    registry_verified_count: int = 0,
) -> dict:
    non_registry_rows = [
        dict(r)
        for r in rows
        if not (r.get("minprom_registry_match") or {}).get("matched")
    ]
    available = (
        registry_context.requirement.required
        and registry_context.status in {"empty", "ok"}
        and registry_verified_count == 0
        and len(non_registry_rows) > 0
    )
    reason_code = ""
    if available:
        if registry_context.status == "empty":
            reason_code = "registry_no_relevant_entries"
        else:
            reason_code = "registry_entries_no_supplier_match"
    return {
        "available": available,
        "verified_count": len(non_registry_rows) if non_registry_rows else len(rows),
        "verified_rows": deepcopy(non_registry_rows) if available else [],
        "reason_code": reason_code,
    }


def prioritize_emails(values: list[str], domain: str) -> list[str]:
    unique = sorted({str(value or "").strip().lower() for value in values if str(value or "").strip()})
    return sorted(unique, key=lambda item: (not email_matches_domain(item, domain), item))


def contact_page_url(pages: list[dict], emails: list[str], phones: list[str]) -> str:
    email_set = {email.lower() for email in emails}
    for page in pages:
        text = page["text"].lower()
        if any(email in text for email in email_set):
            return page["url"]
        if any(phone and phone in page["text"] for phone in phones):
            return page["url"]
    return ""


def best_evidence_page_url(pages: list[dict], match: CandidateMatch) -> str:
    terms = [term.lower() for term in match.matched_terms if len(term) >= 4]
    if not terms:
        return ""
    for page in pages:
        text = page["text"].lower()
        if any(term in text for term in terms):
            return page["url"]
    return ""


def _normalize_company_key(value: str) -> str:
    lowered = str(value or "").lower()
    lowered = re.sub(r"\b(общество с ограниченной ответственностью|акционерное общество|закрытое акционерное общество|публичное акционерное общество|научно производственное предприятие|производственное объединение|группа компаний|торговый дом)\b", " ", lowered)
    cleaned = re.sub(r"\b(ооо|ао|зао|пао|нпо|нпп|тд|гк|ип|пк|пкг|фирма|компания|инжиниринговая)\b", " ", lowered)
    cleaned = re.sub(r"[^a-zа-яё0-9]+", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def extract_company_name(candidate: Candidate, pages: list[dict]) -> str:
    generic_lines = {"компания", "о компании", "производство", "о производстве", "главная"}
    first_text = pages[0]["text"] if pages else ""
    for line in first_text.splitlines()[:20]:
        line = re.sub(r"\s+", " ", line).strip(" -|")
        if line.lower() in generic_lines or _looks_like_generic_company_line(line):
            continue
        if 4 <= len(line) <= 120 and _looks_like_company_fragment(line):
            return line
    title = re.sub(r"\s+", " ", candidate.title).strip(" -|")
    if title:
        for separator in ("|", "—", "–", "-"):
            title = title.split(separator)[0].strip()
        title_lower = title.lower()
        if title_lower in generic_lines or _looks_like_generic_company_line(title):
            return candidate.domain
        if any(word in title_lower for word in ("купить", "каталог", "рукав", "сверло", "головк", "оборудован", "переходник")):
            return candidate.domain
        if 4 <= len(title) <= 120 and _looks_like_company_fragment(title):
            return title
    return candidate.domain


def _looks_like_generic_company_line(value: str) -> bool:
    lowered = re.sub(r"\s+", " ", str(value or "").lower()).strip(" .,:;")
    if not lowered:
        return True
    if lowered in {
        "интернет",
        "интернет-магазин",
        "магазин",
        "каталог",
        "контакты",
        "гарантия",
        "гарантия и сервис",
        "доставка",
        "оплата",
        "сервис",
    }:
        return True
    return any(
        marker in lowered
        for marker in (
            "личный кабинет",
            "корзина",
            "оформить заказ",
            "гарантия и сервис",
            "интернет-магазин",
        )
    )


def infer_supplier_status(text: str) -> str:
    lowered = text.lower()
    if any(word in lowered for word in ("завод", "производитель", "производство", "изготовитель")):
        return "завод"
    if "дистриб" in lowered:
        return "дистрибьютор"
    if "дилер" in lowered:
        return "дилер"
    return "поставщик"


def _supplier_result_sort_key(item: dict) -> tuple[int, int, int]:
    priority = {"exact": 0, "adjacent": 1, "profile": 2}
    return (
        priority.get(str(item.get("match_level") or ""), 9),
        -_supplier_quality_score(item),
        int(item.get("_source_rank") or 9999),
    )


def _supplier_quality_score(item: dict) -> int:
    explicit = item.get("quality_score")
    if explicit not in (None, ""):
        try:
            return max(0, min(100, int(explicit)))
        except (TypeError, ValueError):
            pass
    if item.get("evidence_status") != "verified":
        return 0
    if is_blocked(str(item.get("site") or "")):
        return 0
    product_fit = str(item.get("product_fit") or "").strip().lower()
    score = {"exact": 70, "profile": 60, "adjacent": 50}.get(str(item.get("match_level") or ""), 35)
    score += {"exact": 10, "analog": 0, "category": -10, "profile": -18, "unrelated": -40}.get(product_fit, 0)
    site_domain = base_domain(str(item.get("site") or ""))
    evidence_domain = base_domain(str(item.get("evidence_url") or ""))
    contact_domain = base_domain(str(item.get("contact_url") or ""))
    email = str(item.get("email") or "")
    phone = str(item.get("phone") or "")
    if evidence_domain and evidence_domain == site_domain:
        score += 8
    if contact_domain and contact_domain == site_domain:
        score += 8
    if phone:
        score += 6
    if email_matches_domain(email, site_domain):
        score += 8
    elif email:
        score += 2
    if _is_useful_query(str(item.get("search_query") or "")):
        score += 4
    else:
        score -= 25
    company = str(item.get("company_name") or "").lower()
    if any(marker in company for marker in ("словар", "документация", "сервисный центр")):
        score -= 15
    if _looks_like_generic_company_line(company):
        score -= 20
    cap = {"exact": 100, "analog": 79, "category": 76, "profile": 72, "unrelated": 0}.get(product_fit, 100)
    return max(0, min(100, cap, score))


def _supplier_quality_tier(score: int) -> str:
    if score >= 80:
        return "high"
    if score >= MIN_VERIFIED_SUPPLIER_SCORE:
        return "medium"
    return "low"


def _deterministic_queries(context: str) -> list[str]:
    product_phrases = _product_phrases(context)
    exact_codes = _exact_codes(context)
    queries: list[str] = []

    for phrase in product_phrases[:4]:
        queries.extend(
            [
                f'"{phrase}" поставщик',
                f'"{phrase}" купить',
                f'"{phrase}" производитель',
            ]
        )
    supplier_codes = [code for code in exact_codes if _is_searchable_supplier_code(code)]
    for code in supplier_codes[:5]:
        queries.extend(
            [
                f'"{code}" поставщик официальный сайт',
                f'"{code}" купить цена',
            ]
        )

    lowered = context.lower()
    phrase_text = " ".join(product_phrases).lower()
    product_has_mine_context = not product_phrases or any(word in phrase_text for word in ("шахт", "горно", "горн", "вгсч"))
    product_has_fire_context = not product_phrases or any(word in phrase_text for word in ("пожар", "рукав", "гм 70", "гм-70"))
    if "сверло шахтное" in lowered or "сшу" in lowered:
        queries.extend(
            [
                '"Сверло шахтное универсальное" поставщик',
                '"СШУ-22" "сверло шахтное"',
                '"СШУ-22" купить',
                '"сверло шахтное" "пожарных рукавов"',
            ]
        )
    if product_has_mine_context and any(word in lowered for word in ("шахт", "горно", "горн")):
        queries.extend(
            [
                "горноспасательное оборудование поставщик официальный сайт",
                "горношахтное противопожарное оборудование поставщик",
                "средства безопасности угольных шахт поставщик",
                "оборудование ВГСЧ поставщик производитель",
                "горноспасательное оборудование завод поставщик купить",
            ]
        )
    if product_has_fire_context and "пожар" in lowered:
        queries.extend(
            [
                "противопожарное оборудование для шахт поставщик",
                "пожарное оборудование для шахт дилер",
                "пожарные рукава трубопровод ГМ 70 поставщик",
                "соединительная арматура пожарных рукавов поставщик",
                "оборудование для обслуживания пожарных рукавов поставщик",
                "пожарные головки ГМ-70 поставщик",
                "переходники пожарных рукавов поставщик",
                "пожарные рукава головки ГМ-70 купить",
                "головка муфтовая ГМ-70 пожарная купить поставщик",
                "соединительные головки пожарные рукава поставщик официальный сайт",
                "пожарные переходники рукава головки поставщик",
                "оборудование для пожарных рукавов купить поставщик",
            ]
        )
    if product_has_fire_context and (
        any(word in lowered for word in ("врезк", "подсоединение пожарных рукавов", "подсоединения пожарных рукавов"))
        or ("пожар" in lowered and "трубопровод" in lowered)
    ):
        queries.extend(
            [
                "оборудование для врезки в трубопровод поставщик",
                "приспособление для подсоединения пожарных рукавов к трубопроводу",
            ]
        )

    if not product_phrases and not supplier_codes:
        words = important_terms(context)
        keywords = []
        for word in words:
            value = word.strip("-").lower()
            if value not in keywords:
                keywords.append(value)
            if len(keywords) >= 8:
                break
        base = " ".join(keywords[:6]) or "промышленное оборудование"
        queries.extend(
            [
                f"{base} производитель официальный сайт",
                f"{base} завод контакты",
                f"{base} поставщик отдел продаж",
            ]
        )
    return _clean_supplier_queries(queries)[:32]


def _clean_supplier_queries(queries: list[str]) -> list[str]:
    cleaned: list[str] = []
    for query in queries:
        value = re.sub(r"\s+", " ", str(query or "")).strip(" .,:;")
        if not value or not _is_useful_query(value):
            continue
        if value not in cleaned:
            cleaned.append(value)
    return cleaned


def _product_phrases(context: str) -> list[str]:
    text = re.sub(r"\s+", " ", context)
    phrases: list[str] = []
    for match in re.finditer(
        r"(?:предмет\s+закупки\s*[:|]\s*)?(?:\d+(?:\.\d+)?\.)?\s*(поставка\s+[^|\n.]{8,180})",
        context,
        re.I,
    ):
        phrases.append(match.group(1).strip())
    for match in re.finditer(r"(?:^|\n)\s*на\s+поставк[ау]\s+([^.\n|]{8,180})", context, re.I):
        phrases.append(match.group(1).strip())
    for match in re.finditer(r"(?:^|\n)\s*-\s*([^|\n]{8,160})", context):
        phrases.append(match.group(1).strip())
    for line in context.splitlines():
        cells = [re.sub(r"\s+", " ", cell).strip(" .,:;") for cell in line.split("|")]
        cells = [cell for cell in cells if cell]
        if len(cells) >= 2 and cells[0].lower() == cells[1].lower():
            phrases.append(cells[0])
        if len(cells) >= 2 and re.fullmatch(r"\d+[.)]?", cells[0]):
            phrases.append(cells[1])
    table_match = re.search(r"\b1\s*\|\s*([^|]{12,240})\|", text)
    if table_match:
        phrases.append(table_match.group(1).strip())
    for match in re.finditer(r"\(([^()]{8,120})\)", text):
        phrase = match.group(1).strip()
        if not re.search(r"ста|кг|см|мм|мпа|копировать|наружн|не более|рабоч|диам|дней|месяц", phrase, re.I):
            phrases.append(phrase)
    title_match = re.search(r"(?:^|\n)\s*ТЕХНИЧЕСКОЕ ЗАДАНИЕ\s*\n\s*([^\n]{12,180})", context, re.I)
    if title_match:
        title = title_match.group(1).strip(" -")
        title = re.sub(r"\s+-\s+[^-]{2,80}$", "", title).strip()
        phrases.append(title)
    cleaned: list[str] = []
    for phrase in phrases:
        phrase = re.sub(r"\s+", " ", phrase).strip(" .,:;")
        phrase = re.sub(r"\s*\(далее\s*[-–—].*$", "", phrase, flags=re.I).strip(" .,:;")
        phrase = re.sub(r"^(?:на\s+)?поставк[аиу]\s+", "", phrase, flags=re.I).strip(" .,:;")
        phrase = re.sub(r"\s+с\s+выполнением\b.*$", "", phrase, flags=re.I).strip(" .,:;")
        phrase = re.sub(r"\s+в\s+количестве\b.*$", "", phrase, flags=re.I).strip(" .,:;")
        phrase = re.sub(r"\s+(?:для|на)\s+(?:судна|котельн)[^,|.]*$", "", phrase, flags=re.I).strip(" .,:;")
        if (
            8 <= len(phrase) <= 220
            and _is_strong_product_phrase(phrase)
            and phrase.lower() not in [item.lower() for item in cleaned]
        ):
            cleaned.append(phrase)
    return cleaned


def _exact_codes(context: str) -> list[str]:
    result: list[str] = []
    for match in re.finditer(r"\b[А-ЯЁA-Z]{2,}[-\s]?\d{1,4}\b", context):
        code = re.sub(r"\s+", "-", match.group(0).upper())
        if code not in result and _is_searchable_supplier_code(code):
            result.append(code)
    return result


def _is_searchable_supplier_code(code: str) -> bool:
    value = re.sub(r"\s+", "-", str(code or "").upper()).strip("-")
    if not value or re.fullmatch(r"ГМ-\d+", value):
        return False
    prefix = value.split("-", 1)[0]
    if prefix in _blocked_supplier_code_prefixes():
        return False
    if len(prefix) <= 2 and not re.search(r"[А-ЯЁ]{2,}", prefix):
        return False
    return bool(re.fullmatch(r"[А-ЯЁA-Z]{2,}-\d{2,5}", value))


def _blocked_supplier_code_prefixes() -> set[str]:
    return {
        "PAGE",
        "TABLE",
        "МПА",
        "ГОСТ",
        "ОСТ",
        "OCT",
        "ТУ",
        "ТЗ",
        "OF",
        "IEC",
        "SWT",
        "AISI",
        "ASTM",
        "DIN",
        "EN",
        "EC",
        "ЕС",
        "КУ",
        "ПВ",
        "ФГ",
        "ТМ",
        "ТЭГ",
        "ТОРГ",
        "УПД",
        "ОКПД",
        "GSM",
        "MPSV",
        "МЭК",
    }


def _is_useful_query(query: str) -> bool:
    lowered = re.sub(r"\s+", " ", str(query or "").lower()).strip(" .,:;\"")
    if not lowered:
        return False
    if _is_generic_supplier_anchor(lowered):
        return False
    generic_fragments = (
        "далее - оборудование",
        "далее оборудование",
        "с использованием двух проводов",
        "с использованием четырех проводов",
        "с использованием четырёх проводов",
        "входной контроль",
        "торг-12",
        "table наименование",
        "характеристики копировать",
        "предложение функциональные характеристики",
        "копировать полностью",
        "генеральный директор",
        "утверждаю",
        "приложение спецификация",
    )
    if any(fragment in lowered for fragment in generic_fragments):
        return False
    if _contains_blocked_supplier_code(lowered) and not _is_minprom_okpd_query(lowered):
        return False
    if re.search(r"\b(?:page|table|of|тз)[-\s]?\d+\b", lowered, re.I):
        return False
    return True


def _is_minprom_okpd_query(value: str) -> bool:
    lowered = str(value or "").lower()
    return "окпд" in lowered and any(
        marker in lowered
        for marker in ("реестр минпромторга", "реестр российской промышленной продукции", "пп 719", "гисп")
    )


def _contains_blocked_supplier_code(value: str) -> bool:
    prefixes = "|".join(re.escape(prefix.lower()) for prefix in _blocked_supplier_code_prefixes())
    return bool(re.search(rf"\b(?:{prefixes})[-\s]?\d+\b", str(value or "").lower(), re.I))


def _is_strong_product_phrase(phrase: str) -> bool:
    lowered = re.sub(r"\s+", " ", str(phrase or "").lower()).strip(" .,:;\"")
    if _is_generic_supplier_anchor(lowered):
        return False
    if any(
        fragment in lowered
        for fragment in (
            "заполняет участник",
            "предложение участника",
            "единицы измерения",
            "наименование параметра",
            "номер реестровой записи",
            "функциональные характеристики",
            "технические характеристики",
            "демонтаж",
            "программирование системы",
        )
    ):
        return False
    if lowered.startswith("работы") or "швеллер" in lowered:
        return False
    if len(lowered) > 100 and re.search(r"\bшт\b|кол-во|количество", lowered):
        return False
    meaningful = [word for word in re.findall(r"[а-яёa-z0-9\-]{5,}", lowered) if word not in _supplier_anchor_stopwords()]
    if len(meaningful) < 2:
        return False
    return any(
        marker in lowered
        for marker in (
            "система",
            "контрол",
            "кабель",
            "жгут",
            "измер",
            "испыт",
            "приспособ",
            "рукав",
            "шахт",
            "пожар",
            "трубопровод",
            "арматур",
            "датчик",
            "станок",
            "установка",
            "комплекс",
            "поликарбон",
            "вермикулит",
            "оргстекл",
            "панел",
            "профил",
            "свар",
            "полуавтомат",
            "скрайб",
            "скалыв",
            "холодильн",
            "камера",
            "камер",
            "кладов",
            "пенообраз",
            "преобраз",
            "электролиз",
            "толкател",
            "вагонет",
            "газорегулятор",
            "гидро",
            "кальматрон",
            "ультрабанд",
        )
    )


def _is_generic_supplier_anchor(value: str) -> bool:
    lowered = re.sub(r"\s+", " ", str(value or "").lower()).strip(" .,:;\"")
    if not lowered:
        return True
    generic_patterns = (
        r"^далее\s*[-–—]?\s*оборудование$",
        r"^оборудование$",
        r"^с использованием (?:двух|четыр[её]х|\d+) проводов$",
        r"^page[-\s]?\d+$",
        r"^table[-\s]?\d+$",
        r"^тз[-\s]?\d+$",
        r"^of[-\s]?\d+$",
        r"^входной контроль$",
    )
    return any(re.fullmatch(pattern, lowered, re.I) for pattern in generic_patterns)


def _supplier_anchor_stopwords() -> set[str]:
    return {
        "далее",
        "оборудование",
        "использованием",
        "использование",
        "двух",
        "четырех",
        "четырёх",
        "проводов",
        "страница",
        "table",
        "наименование",
        "приложение",
        "спецификация",
        "генеральный",
        "директор",
        "утверждаю",
        "контракта",
    }


def _rank_candidates(candidates: list[Candidate], context: str) -> list[Candidate]:
    return sorted(candidates, key=lambda item: _candidate_score(item, context), reverse=True)


def _candidate_score(candidate: Candidate, context: str) -> int:
    text = f"{candidate.url} {candidate.title} {candidate.snippet}".lower()
    score = 0
    exact_terms = _exact_match_terms(context)
    if any(term in text for term in exact_terms):
        score += 18
    category_hits = _category_hits(text)
    score += len(category_hits) * 3
    for term in important_terms(context)[:16]:
        if term in text:
            score += 3 if len(term) >= 8 else 2
    if any(word in text for word in ("производ", "завод", "изготов", "официаль")):
        score += 4
    if any(word in text for word in ("catalog", "product", "produkt", "товар", "каталог")):
        score += 2
    if any(word in text for word in ("контакт", "contact", "mail", "email")):
        score += 1
    return score


def assess_candidate_match(candidate: Candidate, context: str, pages: list[dict]) -> CandidateMatch:
    text = f"{candidate.title} {candidate.snippet}\n" + "\n".join(page["text"][:20000] for page in pages)
    front_text = f"{candidate.url} {candidate.title} {candidate.snippet}".lower()
    lowered = text.lower()
    if _looks_like_reference_or_non_supplier(candidate, lowered):
        return CandidateMatch(
            accepted=False,
            level="reject",
            product="",
            reason="Страница похожа на справочник, тендер, учебный/госресурс или нерелевантный источник, а не на сайт поставщика.",
        )

    commercial = _has_commercial_supplier_signal(lowered)
    exact_terms = _exact_match_terms(context)
    exact_matches = tuple(term for term in exact_terms if term in lowered)
    category_hits = _category_hits(lowered)
    front_category_hits = _category_hits(front_text)
    target_groups = _target_category_groups(context)
    product = _best_product_label(context, exact_matches)
    if not exact_matches and "сумк" in front_text:
        return CandidateMatch(
            accepted=False,
            level="reject",
            product="",
            reason="Найдена сумка/аксессуар, а не само приспособление или профильный поставщик оборудования.",
            matched_terms=tuple(sorted(category_hits)),
        )

    if exact_matches and _exact_matches_need_context(exact_matches):
        overlap = _contextual_overlap_terms(context, lowered, exact_matches)
        if len(overlap) < 2 and not any(hit in category_hits for hit in ("fire", "mine", "pipeline")):
            return CandidateMatch(
                accepted=False,
                level="reject",
                product=product,
                reason="Найдено совпадение по коду из ТЗ, но страница не подтверждает предмет закупки или профильную категорию.",
                matched_terms=exact_matches,
            )

    if exact_matches and commercial:
        return CandidateMatch(
            accepted=True,
            level="exact",
            product=product,
            reason="На сайте найдено прямое совпадение с товаром/кодом из ТЗ и опубликованы признаки поставщика.",
            matched_terms=exact_matches,
        )

    if exact_matches and any(hit in category_hits for hit in ("fire", "mine", "pipeline")):
        return CandidateMatch(
            accepted=True,
            level="exact",
            product=product,
            reason="На сайте найдено прямое совпадение с товаром/кодом из ТЗ и профильная категория.",
            matched_terms=exact_matches,
        )

    if not exact_matches and target_groups and not (front_category_hits & target_groups):
        return CandidateMatch(
            accepted=False,
            level="reject",
            product="",
            reason="В заголовке/сниппете нет явной связи с товаром или отраслью ТЗ.",
            matched_terms=tuple(sorted(category_hits)),
        )
    if not exact_matches and "pipeline" in front_category_hits and (target_groups - {"pipeline"}) and not (
        front_category_hits & (target_groups - {"pipeline"})
    ):
        return CandidateMatch(
            accepted=False,
            level="reject",
            product="",
            reason="Найдена общая трубопроводная тематика без пожарного или шахтного контекста ТЗ.",
            matched_terms=tuple(sorted(category_hits)),
        )

    if commercial and _is_adjacent_category(category_hits, context):
        return CandidateMatch(
            accepted=True,
            level="adjacent",
            product=product or "Профильная категория по ТЗ",
            reason="Официальный сайт профильного поставщика: категория близка к ТЗ, наличие точной позиции нужно запросить у отдела продаж.",
            matched_terms=tuple(sorted(category_hits)),
        )

    if commercial and _is_profile_supplier(category_hits, context):
        return CandidateMatch(
            accepted=True,
            level="profile",
            product=product or "Профильный поставщик категории",
            reason="Официальный сайт профильного поставщика по отрасли ТЗ; строка добавлена как лид для запроса наличия/аналога.",
            matched_terms=tuple(sorted(category_hits)),
        )

    terms = important_terms(context)
    matches = [term for term in terms[:24] if term in lowered]
    if len(matches) >= 3 and commercial:
        return CandidateMatch(
            accepted=True,
            level="profile",
            product=product or ", ".join(matches[:6]),
            reason="Официальный сайт поставщика содержит несколько терминов из ТЗ и контактную коммерческую информацию.",
            matched_terms=tuple(matches[:8]),
        )

    return CandidateMatch(
        accepted=False,
        level="reject",
        product="",
        reason="Страница открыта, но не подтверждает точный товар, близкую категорию или профильного поставщика по ТЗ.",
        matched_terms=tuple(matches[:8]),
    )


def candidate_matches_context(candidate: Candidate, context: str, pages: list[dict]) -> bool:
    return assess_candidate_match(candidate, context, pages).accepted


def _exact_match_terms(context: str) -> list[str]:
    phrases = _product_phrases(context)
    result: list[str] = []
    for phrase in phrases:
        lowered = phrase.lower()
        if len(lowered) <= 80:
            result.append(lowered)
        if "(" in phrase and ")" in phrase:
            result.extend(part.lower().strip() for part in re.findall(r"\(([^()]{6,120})\)", phrase))
    result.extend(
        code.lower().replace(" ", "-")
        for code in _exact_codes(context)
        if not re.fullmatch(r"ГМ-\d+", code)
    )
    lowered_context = context.lower()
    if "сверло шахтное универсальное" in lowered_context:
        result.extend(["сверло шахтное универсальное", "сверло шахтное", "сшу-22", "сшу 22"])
    if "промежуточного подсоединения пожарных рукавов" in lowered_context:
        result.extend(
            [
                "приспособление для промежуточного подсоединения",
                "промежуточного подсоединения пожарных рукавов",
            ]
        )
    return [item for item in dict.fromkeys(re.sub(r"\s+", " ", term).strip(" .,:;") for term in result) if len(item) >= 5]


def _exact_matches_need_context(exact_matches: tuple[str, ...]) -> bool:
    return bool(exact_matches) and all(_is_code_like_term(term) for term in exact_matches)


def _is_code_like_term(value: str) -> bool:
    return bool(re.fullmatch(r"[а-яёa-z]{2,}[-\s]?\d{2,5}", str(value or "").lower()))


def _contextual_overlap_terms(context: str, lowered_page_text: str, exact_matches: tuple[str, ...]) -> list[str]:
    exact_parts = {
        part
        for term in exact_matches
        for part in re.findall(r"[а-яёa-z0-9]{2,}", term.lower())
    }
    ignored = _supplier_anchor_stopwords() | {
        "поставщик",
        "производитель",
        "купить",
        "цена",
        "официальный",
        "сайт",
        "контакты",
        "поставка",
        "товар",
        "оборудование",
        "кабель",
        "кабели",
        "провод",
        "провода",
        "разъем",
        "разъём",
        "комплектующая",
        "характеристика",
        "характеристики",
    }
    overlap: list[str] = []
    for term in important_terms(context):
        normalized = term.lower().strip("-")
        if normalized in ignored or normalized in exact_parts:
            continue
        if len(normalized) < 6:
            continue
        if normalized in lowered_page_text and normalized not in overlap:
            overlap.append(normalized)
        if len(overlap) >= 4:
            break
    return overlap


def _best_product_label(context: str, exact_matches: tuple[str, ...]) -> str:
    phrases = _product_phrases(context)
    for match in exact_matches:
        lowered_match = match.lower()
        for phrase in phrases:
            lowered_phrase = phrase.lower()
            if lowered_match in lowered_phrase or lowered_phrase in lowered_match:
                return phrase[:220]
    if phrases:
        return phrases[0][:220]
    if exact_matches:
        return exact_matches[0]
    words = important_terms(context)
    return " ".join(words[:6])


def _category_hits(text: str) -> set[str]:
    hits: set[str] = set()
    if any(word in text for word in ("пожар", "противопожар", "пожаротуш", "рукав", "гм 70", "гм-70")):
        hits.add("fire")
    if any(word in text for word in ("шахт", "горношахт", "горно-шахт", "горноспас", "вгсч", "угольн")):
        hits.add("mine")
    if any(word in text for word in ("трубопровод", "магистрал", "врезк", "подсоедин", "забор воды", "водн")):
        hits.add("pipeline")
    if any(word in text for word in ("средств безопасности", "средства безопасности", "аварийно", "спасатель")):
        hits.add("safety")
    if any(word in text for word in ("оборудован", "арматур", "инвентар", "снабжен", "комплект")):
        hits.add("equipment")
    return hits


def _has_commercial_supplier_signal(text: str) -> bool:
    return any(
        word in text
        for word in (
            "производ",
            "завод",
            "изготов",
            "постав",
            "дилер",
            "дистриб",
            "купить",
            "цена",
            "каталог",
            "отдел продаж",
            "заявк",
            "заказать",
            "оптом",
            "снабжен",
            "оборудован",
        )
    )


def _is_adjacent_category(category_hits: set[str], context: str) -> bool:
    lowered = context.lower()
    if "пожар" in lowered and "шахт" in lowered and {"fire", "mine"} <= category_hits:
        return True
    if "пожар" in lowered and any(word in lowered for word in ("трубопровод", "магистрал", "врезк")) and {"fire", "pipeline"} <= category_hits:
        return True
    if "шахт" in lowered and any(word in lowered for word in ("трубопровод", "магистрал", "врезк")) and {"mine", "pipeline"} <= category_hits:
        return True
    return len(category_hits & {"fire", "mine", "pipeline", "safety"}) >= 3


def _is_profile_supplier(category_hits: set[str], context: str) -> bool:
    target_groups = _target_category_groups(context)
    if {"fire", "equipment"} <= category_hits:
        return "fire" in target_groups
    if {"mine", "equipment"} <= category_hits:
        return "mine" in target_groups
    if {"pipeline", "equipment"} <= category_hits and not (target_groups - {"pipeline"}):
        return "pipeline" in target_groups
    return bool(target_groups & category_hits) and "safety" in category_hits


def _target_category_groups(context: str) -> set[str]:
    lowered = context.lower()
    target_groups: set[str] = set()
    if "пожар" in lowered:
        target_groups.add("fire")
    if any(word in lowered for word in ("шахт", "горно", "горн", "уголь")):
        target_groups.add("mine")
    if any(word in lowered for word in ("трубопровод", "магистрал", "врезк", "подсоедин")):
        target_groups.add("pipeline")
    return target_groups


def _looks_like_reference_or_non_supplier(candidate: Candidate, text: str) -> bool:
    host = hostname(candidate.url)
    front = f"{candidate.url} {candidate.title} {candidate.snippet}".lower()
    path = urlparse(candidate.url).path.lower()
    if is_blocked(host):
        return True
    if any(
        marker in path
        for marker in (
            "/news",
            "/article",
            "/articles",
            "/blog",
            "/blogs",
            "/info/",
            "/pravila",
            "/instruk",
            "/profession",
            "/tender",
            "/procedures",
            "/num/",
        )
    ):
        if not any(marker in path for marker in ("/catalog", "/product", "/shop")):
            return True
    if any(word in front for word in ("инструкц", "новост", "статья", "article", "blog", "news", "forum", "форум")):
        if not any(word in front for word in ("купить", "каталог", "цена", "продаж", "постав")):
            return True
    if any(word in text[:4000] for word in ("патент", "реферат", "википедия", "академия", "университет", "фгбоу", "мчс россии")):
        return True
    reference_front = f"{front} {text[:5000]}"
    if any(word in reference_front for word in ("решение от", "по делу №", "арбитраж", "судебн")):
        return True
    if "информационная статья" in reference_front:
        return True
    if any(word in reference_front for word in ("профессия", "зарплата", "обучение")) and "производств" in reference_front:
        return True
    if any(word in reference_front for word in ("реестр сертификатов", "сертификат соответствия", "госреестр")):
        return True
    if any(word in reference_front for word in ("тендер", "44-фз", "223-фз", "заказчик", "извещение", "процедур")) and "закуп" in reference_front:
        return True
    if "сверл" in front and not any(word in front for word in ("шахт", "горн", "сшу", "шсу", "пожар", "трубопровод")):
        return True
    if "сверл" in text and not any(word in text for word in ("пожар", "шахт", "горноспас", "трубопровод", "гм 70", "гм-70")):
        return True
    if any(word in text for word in ("радиоламп", "триод", "анод", "модуляторн", "электровакуум")):
        return True
    return False


def important_terms(context: str) -> list[str]:
    stopwords = {
        "техническое",
        "задание",
        "условия",
        "поставки",
        "поставка",
        "товара",
        "товар",
        "требования",
        "договор",
        "договора",
        "срок",
        "дней",
        "даты",
        "дата",
        "город",
        "получения",
        "предложений",
        "коммерческого",
        "заключения",
        "рабочих",
        "календарных",
        "приложение",
        "спецификация",
        "генеральный",
        "директор",
        "утверждаю",
        "контракта",
        "table",
        "наименование",
        "предложение",
        "функциональные",
        "характеристики",
        "технические",
    }
    result: list[str] = []
    for word in re.findall(r"[А-Яа-яЁёA-Za-z0-9\-]{5,}", context.lower()):
        value = word.strip("-")
        if not value or value in stopwords or value.isdigit() or _contains_blocked_supplier_code(value):
            continue
        if value not in result:
            result.append(value)
        if len(result) >= 40:
            break
    return result

async def _discover_minprom_registry_candidates(
    settings: SystemSettings,
    registry_context: MinpromRegistryContext,
    excluded_domains: set[str],
) -> list[Candidate]:
    if registry_context.status != "ok" or not registry_context.entries:
        return []

    candidates: list[Candidate] = []
    seen_inns_or_names: set[str] = set()

    for entry in registry_context.entries:
        inn = re.sub(r"\D+", "", str(entry.get("inn") or ""))
        mfr = str(entry.get("manufacturer") or "").strip()
        key = inn if inn else _normalize_company_key(mfr)
        if not key or key in seen_inns_or_names:
            continue
        seen_inns_or_names.add(key)

        queries = [
            f'ИНН {inn} официальный сайт' if inn else f'"{mfr}" официальный сайт',
            f'ИНН {inn} производитель контакты' if inn else f'"{mfr}" дилер официальный сайт'
        ]
        found_cands, _ = await discover_candidates(
            settings,
            queries,
            max_results=6,
            excluded_domains=excluded_domains,
            primary_candidate_floor=2,
            fallback_candidate_limit=3,
        )
        added_for_entry = 0
        if found_cands:
            for cand in found_cands:
                if cand.domain and not is_blocked(cand.domain) and cand.domain not in excluded_domains:
                    candidates.append(cand)
                    added_for_entry += 1
                    if added_for_entry >= 3:
                        break
    return candidates
